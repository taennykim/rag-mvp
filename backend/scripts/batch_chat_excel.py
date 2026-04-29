import argparse
import json
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent

DEFAULT_INPUT_PATH = SCRIPT_DIR / "questions.xlsx"
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "questions_result.xlsx"

DEFAULT_CHAT_API_URL = "http://localhost:8000/chat"
DEFAULT_TIMEOUT_SECONDS = 180
DEFAULT_SLEEP_SECONDS = 0.2


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def find_or_create_columns(ws) -> dict[str, int]:
    """
    Excel 1행에서 컬럼을 찾는다.

    필수 컬럼:
    - no
    - Question

    없으면 자동 생성:
    - Response
    - Context
    """
    header_map: dict[str, int] = {}

    for cell in ws[1]:
        header = normalize_header(cell.value)
        if header:
            header_map[header] = cell.column

    if "no" not in header_map:
        raise ValueError("Excel 1행에 'no' 컬럼이 필요합니다.")

    if "question" not in header_map:
        raise ValueError("Excel 1행에 'Question' 컬럼이 필요합니다.")

    if "response" not in header_map:
        col = ws.max_column + 1
        ws.cell(row=1, column=col).value = "Response"
        header_map["response"] = col

    if "context" not in header_map:
        col = ws.max_column + 1
        ws.cell(row=1, column=col).value = "Context"
        header_map["context"] = col

    return header_map


def build_chat_payload(question: str) -> dict[str, Any]:
    """
    rag-mvp의 ChatRequest는 테스트 기준 query 필드를 사용한다.
    /chat API가 question 필드를 요구한다면 여기만 {"question": question}으로 바꾸면 된다.
    """
    return {
        "query": question,
    }


def call_chat_api(api_url: str, question: str, timeout_seconds: int) -> dict[str, Any]:
    res = requests.post(
        api_url,
        json=build_chat_payload(question),
        timeout=timeout_seconds,
    )
    res.raise_for_status()

    try:
        return res.json()
    except Exception as exc:
        raise ValueError(f"/chat API 응답이 JSON이 아닙니다. status={res.status_code}, body={res.text[:500]}") from exc


def pick_first(data: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in data and data[key] is not None:
            return data[key]
    return None


def extract_answer(api_response: dict[str, Any]) -> str:
    """
    /chat 응답에서 답변 텍스트를 추출한다.

    우선 기대 형태:
    {
      "answer": "..."
    }

    그 외 response/result/message/content 형태도 일부 대응한다.
    """
    answer = pick_first(
        api_response,
        [
            "answer",
            "response",
            "Response",
            "result",
            "message",
            "content",
        ],
    )

    if isinstance(answer, str):
        return answer.strip()

    if isinstance(answer, dict):
        nested = pick_first(
            answer,
            [
                "answer",
                "response",
                "content",
                "message",
                "text",
            ],
        )
        if isinstance(nested, str):
            return nested.strip()

    if answer is None:
        return ""

    return json.dumps(answer, ensure_ascii=False, indent=2)


def extract_retrieved_chunks(api_response: dict[str, Any]) -> list[dict[str, Any]]:
    """
    /chat 응답에서 retrieved_chunks 리스트를 추출한다.

    우선 기대 형태:
    {
      "retrieved_chunks": [
        {
          "document_name": "..."
        }
      ]
    }

    result/data 아래로 감싸진 형태도 대응한다.
    """
    candidates: list[Any] = [
        api_response.get("retrieved_chunks"),
        api_response.get("retrievedChunks"),
    ]

    result = api_response.get("result")
    if isinstance(result, dict):
        candidates.extend(
            [
                result.get("retrieved_chunks"),
                result.get("retrievedChunks"),
            ]
        )

    data = api_response.get("data")
    if isinstance(data, dict):
        candidates.extend(
            [
                data.get("retrieved_chunks"),
                data.get("retrievedChunks"),
            ]
        )

    for candidate in candidates:
        if isinstance(candidate, list):
            return [item for item in candidate if isinstance(item, dict)]

    return []


def extract_document_name(chunk: dict[str, Any]) -> str | None:
    """
    retrieved_chunks[].document_name 추출.

    우선순위:
    1. chunk["document_name"]
    2. chunk["documentName"]
    3. chunk["metadata"]["document_name"]
    4. chunk["metadata"]["documentName"]
    """
    document_name = pick_first(
        chunk,
        [
            "document_name",
            "documentName",
        ],
    )

    if isinstance(document_name, str) and document_name.strip():
        return document_name.strip()

    metadata = chunk.get("metadata")
    if isinstance(metadata, dict):
        metadata_document_name = pick_first(
            metadata,
            [
                "document_name",
                "documentName",
            ],
        )
        if isinstance(metadata_document_name, str) and metadata_document_name.strip():
            return metadata_document_name.strip()

    return None


def extract_context(api_response: dict[str, Any]) -> str:
    """
    Context 컬럼에는 retrieved_chunks[].document_name 리스트를 넣는다.
    중복은 제거하고 검색 결과 순서는 유지한다.
    """
    chunks = extract_retrieved_chunks(api_response)

    names: list[str] = []
    seen: set[str] = set()

    for chunk in chunks:
        name = extract_document_name(chunk)
        if not name:
            continue

        if name not in seen:
            seen.add(name)
            names.append(name)

    return "\n".join(names)


def autosize_columns(ws, max_width: int = 100) -> None:
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_len = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            for line in str(cell.value).splitlines():
                max_len = max(max_len, len(line))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def apply_cell_style(ws, row: int, response_col: int, context_col: int) -> None:
    ws.cell(row=row, column=response_col).alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )
    ws.cell(row=row, column=context_col).alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )


def process_excel(
    input_path: Path,
    output_path: Path,
    api_url: str,
    timeout_seconds: int,
    sleep_seconds: float,
) -> None:
    wb = load_workbook(input_path)
    ws = wb.active

    cols = find_or_create_columns(ws)

    no_col = cols["no"]
    question_col = cols["question"]
    response_col = cols["response"]
    context_col = cols["context"]

    ws.cell(row=1, column=response_col).value = "Response"
    ws.cell(row=1, column=context_col).value = "Context"

    total_rows = ws.max_row - 1
    print(f"입력 파일: {input_path}")
    print(f"출력 파일: {output_path}")
    print(f"API URL: {api_url}")
    print(f"처리 대상 row 수: {total_rows}")

    for row in range(2, ws.max_row + 1):
        no_value = ws.cell(row=row, column=no_col).value
        question_value = ws.cell(row=row, column=question_col).value

        if question_value is None or str(question_value).strip() == "":
            continue

        question = str(question_value).strip()
        print(f"[row={row}, no={no_value}] 처리 시작: {question[:100]}")

        try:
            api_response = call_chat_api(
                api_url=api_url,
                question=question,
                timeout_seconds=timeout_seconds,
            )

            answer = extract_answer(api_response)
            context = extract_context(api_response)

            ws.cell(row=row, column=response_col).value = answer
            ws.cell(row=row, column=context_col).value = context
            apply_cell_style(ws, row, response_col, context_col)

            print(f"[row={row}, no={no_value}] 완료")

        except Exception as e:
            error_message = f"ERROR: {type(e).__name__}: {e}"

            ws.cell(row=row, column=response_col).value = error_message
            ws.cell(row=row, column=context_col).value = ""
            apply_cell_style(ws, row, response_col, context_col)

            print(f"[row={row}, no={no_value}] 실패: {error_message}")

        time.sleep(sleep_seconds)

    autosize_columns(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"결과 파일 저장 완료: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel의 Question을 rag-mvp /chat API로 배치 실행하고 Response, Context를 채웁니다."
    )

    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT_PATH),
        help=f"입력 Excel 파일 경로. 기본값: {DEFAULT_INPUT_PATH}",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help=f"결과 Excel 파일 경로. 기본값: {DEFAULT_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--api-url",
        default=DEFAULT_CHAT_API_URL,
        help=f"/chat API URL. 기본값: {DEFAULT_CHAT_API_URL}",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT_SECONDS,
        help=f"질문 1건당 timeout 초. 기본값: {DEFAULT_TIMEOUT_SECONDS}",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=DEFAULT_SLEEP_SECONDS,
        help=f"row 사이 대기 초. 기본값: {DEFAULT_SLEEP_SECONDS}",
    )

    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"입력 Excel 파일이 없습니다: {input_path}")

    process_excel(
        input_path=input_path,
        output_path=output_path,
        api_url=args.api_url,
        timeout_seconds=args.timeout,
        sleep_seconds=args.sleep,
    )


if __name__ == "__main__":
    main()
