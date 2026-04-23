from datetime import datetime, timezone
import hashlib
from io import BytesIO
import json
from functools import lru_cache
import logging
from logging.handlers import RotatingFileHandler
import os
from pathlib import Path
import re
import shutil
import socket
import subprocess
import time
from urllib import error as urllib_error
from urllib import request as urllib_request
from uuid import uuid4
import xml.etree.ElementTree as ET
from zipfile import ZipFile

import fitz
from docx import Document
from docx.document import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table, _Cell
from docx.text.paragraph import Paragraph
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.exception_handlers import http_exception_handler
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from app.query_routing import (
    get_document_hint_expansions,
    get_question_type_expansions,
    infer_document_hint_from_rules,
    infer_question_type_from_rules,
)


app = FastAPI(title="Insurance Document RAG MVP API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
        "http://10.160.98.178:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_requests(request: Request, call_next):
    started_at = time.perf_counter()
    logger.info("request_started %s", summarize_request(request))
    try:
        response = await call_next(request)
    except Exception:
        duration_ms = int((time.perf_counter() - started_at) * 1000)
        logger.exception("request_failed %s duration_ms=%s", summarize_request(request), duration_ms)
        raise

    duration_ms = int((time.perf_counter() - started_at) * 1000)
    logger.info(
        "request_completed %s status_code=%s duration_ms=%s",
        summarize_request(request),
        response.status_code,
        duration_ms,
    )
    response.headers["X-Backend-Log-Path"] = str(APP_LOG_PATH)
    return response


@app.exception_handler(HTTPException)
async def handle_http_exception(request: Request, exc: HTTPException):
    logger.warning(
        "http_error %s status_code=%s detail=%s",
        summarize_request(request),
        exc.status_code,
        exc.detail,
    )
    response = await http_exception_handler(request, exc)
    response.headers["X-Backend-Log-Path"] = str(APP_LOG_PATH)
    return response


@app.exception_handler(Exception)
async def handle_unexpected_exception(request: Request, exc: Exception):
    logger.exception("unexpected_error %s", summarize_request(request))
    return JSONResponse(
        status_code=500,
        content={
            "detail": "Unexpected server error.",
            "log_path": str(APP_LOG_PATH),
        },
        headers={"X-Backend-Log-Path": str(APP_LOG_PATH)},
    )

BASE_DIR = Path(__file__).resolve().parent.parent
LOG_DIR = BASE_DIR / "logs"
APP_LOG_PATH = LOG_DIR / "app.log"
UPLOAD_DIR = BASE_DIR / "data" / "uploads"
DEFAULT_FILE_DIR = BASE_DIR / "data" / "default-files"
PARSE_METADATA_DIR = BASE_DIR / "data" / "parse-metadata"
CHUNK_METADATA_DIR = BASE_DIR / "data" / "chunk-metadata"
MARKDOWN_OUTPUT_DIR = BASE_DIR / "data" / "markdown"
CHROMA_DIR = BASE_DIR / "data" / "chroma"
ALLOWED_EXTENSIONS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
PDF_SIGNATURE = b"%PDF-"
ZIP_SIGNATURE = b"PK\x03\x04"
CFBF_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
PREVIEW_LENGTH = 300
QUALITY_WARNING_THRESHOLD = 0.8
PDF_GARBLED_SUSPICIOUS_CHAR_RATIO_THRESHOLD = 0.35
PDF_GARBLED_SUSPICIOUS_CHAR_RATIO_DELTA_THRESHOLD = 0.15
DEFAULT_EXTERNAL_SEARCH_API_ENDPOINT = "http://10.160.98.123:8000/api/search"
DEFAULT_EXTERNAL_SEARCH_TIMEOUT_SEC = 15
PDF_GARBLED_CONTROL_CHAR_RATIO_THRESHOLD = 0.005
PDF_GARBLED_LENGTH_RATIO_THRESHOLD = 0.6
WORD_NAMESPACE = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
CHUNK_TARGET_LENGTH = 800
CHUNK_OVERLAP_LENGTH = 120
CHUNK_PREVIEW_LENGTH = 160
EMBEDDING_DIMENSION = 256
RETRIEVAL_CANDIDATE_MULTIPLIER = 5
RETRIEVAL_MAX_CANDIDATES = 50
MIN_STANDALONE_QUERY_LENGTH = 15
MAX_STANDALONE_QUERY_LENGTH = 150
PRIMARY_PARSER_DOCLING = "docling"
PRIMARY_PARSER_DOCLING_MARKDOWN = "docling-markdown"
PRIMARY_PARSER_LEGACY_AUTO = "legacy-auto"
FALLBACK_PARSER_EXTENSION_DEFAULT = "extension-default"
FALLBACK_PARSER_NONE = "none"
FALLBACK_PARSER_PYMUPDF = "pymupdf"
FALLBACK_PARSER_PYTHON_DOCX = "python-docx"
FALLBACK_PARSER_DOC = "doc-parser"
FALLBACK_PARSER_EXCEL = "excel-parser"
EMBEDDING_PROVIDER_HASH = "hash"
EMBEDDING_PROVIDER_AZURE_OPENAI = "azure_openai"
DEFAULT_AZURE_OPENAI_EMBEDDING_DEPLOYMENT = "text-embedding-3-small"
DEFAULT_AZURE_OPENAI_API_VERSION = "2024-02-01"
DEFAULT_QUERY_REWRITE_MODEL = "gpt-4o-mini"
DEFAULT_ANSWER_MODEL = "gpt-4o"
CUSTOM_QUERY_REWRITE_MODEL = "custom"
AZURE_OPENAI_EMBEDDING_BATCH_SIZE = 32
DEFAULT_CHAT_TOP_K = 30
DEFAULT_CHAT_FINAL_K = 10
DEFAULT_CHAT_TEMPERATURE = 0.3
DEFAULT_CHAT_TOP_P = 0.9
DEFAULT_CHAT_MAX_TOKENS = 700
DEFAULT_QUERY_REWRITE_TIMEOUT_SEC = 120
QUERY_REWRITE_RESPONSE_LOG_PREVIEW_CHARS = 800
STREAM_DELTA_MIN_CHARS = 10
STREAM_DELTA_FLUSH_INTERVAL_SEC = 0.025
STREAM_REWRITE_DELTA_MIN_CHARS = 4
CHAT_RETRY_ANSWER_TEXT = "재 시도 중입니다."
RUNTIME_ENV_FILE = BASE_DIR / ".env.runtime"
QUERY_REWRITE_SPEC_PATH = BASE_DIR.parent / "docs" / "query-rewrite-spec.md"
ANSWER_GENERATION_SPEC_PATH = BASE_DIR.parent / "docs" / "answer-generation-spec.md"
SEARCH_EVAL_MIN_TOP_CHUNK_TEXT_LENGTH = 180
SEARCH_EVAL_MULTI_HIT_SAME_DOCUMENT_THRESHOLD = 2
UNKNOWN_DOCUMENT_NAME = "Unknown document"
UNKNOWN_HEADER_PATH = "Unknown section"
EMPTY_CONTENT_PLACEHOLDER = "[no content]"
UNKNOWN_RRF_SCORE: float | None = None
SEARCH_EVAL_CONTEXT_KEYWORDS: tuple[str, ...] = (
    "단,",
    "예외",
    "유의",
    "조건",
    "면책",
    "제외",
    "제출서류",
)
STATISTICS_KEYWORDS: tuple[str, ...] = (
    "통계",
    "수치",
    "평균",
    "기준",
    "연도",
    "인당",
    "비율",
    "건수",
    "금액",
    "진료비",
    "수술비",
    "발생 건수",
    "발생건수",
    "자료",
    "현황",
)
STATISTICS_REQUIRED_PRESERVATION_KEYWORDS: tuple[str, ...] = (
    "평균",
    "통계",
    "기준",
    "연도",
    "인당",
    "비율",
    "건수",
    "금액",
    "진료비",
    "수술비",
    "발생 건수",
    "발생건수",
)
POLICY_OR_CLAIM_KEYWORDS: tuple[str, ...] = (
    "보장",
    "면책",
    "청구",
    "약관",
    "가능 여부",
    "가능여부",
    "보험금",
    "지급",
)
STATISTICS_DOCUMENT_HINTS: tuple[str, ...] = ("statistics_table", "statistics_report")
STATISTICS_ROUTING_CHUNK_TYPES: tuple[str, ...] = ("statistics_table", "statistics_report")
SEARCH_DOCUMENT_TYPES: tuple[str, ...] = (
    "policy",
    "calculation_guide",
    "business_guide",
    "statistics_table",
)
DOCUMENT_HINT_TO_DOCUMENT_TYPE: dict[str, str] = {
    "terms": "policy",
    "pricing_method": "calculation_guide",
    "change_process": "business_guide",
    "statistics_table": "statistics_table",
    "statistics_report": "statistics_table",
}
YEAR_TOKEN_PATTERN = re.compile(r"(?<!\\d)((?:19|20)\\d{2})\\s*년?")
PROCEDURE_TOKEN_PATTERN = re.compile(r"([0-9A-Za-z가-힣]+(?:\\s*[0-9A-Za-z가-힣]+){0,2}\\s*수술)")
PRODUCT_QUERY_CANDIDATE_PATTERN = re.compile(r"([0-9A-Za-z가-힣()_-]{4,80}?)(?:에서|의|은|는|이|가|도|를|을|과|와|로|으로)")
PRODUCT_SUFFIX_CANDIDATE_PATTERN = re.compile(r"([0-9A-Za-z가-힣()_-]{4,80}?(?:보험|공제|플랜|케어))")
QUOTED_PRODUCT_PATTERN = re.compile(r"[\"'“”‘’]([^\"'“”‘’]{4,80})[\"'“”‘’]")
PRODUCT_DOC_STOPWORDS: tuple[str, ...] = (
    "약관",
    "보통약관",
    "특별약관",
    "산출방법서",
    "사업방법서",
    "상품요약서",
    "요약서",
    "안내",
    "청약서",
    "상품설명서",
    "설명서",
    "template",
    "sample",
)


def load_runtime_env_file() -> None:
    if not RUNTIME_ENV_FILE.is_file():
        return

    for raw_line in RUNTIME_ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_runtime_env_file()


@lru_cache(maxsize=1)
def load_query_rewrite_spec() -> str:
    if not QUERY_REWRITE_SPEC_PATH.is_file():
        return ""
    return QUERY_REWRITE_SPEC_PATH.read_text(encoding="utf-8").strip()


@lru_cache(maxsize=1)
def load_query_rewrite_system_prompt() -> str:
    spec_text = load_query_rewrite_spec()
    if not spec_text:
        return ""

    section_match = re.search(
        r"^##\s*11\.\s*LLM System Prompt\s*$" r"(.*?)" r"(?=^##\s*\d+\.|\Z)",
        spec_text,
        flags=re.MULTILINE | re.DOTALL,
    )
    if not section_match:
        return ""

    section_body = section_match.group(1).strip()
    code_block_match = re.search(r"```(?:text)?\s*(.*?)\s*```", section_body, flags=re.DOTALL)
    if not code_block_match:
        return ""

    return code_block_match.group(1).strip()


@lru_cache(maxsize=1)
def load_answer_generation_spec() -> str:
    if not ANSWER_GENERATION_SPEC_PATH.is_file():
        return ""
    return ANSWER_GENERATION_SPEC_PATH.read_text(encoding="utf-8").strip()


def summarize_llm_response_preview(text: str, *, limit: int = QUERY_REWRITE_RESPONSE_LOG_PREVIEW_CHARS) -> str:
    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) <= limit:
        return normalized
    return f"{normalized[:limit]}..."


def configure_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("rag-mvp")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s")

    file_handler = RotatingFileHandler(APP_LOG_PATH, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.propagate = False
    return logger


logger = configure_logging()


def ensure_upload_dir() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def ensure_default_file_dir() -> None:
    DEFAULT_FILE_DIR.mkdir(parents=True, exist_ok=True)


def ensure_parse_metadata_dir() -> None:
    PARSE_METADATA_DIR.mkdir(parents=True, exist_ok=True)


def ensure_chunk_metadata_dir() -> None:
    CHUNK_METADATA_DIR.mkdir(parents=True, exist_ok=True)


def ensure_markdown_output_dir() -> None:
    MARKDOWN_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def ensure_chroma_dir() -> None:
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)


def summarize_request(request: Request) -> str:
    return f"{request.method} {request.url.path}"


def summarize_parser_selection(primary_parser: str, fallback_parser: str) -> str:
    return f"primary={primary_parser}, fallback={fallback_parser}"


def normalize_fallback_parser(primary_parser: str, fallback_parser: str) -> str:
    if primary_parser == PRIMARY_PARSER_DOCLING_MARKDOWN:
        return FALLBACK_PARSER_NONE
    return fallback_parser


def get_parse_history_from_logs() -> dict[str, dict[str, object]]:
    history: dict[str, dict[str, object]] = {}
    open_attempts: list[dict[str, str]] = []
    log_path = APP_LOG_PATH
    if not log_path.is_file():
        return history

    started_pattern = re.compile(r"parse_started stored_name=(\S+) .* primary=(\S+), fallback=(\S+)")
    completed_pattern = re.compile(
        r"parse_completed stored_name=(\S+) parser_used=(\S+) fallback_used=(True|False)"
    )

    for line in log_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        started_match = started_pattern.search(line)
        if started_match:
            stored_name, primary_parser, fallback_parser = started_match.groups()
            open_attempts.append(
                {
                    "stored_name": stored_name,
                    "primary_parser": primary_parser,
                    "fallback_parser": fallback_parser,
                }
            )
            continue

        completed_match = completed_pattern.search(line)
        if completed_match:
            stored_name, parser_used, fallback_used = completed_match.groups()
            entry = history.setdefault(stored_name, {})
            entry["last_successful_parser_used"] = parser_used
            entry["last_successful_fallback_used"] = fallback_used == "True"
            open_attempts = [attempt for attempt in open_attempts if attempt["stored_name"] != stored_name]
            continue

        if "request_completed POST /parse status_code=400" in line and open_attempts:
            attempt = open_attempts.pop(0)
            entry = history.setdefault(attempt["stored_name"], {})
            entry["last_failed_parser_used"] = f"{attempt['primary_parser']} -> {attempt['fallback_parser']}"
            entry["last_failed_fallback_used"] = attempt["fallback_parser"] != FALLBACK_PARSER_NONE

    return history


def get_extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def get_supported_file_types_message() -> str:
    return "PDF, DOC, DOCX, XLS, and XLSX files are supported."


def get_type_validation_message(expected_extension: str, detected_extension: str | None) -> str:
    expected_label = expected_extension.lstrip(".").upper()
    if detected_extension == ".legacy-office":
        detected_label = "legacy Office binary"
    elif detected_extension:
        detected_label = detected_extension.lstrip(".").upper()
    else:
        detected_label = "unknown file type"

    return f"File extension indicates {expected_label}, but the uploaded content looks like {detected_label}."


def detect_ooxml_extension(content: bytes) -> str | None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            names = set(archive.namelist())
    except Exception:
        return None

    if "word/document.xml" in names:
        return ".docx"
    if "xl/workbook.xml" in names:
        return ".xlsx"
    return None


def detect_actual_extension(filename: str, content: bytes) -> str | None:
    extension = get_extension(filename)
    if extension not in ALLOWED_EXTENSIONS:
        return None

    if content.startswith(PDF_SIGNATURE):
        return ".pdf"

    if content.startswith(ZIP_SIGNATURE):
        return detect_ooxml_extension(content)

    if content.startswith(CFBF_SIGNATURE):
        return ".legacy-office"

    return None


def validate_uploaded_file_type(filename: str, content: bytes) -> None:
    extension = get_extension(filename)
    detected_extension = detect_actual_extension(filename, content)

    if extension == ".pdf" and detected_extension != ".pdf":
        raise HTTPException(status_code=400, detail=get_type_validation_message(extension, detected_extension))

    if extension in {".docx", ".xlsx"} and detected_extension != extension:
        raise HTTPException(status_code=400, detail=get_type_validation_message(extension, detected_extension))

    if extension in {".doc", ".xls"} and detected_extension != ".legacy-office":
        raise HTTPException(status_code=400, detail=get_type_validation_message(extension, detected_extension))


def normalize_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def normalize_chat_text(text: str) -> str:
    return " ".join(text.split())


QUERY_VALIDATION_STOPWORDS = {
    "고객",
    "고객님",
    "상담사",
    "상담원",
    "보험",
    "보험금",
    "관련",
    "경우",
    "내용",
    "문의",
    "질문",
    "부분",
    "그게",
    "그런",
    "이런",
    "저런",
    "친구",
    "나라",
    "사례",
    "기준",
    "무엇",
}


QUESTION_CONVERSATION_ROLE_PREFIXES: tuple[tuple[str, str], ...] = (
    ("customer", "customer:"),
    ("customer", "customer -"),
    ("customer", "customer "),
    ("customer", "고객:"),
    ("customer", "고객 -"),
    ("customer", "고객 "),
    ("customer", "고객님:"),
    ("customer", "고객님 -"),
    ("customer", "고객님 "),
    ("agent", "agent:"),
    ("agent", "agent -"),
    ("agent", "agent "),
    ("agent", "assistant:"),
    ("agent", "assistant -"),
    ("agent", "assistant "),
    ("agent", "상담사:"),
    ("agent", "상담사 -"),
    ("agent", "상담사 "),
    ("agent", "상담원:"),
    ("agent", "상담원 -"),
    ("agent", "상담원 "),
)

CONVERSATIONAL_FILLER_PATTERNS: tuple[str, ...] = (
    "그러니까요",
    "예를 들면",
    "네",
    "음",
    "어",
    "어...",
    "음...",
    "저기",
    "그",
    "그게",
    "약간",
    "혹시",
)

INVALID_REWRITE_PREFIXES: tuple[str, ...] = (":", "：", "...", "…", ",", ".", "?", "!")
VAGUE_REWRITE_MARKERS: tuple[str, ...] = (
    "그러니까요",
    "예를 들면",
    "그런 게",
    "그런건",
    "그런 거",
    "이런 게",
    "이런 거",
    "그게",
    "그거",
    "있는 건지",
    "기간이 있다거나",
    "특정한 이유면",
)

INSURANCE_PRODUCT_CLUE_PATTERNS: tuple[str, ...] = (
    "보험",
    "종신",
    "실손",
    "암보험",
    "건강보험",
    "표준형",
    "계약",
    "특약",
    "가입",
    "보장",
)

DEATH_BENEFIT_CONTEXT_PATTERNS: tuple[str, ...] = (
    "평생 보장",
    "안 나온",
    "안 나오는",
    "못 받",
    "안 된다",
    "안되는",
    "제한",
    "면책",
    "특정한 이유",
)


def normalize_conversation_role(role: str) -> str:
    normalized_role = normalize_chat_text(role).lower()
    if normalized_role in {"customer", "user", "client", "insured"}:
        return "customer"
    if normalized_role in {"agent", "assistant", "advisor", "counselor", "operator"}:
        return "agent"
    return normalized_role or "unknown"


def get_conversation_turn_text(turn: "ConversationTurn") -> str:
    if turn.content and turn.content.strip():
        return turn.content
    if turn.text and turn.text.strip():
        return turn.text
    return ""


def normalize_conversation_context(conversation_context: list["ConversationTurn"]) -> list[dict[str, str]]:
    normalized_turns: list[dict[str, str]] = []
    for turn in conversation_context:
        normalized_content = normalize_chat_text(get_conversation_turn_text(turn))
        if not normalized_content:
            continue
        normalized_turns.append(
            {
                "role": normalize_conversation_role(turn.role),
                "content": normalized_content,
            }
        )
    return normalized_turns


def parse_conversation_context_from_query_text(query_text: str) -> list[dict[str, str]]:
    parsed_turns: list[dict[str, str]] = []
    current_role: str | None = None
    current_content_lines: list[str] = []

    def flush_current_turn() -> None:
        nonlocal current_role, current_content_lines
        if not current_role:
            current_content_lines = []
            return

        normalized_content = normalize_chat_text(" ".join(current_content_lines))
        if normalized_content:
            parsed_turns.append({"role": current_role, "content": normalized_content})
        current_role = None
        current_content_lines = []

    for raw_line in query_text.splitlines():
        stripped_line = raw_line.strip()
        if not stripped_line:
            continue

        lowered_line = stripped_line.lower()
        matched_prefix: tuple[str, str] | None = None
        for role, prefix in QUESTION_CONVERSATION_ROLE_PREFIXES:
            if lowered_line.startswith(prefix):
                matched_prefix = (role, prefix)
                break

        if matched_prefix:
            flush_current_turn()
            role, prefix = matched_prefix
            current_role = role
            current_content_lines = [stripped_line[len(prefix) :].strip().lstrip(":：- ").strip()]
            continue

        if current_role:
            current_content_lines.append(stripped_line)

    flush_current_turn()
    return parsed_turns


def extract_last_customer_message(normalized_turns: list[dict[str, str]]) -> str | None:
    for turn in reversed(normalized_turns):
        if turn.get("role") == "customer" and turn.get("content"):
            return turn["content"]
    return None


def contains_sensitive_identity_info(text: str) -> bool:
    normalized = normalize_chat_text(text)
    if not normalized:
        return False
    sensitive_patterns = (
        r"\b생년월일\b",
        r"\b주민등록\b",
        r"\b성함\b",
        r"\b이름\b",
        r"\b본인 확인\b",
        r"\b정보 확인\b",
        r"\d{2}년\s*\d{1,2}월",
        r"\d{6}-?\d{7}",
        r"[가-힣]\s*○+\s*(?:이고|입니다)",
        r"[가-힣]\s*[xX]+\s*(?:이고|입니다)",
    )
    return any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in sensitive_patterns)


def extract_last_meaningful_customer_message(normalized_turns: list[dict[str, str]]) -> str | None:
    for turn in reversed(normalized_turns):
        if turn.get("role") != "customer":
            continue
        content = turn.get("content") or ""
        if not content:
            continue
        if contains_sensitive_identity_info(content) and not contains_insurance_product_clues(content):
            continue
        return content
    return extract_last_customer_message(normalized_turns)


def strip_conversational_fillers(text: str) -> str:
    normalized = normalize_chat_text(text)
    if not normalized:
        return ""

    cleaned = normalized
    for pattern in CONVERSATIONAL_FILLER_PATTERNS:
        cleaned = re.sub(rf"(^|[\s,.\u2026]){re.escape(pattern)}(?=$|[\s,.\u2026])", " ", cleaned)
    cleaned = re.sub(r"[.,]{2,}", " ", cleaned)
    cleaned = re.sub(r"[~…]+", " ", cleaned)
    cleaned = cleaned.replace("…", " ")
    return normalize_chat_text(cleaned)


def is_vague_customer_message(text: str) -> bool:
    normalized = strip_conversational_fillers(text)
    if not normalized:
        return True

    if len(normalized) < 28:
        return True

    if any(marker in normalized for marker in VAGUE_REWRITE_MARKERS):
        return True

    keyword_count = len(extract_query_keywords(normalized))
    return keyword_count < 2


def contains_insurance_product_clues(text: str) -> bool:
    normalized = normalize_chat_text(text)
    return any(pattern in normalized for pattern in INSURANCE_PRODUCT_CLUE_PATTERNS)


def infer_domain_specific_focus(normalized_turns: list[dict[str, str]]) -> str | None:
    customer_text = " ".join(
        strip_conversational_fillers(turn.get("content", ""))
        for turn in normalized_turns
        if turn.get("role") == "customer" and turn.get("content")
    )
    if not customer_text:
        return None

    if "종신" in customer_text and any(pattern in customer_text for pattern in DEATH_BENEFIT_CONTEXT_PATTERNS):
        return "사망보장"

    return None


def build_customer_rewrite_focus_text(normalized_turns: list[dict[str, str]], last_customer_message: str | None) -> str:
    cleaned_last = strip_conversational_fillers(last_customer_message or "")

    if not normalized_turns:
        return cleaned_last

    customer_turns = [
        strip_conversational_fillers(turn.get("content", ""))
        for turn in normalized_turns
        if turn.get("role") == "customer" and turn.get("content")
    ]
    customer_turns = [turn for turn in customer_turns if turn]
    if not customer_turns:
        return cleaned_last

    if cleaned_last and not is_vague_customer_message(cleaned_last):
        return cleaned_last

    trailing_customer_turns = customer_turns[-3:]
    informative_turn = ""
    for turn in reversed(customer_turns[:-1]):
        if contains_insurance_product_clues(turn):
            informative_turn = turn
            break

    candidate_turns = trailing_customer_turns
    if informative_turn:
        candidate_turns = [informative_turn, *candidate_turns]

    unique_turns: list[str] = []
    seen: set[str] = set()
    for turn in candidate_turns[-4:]:
        if turn not in seen:
            seen.add(turn)
            unique_turns.append(turn)

    joined = normalize_chat_text(" ".join(unique_turns))
    domain_focus = infer_domain_specific_focus(normalized_turns)
    if domain_focus and domain_focus not in joined:
        joined = normalize_chat_text(f"{joined} {domain_focus}")
    return joined


def ensure_question_sentence(text: str) -> str:
    normalized = normalize_chat_text(text).strip()
    if not normalized:
        return ""
    if normalized.endswith("?"):
        return normalized
    normalized = normalized.rstrip(".! ")
    return f"{normalized}?"


def normalize_direct_user_question(text: str) -> str:
    normalized = normalize_chat_text(text)
    if not normalized:
        return ""
    normalized = normalized.rstrip(" .!")
    normalized = re.sub(r"(알려\s*줘요?|보여\s*줘요?|가르쳐\s*줘요?)\s*$", "", normalized)
    normalized = re.sub(r"(궁금해요|궁금합니다)\s*$", "", normalized)
    normalized = re.sub(r"(할 수 있나요|되는지요|되나요)\s*$", "할 수 있나요", normalized)
    normalized = normalized.rstrip(" .!")
    return ensure_question_sentence(normalized)


def extract_query_keywords(*texts: str) -> list[str]:
    keywords: list[str] = []
    seen: set[str] = set()
    for text in texts:
        for candidate in re.findall(r"[0-9A-Za-z가-힣]+", text):
            normalized = candidate.strip().lower()
            if len(normalized) < 2 or normalized in QUERY_VALIDATION_STOPWORDS:
                continue
            if normalized in seen:
                continue
            seen.add(normalized)
            keywords.append(candidate.strip())
    return keywords


def extract_year_tokens(text: str) -> list[str]:
    return [match.group(1) for match in YEAR_TOKEN_PATTERN.finditer(text or "")]


def extract_matching_keywords(text: str, keywords: tuple[str, ...]) -> list[str]:
    normalized_text = text or ""
    matched: list[str] = []
    for keyword in keywords:
        if keyword in normalized_text and keyword not in matched:
            matched.append(keyword)
    return matched


def is_statistics_or_numeric_query(text: str) -> bool:
    matched_keywords = extract_matching_keywords(text, STATISTICS_KEYWORDS)
    if "통계" in matched_keywords or "수치" in matched_keywords:
        return True
    return len(matched_keywords) >= 2


def contains_policy_or_claim_focus(text: str) -> bool:
    return any(keyword in (text or "") for keyword in POLICY_OR_CLAIM_KEYWORDS)


def extract_statistics_entities(text: str) -> dict[str, str]:
    normalized_text = normalize_chat_text(text)
    entities: dict[str, str] = {}
    years = extract_year_tokens(normalized_text)
    if years:
        entities["year"] = years[0]

    metric_keywords = extract_matching_keywords(normalized_text, STATISTICS_REQUIRED_PRESERVATION_KEYWORDS)
    if metric_keywords:
        entities["metric"] = ", ".join(metric_keywords[:3])

    procedure_match = PROCEDURE_TOKEN_PATTERN.search(normalized_text)
    if procedure_match:
        procedure = normalize_chat_text(procedure_match.group(1))
        entities["procedure"] = procedure
        entities.setdefault("target", procedure)
        entities.setdefault("topic", procedure)

    if "topic" not in entities and metric_keywords:
        entities["topic"] = metric_keywords[0]
    return entities


def get_statistics_query_expansions(query: str, entities: dict[str, str]) -> list[str]:
    normalized_query = normalize_chat_text(query).strip()
    procedure = (entities.get("procedure") or entities.get("target") or entities.get("topic") or "").strip()
    year = (entities.get("year") or "").strip()
    metric = (entities.get("metric") or "").split(",", 1)[0].strip()

    candidates = [
        normalized_query,
        f"{year}년 {procedure} {metric}".strip(),
        f"{procedure} 통계".strip(),
        f"{year}년 {procedure} 통계".strip(),
    ]
    return [candidate for candidate in candidates if candidate]


def normalize_external_search_chunk_types(raw_chunk_types: list[str]) -> list[str]:
    normalized_chunk_types: list[str] = []
    for chunk_type in raw_chunk_types:
        lowered = chunk_type.strip().lower()
        if "table" in lowered:
            normalized_chunk_types.append("table")
        elif "mixed" in lowered or "report" in lowered:
            normalized_chunk_types.append("mixed")
        elif lowered in {"text", "table", "mixed"}:
            normalized_chunk_types.append(lowered)

    unique_chunk_types: list[str] = []
    seen: set[str] = set()
    for chunk_type in normalized_chunk_types:
        if chunk_type in seen:
            continue
        seen.add(chunk_type)
        unique_chunk_types.append(chunk_type)
    return unique_chunk_types


def normalize_search_document_type(value: str | None) -> str | None:
    if not value:
        return None

    lowered = value.strip().lower()
    if not lowered:
        return None

    alias_map = {
        "policy": "policy",
        "terms": "policy",
        "약관": "policy",
        "규정": "policy",
        "calculation_guide": "calculation_guide",
        "pricing_method": "calculation_guide",
        "산출방법서": "calculation_guide",
        "보험료/해약환급금 산출방법서": "calculation_guide",
        "business_guide": "business_guide",
        "change_process": "business_guide",
        "업무가이드": "business_guide",
        "업무처리": "business_guide",
        "guide": "business_guide",
        "statistics_table": "statistics_table",
        "statistics_report": "statistics_table",
        "statistics": "statistics_table",
        "통계": "statistics_table",
    }
    normalized = alias_map.get(lowered, lowered)
    if normalized in SEARCH_DOCUMENT_TYPES:
        return normalized
    return None


def validate_standalone_search_query(
    query: str,
    *,
    source_texts: list[str],
) -> tuple[str, list[str]]:
    normalized_query = ensure_question_sentence(query)
    failure_reasons: list[str] = []

    if not normalized_query:
        failure_reasons.append("empty")
        return normalized_query, failure_reasons

    if normalized_query.startswith(INVALID_REWRITE_PREFIXES):
        failure_reasons.append("invalid_prefix")

    query_length = len(normalized_query)
    if query_length < MIN_STANDALONE_QUERY_LENGTH:
        failure_reasons.append("too_short")
    if query_length > MAX_STANDALONE_QUERY_LENGTH:
        failure_reasons.append("too_long")
    if not normalized_query.endswith("?"):
        failure_reasons.append("not_question")

    filler_hit_count = sum(1 for marker in CONVERSATIONAL_FILLER_PATTERNS if marker in normalized_query)
    if filler_hit_count >= 2:
        failure_reasons.append("contains_conversational_fillers")

    if any(marker in normalized_query for marker in VAGUE_REWRITE_MARKERS):
        query_keywords = extract_query_keywords(normalized_query)
        has_product_or_coverage = contains_insurance_product_clues(normalized_query) or "보장" in normalized_query
        if not has_product_or_coverage or len(query_keywords) < 3:
            failure_reasons.append("vague_referential_query")

    source_keywords = extract_query_keywords(*source_texts)
    source_text = normalize_chat_text(" ".join(source_texts))
    source_years = extract_year_tokens(source_text)
    required_statistics_keywords = extract_matching_keywords(source_text, STATISTICS_REQUIRED_PRESERVATION_KEYWORDS)
    source_is_statistics_query = is_statistics_or_numeric_query(source_text)

    if source_keywords:
        query_lower = normalized_query.lower()
        has_product_or_coverage = contains_insurance_product_clues(normalized_query) or "보장" in normalized_query
        if not has_product_or_coverage and not any(keyword.lower() in query_lower for keyword in source_keywords[:8]):
            failure_reasons.append("missing_core_keyword")

    should_preserve_year = source_is_statistics_query or "가입" in source_text or "기준" in source_text
    if source_years and should_preserve_year and not any(year in normalized_query for year in source_years):
        failure_reasons.append("missing_year")

    if required_statistics_keywords and not any(keyword in normalized_query for keyword in required_statistics_keywords):
        failure_reasons.append("missing_statistics_keyword")

    if source_is_statistics_query:
        rewritten_is_statistics_query = is_statistics_or_numeric_query(normalized_query)
        if not rewritten_is_statistics_query:
            failure_reasons.append("statistics_intent_lost")
        if contains_policy_or_claim_focus(normalized_query) and not any(
            keyword in normalized_query for keyword in required_statistics_keywords
        ):
            failure_reasons.append("statistics_query_shifted_to_policy")

    return normalized_query, failure_reasons


def is_standalone_query_rewrite(normalized_conversation: list[dict[str, str]]) -> bool:
    return not normalized_conversation


def choose_best_rewrite_candidate(
    candidates: list[str],
    *,
    source_texts: list[str],
) -> tuple[str, list[str], str | None]:
    for candidate in candidates:
        validated_candidate, reasons = validate_standalone_search_query(candidate, source_texts=source_texts)
        if not reasons:
            return validated_candidate, [], candidate
    if candidates:
        validated_candidate, reasons = validate_standalone_search_query(candidates[0], source_texts=source_texts)
        return validated_candidate, reasons, candidates[0]
    return "", ["empty"], None


def build_metadata_augmented_query(last_customer_message: str | None, metadata: object | None) -> str:
    base_text = normalize_chat_text(last_customer_message or "")
    if not base_text:
        return ""

    augment_parts = [
        value.strip()
        for value in [
            getattr(metadata, "product", "") if metadata else "",
            getattr(metadata, "document_type", "") if metadata else "",
        ]
        if value and value.strip()
    ]
    if not augment_parts:
        return ensure_question_sentence(base_text)

    return ensure_question_sentence(f"{' '.join(augment_parts)} 관련해서 {base_text}")


def tokenize_text(text: str) -> list[str]:
    return re.findall(r"\w+", text.lower())


def count_control_characters(text: str) -> int:
    return sum(1 for char in text if ord(char) < 32 and char not in {"\n", "\r", "\t"})


def count_replacement_characters(text: str) -> int:
    return text.count("\ufffd")


def calculate_suspicious_character_ratio(text: str) -> float:
    meaningful_characters = [char for char in text if not char.isspace()]
    if not meaningful_characters:
        return 0.0

    allowed_pattern = re.compile(r"[0-9A-Za-z가-힣ㄱ-ㅎㅏ-ㅣ.,:;!?()\[\]{}<>\-_/\\%&@#*+=~'\"`|]")
    suspicious_count = sum(1 for char in meaningful_characters if not allowed_pattern.fullmatch(char))
    return suspicious_count / len(meaningful_characters)


def calculate_text_length_ratio(left_text: str, right_text: str) -> float:
    right_length = max(len(right_text.strip()), 1)
    return len(left_text.strip()) / right_length


def get_default_auxiliary_parser(extension: str) -> str:
    if extension == ".pdf":
        return FALLBACK_PARSER_PYMUPDF
    if extension == ".docx":
        return FALLBACK_PARSER_PYTHON_DOCX
    if extension == ".doc":
        return FALLBACK_PARSER_DOC
    if extension in {".xls", ".xlsx"}:
        return FALLBACK_PARSER_EXCEL
    return FALLBACK_PARSER_NONE


def parser_supports_extension(parser_id: str, extension: str) -> bool:
    if parser_id in {PRIMARY_PARSER_DOCLING, PRIMARY_PARSER_DOCLING_MARKDOWN}:
        return extension in {".pdf", ".docx", ".xlsx"}
    if parser_id == PRIMARY_PARSER_LEGACY_AUTO:
        return extension in ALLOWED_EXTENSIONS
    if parser_id == FALLBACK_PARSER_PYMUPDF:
        return extension == ".pdf"
    if parser_id == FALLBACK_PARSER_PYTHON_DOCX:
        return extension == ".docx"
    if parser_id == FALLBACK_PARSER_DOC:
        return extension == ".doc"
    if parser_id == FALLBACK_PARSER_EXCEL:
        return extension in {".xls", ".xlsx"}
    return parser_id == FALLBACK_PARSER_NONE


def get_parser_label(parser_id: str) -> str:
    labels = {
        PRIMARY_PARSER_DOCLING: "Docling",
        PRIMARY_PARSER_DOCLING_MARKDOWN: "Docling(md)",
        PRIMARY_PARSER_LEGACY_AUTO: "Legacy auto parser",
        FALLBACK_PARSER_EXTENSION_DEFAULT: "Extension default parser",
        FALLBACK_PARSER_NONE: "No fallback",
        FALLBACK_PARSER_PYMUPDF: "PyMuPDF (PDF)",
        FALLBACK_PARSER_PYTHON_DOCX: "python-docx (DOCX)",
        FALLBACK_PARSER_DOC: "DOC parser",
        FALLBACK_PARSER_EXCEL: "Excel parser",
    }
    return labels.get(parser_id, parser_id)


def is_docling_available() -> bool:
    try:
        import docling  # noqa: F401
    except ModuleNotFoundError:
        return False
    return True


def is_command_available(command: str) -> bool:
    return shutil.which(command) is not None


def is_excel_parser_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        import xlrd  # noqa: F401
    except ModuleNotFoundError:
        return False
    return True


@lru_cache(maxsize=1)
def get_docling_converter():
    try:
        from docling.document_converter import DocumentConverter
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("Docling is not installed in the current environment.") from exc

    return DocumentConverter()


def get_parser_catalog() -> dict[str, object]:
    docling_available = is_docling_available()
    excel_parser_available = is_excel_parser_available()
    doc_parser_available = is_command_available("antiword")
    return {
        "default_primary_parser": PRIMARY_PARSER_LEGACY_AUTO,
        "default_fallback_parser": FALLBACK_PARSER_EXTENSION_DEFAULT,
        "primary_parsers": [
            {
                "id": PRIMARY_PARSER_DOCLING,
                "label": "Docling",
                "available": docling_available,
                "description": "비교 검증용 파서. PDF, DOCX, XLSX를 Docling으로 변환합니다.",
                "supported_extensions": [".pdf", ".docx", ".xlsx"],
            },
            {
                "id": PRIMARY_PARSER_DOCLING_MARKDOWN,
                "label": "Docling(md)",
                "available": docling_available,
                "description": "Docling만 사용하고, 선택된 파일을 Markdown 전용 산출물로 변환합니다.",
                "supported_extensions": [".pdf", ".docx", ".xlsx"],
            },
            {
                "id": PRIMARY_PARSER_LEGACY_AUTO,
                "label": "Legacy auto parser",
                "available": True,
                "description": "기본 운영 파서. 파일 확장자에 따라 검증된 내장 파서를 바로 사용합니다.",
                "supported_extensions": [".pdf", ".doc", ".docx", ".xls", ".xlsx"],
            },
        ],
        "fallback_parsers": [
            {
                "id": FALLBACK_PARSER_EXTENSION_DEFAULT,
                "label": "Extension default parser",
                "available": True,
                "description": "PDF는 PyMuPDF, DOCX는 python-docx, DOC는 antiword, Excel은 openpyxl/xlrd를 자동 선택합니다.",
                "supported_extensions": [".pdf", ".docx", ".doc", ".xls", ".xlsx"],
            },
            {
                "id": FALLBACK_PARSER_PYMUPDF,
                "label": "PyMuPDF (PDF)",
                "available": True,
                "description": "PDF 전용 보조 파서입니다.",
                "supported_extensions": [".pdf"],
            },
            {
                "id": FALLBACK_PARSER_PYTHON_DOCX,
                "label": "python-docx (DOCX)",
                "available": True,
                "description": "DOCX 전용 보조 파서입니다.",
                "supported_extensions": [".docx"],
            },
            {
                "id": FALLBACK_PARSER_DOC,
                "label": "DOC parser",
                "available": doc_parser_available,
                "description": "DOC 전용 보조 파서입니다. antiword를 사용합니다.",
                "supported_extensions": [".doc"],
            },
            {
                "id": FALLBACK_PARSER_EXCEL,
                "label": "Excel parser",
                "available": excel_parser_available,
                "description": "XLS/XLSX 보조 파서입니다. openpyxl과 xlrd를 사용합니다.",
                "supported_extensions": [".xls", ".xlsx"],
            },
            {
                "id": FALLBACK_PARSER_NONE,
                "label": "No fallback",
                "available": True,
                "description": "기본 파서 실패 시 보조 파서를 사용하지 않습니다.",
                "supported_extensions": [],
            },
        ],
    }


def build_token_frequency(tokens: list[str]) -> dict[str, int]:
    frequency: dict[str, int] = {}
    for token in tokens:
        frequency[token] = frequency.get(token, 0) + 1
    return frequency


def normalize_for_ngrams(text: str) -> str:
    return re.sub(r"\s+", "", text.lower())


def build_character_ngrams(text: str, min_n: int = 2, max_n: int = 4) -> set[str]:
    normalized = normalize_for_ngrams(text)
    ngrams: set[str] = set()

    for size in range(min_n, max_n + 1):
        if len(normalized) < size:
            continue
        for index in range(len(normalized) - size + 1):
            ngrams.add(normalized[index : index + size])

    return ngrams


def build_file_metadata(path: Path) -> dict[str, object]:
    return {
        "stored_name": path.name,
        "original_name": path.name.split("__", 1)[1] if "__" in path.name else path.name,
        "size_bytes": path.stat().st_size,
        "uploaded_at": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(),
    }


def get_parse_summary_path(path: Path) -> Path:
    ensure_parse_metadata_dir()
    return PARSE_METADATA_DIR / f"{path.name}.json"


def get_chunk_summary_path(path: Path) -> Path:
    ensure_chunk_metadata_dir()
    return CHUNK_METADATA_DIR / f"{path.name}.json"


def read_json_file(path: Path) -> dict[str, object] | None:
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def write_parse_summary(
    path: Path,
    *,
    status: str,
    parser_used: str,
    fallback_used: bool,
    text_length: int,
    file_type: str,
    preview: str | None = None,
    markdown_path: str | None = None,
    error_detail: str | None = None,
) -> Path:
    summary_path = get_parse_summary_path(path)
    summary = {
        "stored_name": path.name,
        "original_name": path.name.split("__", 1)[1] if "__" in path.name else path.name,
        "status": status,
        "file_type": file_type,
        "parser_used": parser_used,
        "fallback_used": fallback_used,
        "text_length": text_length,
        "preview": preview,
        "markdown_path": markdown_path,
        "error_detail": error_detail,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary_path


def write_parse_failure_summary(
    path: Path,
    *,
    file_type: str,
    parser_used: str,
    fallback_used: bool,
    error_detail: str,
) -> Path:
    return write_parse_summary(
        path,
        status="failed",
        parser_used=parser_used,
        fallback_used=fallback_used,
        text_length=0,
        file_type=file_type,
        preview=None,
        markdown_path=None,
        error_detail=error_detail,
    )


def update_parse_quality_summary(
    path: Path,
    *,
    parser_used: str,
    fallback_used: bool,
    metrics: dict[str, object],
) -> None:
    summary_path = get_parse_summary_path(path)
    summary = read_json_file(summary_path) or {
        "stored_name": path.name,
        "original_name": path.name.split("__", 1)[1] if "__" in path.name else path.name,
        "status": "completed",
        "file_type": get_extension(path.name),
    }
    summary.update(
        {
            "parser_used": parser_used,
            "fallback_used": fallback_used,
            "quality_checked_at": datetime.now(timezone.utc).isoformat(),
            "jaccard_similarity": metrics.get("jaccard_similarity"),
            "levenshtein_distance": metrics.get("levenshtein_distance"),
            "quality_warning": metrics.get("quality_warning"),
            "quality_warning_message": metrics.get("quality_warning_message"),
            "quality_warning_reasons": metrics.get("quality_warning_reasons"),
            "pdf_garbled_detected": metrics.get("pdf_garbled_detected"),
            "pdf_suspicious_char_ratio": metrics.get("pdf_suspicious_char_ratio"),
            "pdf_reference_suspicious_char_ratio": metrics.get("pdf_reference_suspicious_char_ratio"),
            "pdf_suspicious_char_ratio_delta": metrics.get("pdf_suspicious_char_ratio_delta"),
            "pdf_replacement_character_count": metrics.get("pdf_replacement_character_count"),
            "pdf_control_character_count": metrics.get("pdf_control_character_count"),
            "pdf_text_length_ratio": metrics.get("pdf_text_length_ratio"),
        }
    )
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def list_uploaded_files() -> list[dict[str, object]]:
    ensure_upload_dir()
    files: list[dict[str, object]] = []
    for path in sorted(UPLOAD_DIR.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True):
        if not path.is_file():
            continue
        files.append(build_file_metadata(path))
    return files


def list_parsed_files() -> list[dict[str, object]]:
    ensure_parse_metadata_dir()
    files: list[dict[str, object]] = []
    for path in sorted(PARSE_METADATA_DIR.iterdir(), key=lambda item: item.name):
        if not path.is_file() or path.suffix.lower() != ".json":
            continue
        summary = read_json_file(path)
        if summary:
            files.append(summary)
    return files


def list_chunked_files() -> list[dict[str, object]]:
    ensure_chunk_metadata_dir()
    files: list[dict[str, object]] = []
    for path in sorted(CHUNK_METADATA_DIR.iterdir(), key=lambda item: item.name):
        if not path.is_file() or path.suffix.lower() != ".json":
            continue
        summary = read_json_file(path)
        if summary:
            files.append(summary)
    return files


def list_default_files() -> list[dict[str, object]]:
    ensure_default_file_dir()
    files: list[dict[str, object]] = []
    for path in sorted(DEFAULT_FILE_DIR.iterdir(), key=lambda item: item.name):
        if not path.is_file():
            continue
        if get_extension(path.name) not in ALLOWED_EXTENSIONS:
            continue
        files.append(
            {
                "name": path.name,
                "size_bytes": path.stat().st_size,
            }
        )
    return files


def save_content_to_uploads(filename: str, content: bytes, content_type: str | None) -> dict[str, object]:
    ensure_upload_dir()
    stored_name = f"{uuid4().hex}__{Path(filename).name}"
    destination = UPLOAD_DIR / stored_name
    destination.write_bytes(content)
    metadata = build_file_metadata(destination)
    metadata.update(
        {
            "status": "uploaded",
            "content_type": content_type,
        }
    )
    return metadata


def ensure_no_completed_duplicate_upload(filename: str) -> None:
    normalized_name = Path(filename).name
    for file in list_pipeline_files():
        original_name = file.get("original_name")
        index_status = file.get("index_status")
        if original_name == normalized_name and index_status == "completed":
            raise HTTPException(
                status_code=409,
                detail=f"{normalized_name} is already uploaded and indexed. Reset the existing file before uploading it again.",
            )


def get_uploaded_file_path(stored_name: str) -> Path:
    ensure_upload_dir()
    filename = Path(stored_name).name
    path = UPLOAD_DIR / filename

    if not path.is_file():
        raise HTTPException(status_code=404, detail="Uploaded file not found.")

    extension = get_extension(path.name)
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=get_supported_file_types_message())

    return path


def extract_pdf_text(path: Path) -> str:
    document = fitz.open(path)
    try:
        pages = [page.get_text("text") for page in document]
    finally:
        document.close()
    return normalize_text("\n".join(pages))


def extract_docx_text(path: Path) -> str:
    document = Document(path)
    sections: list[str] = []

    sections.extend(extract_docx_container_text(document))

    for section in document.sections:
        sections.extend(extract_docx_container_text(section.header))
        sections.extend(extract_docx_container_text(section.footer))

    return normalize_text("\n".join(section for section in sections if section.strip()))


def extract_doc_text(path: Path) -> str:
    if not is_command_available("antiword"):
        raise RuntimeError("antiword is not installed in the current environment.")

    completed = subprocess.run(
        ["antiword", str(path)],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
    )
    if completed.returncode != 0:
        stderr = completed.stderr.strip()
        raise RuntimeError(stderr or "antiword failed to parse the DOC file.")

    return normalize_text(completed.stdout)


def format_excel_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_xlsx_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("openpyxl is not installed in the current environment.") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheets: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[str] = [f"# Sheet: {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                values = [format_excel_cell(cell) for cell in row]
                values = [value for value in values if value]
                if values:
                    rows.append(" | ".join(values))
            if len(rows) > 1:
                sheets.append("\n".join(rows))
    finally:
        workbook.close()

    return normalize_text("\n\n".join(sheets))


def extract_xls_text(path: Path) -> str:
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("xlrd is not installed in the current environment.") from exc

    workbook = xlrd.open_workbook(path)
    sheets: list[str] = []
    for sheet in workbook.sheets():
        rows: list[str] = [f"# Sheet: {sheet.name}"]
        for row_index in range(sheet.nrows):
            values = [format_excel_cell(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            values = [value for value in values if value]
            if values:
                rows.append(" | ".join(values))
        if len(rows) > 1:
            sheets.append("\n".join(rows))

    return normalize_text("\n\n".join(sheets))


def extract_excel_text(path: Path) -> str:
    extension = get_extension(path.name)
    if extension == ".xlsx":
        return extract_xlsx_text(path)
    if extension == ".xls":
        return extract_xls_text(path)
    raise HTTPException(status_code=400, detail="Excel parser is only available for XLS/XLSX files.")


def iter_docx_blocks(parent: DocxDocument | _Cell | object) -> list[Paragraph | Table]:
    if isinstance(parent, DocxDocument):
        parent_element = parent.element.body
    elif hasattr(parent, "_tc"):
        parent_element = parent._tc
    else:
        parent_element = parent._element

    blocks: list[Paragraph | Table] = []
    for child in parent_element.iterchildren():
        if isinstance(child, CT_P):
            blocks.append(Paragraph(child, parent))
            continue
        if isinstance(child, CT_Tbl):
            blocks.append(Table(child, parent))
    return blocks


def extract_docx_table_text(table: Table) -> str:
    rows: list[str] = []

    for row in table.rows:
        cells: list[str] = []
        for cell in row.cells:
            cell_sections = extract_docx_container_text(cell)
            cell_text = " ".join(part for part in cell_sections if part.strip())
            if cell_text:
                cells.append(cell_text)
        if cells:
            rows.append(" | ".join(cells))

    return "\n".join(rows)


def extract_docx_container_text(container: DocxDocument | _Cell | object) -> list[str]:
    parts: list[str] = []

    for block in iter_docx_blocks(container):
        if isinstance(block, Paragraph):
            if block.text.strip():
                parts.append(block.text)
            continue

        table_text = extract_docx_table_text(block)
        if table_text:
            parts.append(table_text)

    return parts


def extract_docling_text(path: Path) -> str:
    converter = get_docling_converter()
    result = converter.convert(path)
    text = result.document.export_to_markdown()
    return normalize_text(text)


def get_markdown_output_path(path: Path) -> Path:
    ensure_markdown_output_dir()
    return MARKDOWN_OUTPUT_DIR / f"{path.stem}.md"


def write_docling_markdown_file(path: Path, markdown_text: str) -> Path:
    output_path = get_markdown_output_path(path)
    output_path.write_text(markdown_text, encoding="utf-8")
    return output_path


def extract_with_parser(path: Path, parser_id: str) -> tuple[str, str]:
    extension = get_extension(path.name)

    if parser_id == PRIMARY_PARSER_DOCLING:
        if not parser_supports_extension(parser_id, extension):
            raise HTTPException(status_code=400, detail=f"Docling does not support {extension} in the current parser policy.")
        return extract_docling_text(path), parser_id

    if parser_id == PRIMARY_PARSER_DOCLING_MARKDOWN:
        if not parser_supports_extension(parser_id, extension):
            raise HTTPException(status_code=400, detail=f"Docling(md) does not support {extension} in the current parser policy.")
        return extract_docling_text(path), parser_id

    if parser_id in {PRIMARY_PARSER_LEGACY_AUTO, FALLBACK_PARSER_EXTENSION_DEFAULT}:
        resolved_parser = get_default_auxiliary_parser(extension)
        return extract_with_parser(path, resolved_parser)

    if parser_id == FALLBACK_PARSER_PYMUPDF:
        if extension != ".pdf":
            raise HTTPException(status_code=400, detail="PyMuPDF fallback is only available for PDF files.")
        return extract_pdf_text(path), parser_id

    if parser_id == FALLBACK_PARSER_PYTHON_DOCX:
        if extension != ".docx":
            raise HTTPException(status_code=400, detail="python-docx fallback is only available for DOCX files.")
        return extract_docx_text(path), parser_id

    if parser_id == FALLBACK_PARSER_DOC:
        if extension != ".doc":
            raise HTTPException(status_code=400, detail="DOC fallback parser is only available for DOC files.")
        return extract_doc_text(path), parser_id

    if parser_id == FALLBACK_PARSER_EXCEL:
        if extension not in {".xls", ".xlsx"}:
            raise HTTPException(status_code=400, detail="Excel fallback parser is only available for XLS/XLSX files.")
        return extract_excel_text(path), parser_id

    if parser_id == FALLBACK_PARSER_NONE:
        raise HTTPException(status_code=400, detail="No fallback parser is configured.")

    raise HTTPException(status_code=400, detail=f"Unsupported parser selection: {parser_id}")


def extract_text_from_file(
    path: Path,
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
) -> tuple[str, str, bool]:
    attempts: list[str] = []
    seen: set[str] = set()

    for parser_id in [primary_parser, fallback_parser]:
        if parser_id in seen:
            continue
        seen.add(parser_id)
        try:
            logger.info("parser_attempt stored_name=%s parser=%s", path.name, parser_id)
            text, parser_used = extract_with_parser(path, parser_id)
            logger.info("parser_success stored_name=%s parser=%s", path.name, parser_used)
            return text, parser_used, parser_used != primary_parser
        except (HTTPException, ModuleNotFoundError, RuntimeError, ValueError) as exc:
            logger.warning("parser_failed stored_name=%s parser=%s reason=%s", path.name, parser_id, str(exc))
            attempts.append(f"{get_parser_label(parser_id)}: {str(exc)}")
            continue

    raise HTTPException(
        status_code=400,
        detail="Parsing failed. " + " | ".join(attempts) if attempts else "Parsing failed.",
    )


def extract_reference_pdf_text(path: Path) -> str:
    document = fitz.open(path)
    try:
        pages: list[str] = []
        for page in document:
            blocks = page.get_text("blocks")
            block_texts = [block[4] for block in blocks if len(block) > 4 and str(block[4]).strip()]
            pages.append("\n".join(block_texts))
    finally:
        document.close()
    return normalize_text("\n".join(pages))


def extract_reference_docx_text(path: Path) -> str:
    parts: list[str] = []
    with ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            if not name.startswith("word/"):
                continue
            if not name.endswith(".xml"):
                continue
            if not (name == "word/document.xml" or name.startswith("word/header") or name.startswith("word/footer")):
                continue

            xml_bytes = archive.read(name)
            root = ET.fromstring(xml_bytes)
            texts = [node.text for node in root.findall(".//w:t", WORD_NAMESPACE) if node.text]
            if texts:
                parts.append(" ".join(texts))

    return normalize_text("\n".join(parts))


def extract_reference_text(path: Path) -> str:
    extension = get_extension(path.name)

    if extension == ".pdf":
        return extract_reference_pdf_text(path)

    if extension == ".docx":
        return extract_reference_docx_text(path)

    if extension == ".doc":
        return extract_doc_text(path)

    if extension in {".xls", ".xlsx"}:
        return extract_excel_text(path)

    raise HTTPException(status_code=400, detail=get_supported_file_types_message())


def calculate_jaccard_similarity(left_text: str, right_text: str) -> float:
    left_tokens = set(tokenize_text(left_text))
    right_tokens = set(tokenize_text(right_text))

    if not left_tokens and not right_tokens:
        return 1.0

    union = left_tokens | right_tokens
    if not union:
        return 0.0

    return len(left_tokens & right_tokens) / len(union)


def calculate_levenshtein_distance(left_tokens: list[str], right_tokens: list[str]) -> int:
    if len(left_tokens) < len(right_tokens):
        left_tokens, right_tokens = right_tokens, left_tokens

    previous = list(range(len(right_tokens) + 1))
    for i, left_token in enumerate(left_tokens, start=1):
        current = [i]
        for j, right_token in enumerate(right_tokens, start=1):
            insert_cost = current[j - 1] + 1
            delete_cost = previous[j] + 1
            replace_cost = previous[j - 1] + (0 if left_token == right_token else 1)
            current.append(min(insert_cost, delete_cost, replace_cost))
        previous = current

    return previous[-1]


def build_pdf_garbled_warnings(parsed_text: str, reference_text: str) -> tuple[bool, list[str], dict[str, object]]:
    suspicious_char_ratio = calculate_suspicious_character_ratio(parsed_text)
    reference_suspicious_char_ratio = calculate_suspicious_character_ratio(reference_text)
    replacement_character_count = count_replacement_characters(parsed_text)
    control_character_count = count_control_characters(parsed_text)
    text_length_ratio = calculate_text_length_ratio(parsed_text, reference_text)
    suspicious_char_ratio_delta = suspicious_char_ratio - reference_suspicious_char_ratio

    warnings: list[str] = []
    if replacement_character_count > 0:
        warnings.append("PDF text contains replacement characters.")
    if control_character_count > 0 and (control_character_count / max(len(parsed_text), 1)) > PDF_GARBLED_CONTROL_CHAR_RATIO_THRESHOLD:
        warnings.append("PDF text contains unexpected control characters.")
    if (
        suspicious_char_ratio > PDF_GARBLED_SUSPICIOUS_CHAR_RATIO_THRESHOLD
        and suspicious_char_ratio_delta > PDF_GARBLED_SUSPICIOUS_CHAR_RATIO_DELTA_THRESHOLD
    ):
        warnings.append("PDF text contains too many suspicious symbols.")
    if reference_text.strip() and text_length_ratio < PDF_GARBLED_LENGTH_RATIO_THRESHOLD:
        warnings.append("PDF text is much shorter than the reference extraction.")

    metrics = {
        "pdf_garbled_detected": bool(warnings),
        "pdf_suspicious_char_ratio": suspicious_char_ratio,
        "pdf_reference_suspicious_char_ratio": reference_suspicious_char_ratio,
        "pdf_suspicious_char_ratio_delta": suspicious_char_ratio_delta,
        "pdf_replacement_character_count": replacement_character_count,
        "pdf_control_character_count": control_character_count,
        "pdf_text_length_ratio": text_length_ratio,
    }
    return bool(warnings), warnings, metrics


def build_quality_metrics(parsed_text: str, reference_text: str, *, file_type: str) -> dict[str, object]:
    parsed_tokens = tokenize_text(parsed_text)
    reference_tokens = tokenize_text(reference_text)
    jaccard_similarity = calculate_jaccard_similarity(parsed_text, reference_text)
    levenshtein_distance = calculate_levenshtein_distance(parsed_tokens, reference_tokens)
    warning_reasons: list[str] = []

    if jaccard_similarity < QUALITY_WARNING_THRESHOLD:
        warning_reasons.append("Reference similarity is below the warning threshold.")

    pdf_garbled_detected = False
    extra_metrics: dict[str, object] = {}
    if file_type == ".pdf":
        pdf_garbled_detected, pdf_warnings, extra_metrics = build_pdf_garbled_warnings(parsed_text, reference_text)
        warning_reasons.extend(pdf_warnings)

    quality_warning = bool(warning_reasons)

    return {
        "jaccard_similarity": jaccard_similarity,
        "levenshtein_distance": levenshtein_distance,
        "quality_warning": quality_warning,
        "quality_warning_message": " / ".join(warning_reasons),
        "quality_warning_reasons": warning_reasons,
        "pdf_garbled_detected": pdf_garbled_detected,
        **extra_metrics,
    }


def build_parsing_result(
    path: Path,
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
) -> dict[str, object]:
    fallback_parser = normalize_fallback_parser(primary_parser, fallback_parser)
    text, parser_used, fallback_used = extract_text_from_file(
        path,
        primary_parser=primary_parser,
        fallback_parser=fallback_parser,
    )

    if not text.strip():
        raise HTTPException(status_code=400, detail="Extracted text is empty.")

    markdown_path: str | None = None
    if parser_used == PRIMARY_PARSER_DOCLING_MARKDOWN:
        markdown_path = str(write_docling_markdown_file(path, text))

    write_parse_summary(
        path,
        status="completed",
        parser_used=parser_used,
        fallback_used=fallback_used,
        text_length=len(text),
        file_type=get_extension(path.name),
        preview=text[:PREVIEW_LENGTH],
        markdown_path=markdown_path,
    )

    metadata = build_file_metadata(path)
    metadata.update(
        {
            "status": "parsed",
            "file_type": get_extension(path.name),
            "text_length": len(text),
            "preview": text[:PREVIEW_LENGTH],
            "extracted_text": text,
            "primary_parser": primary_parser,
            "fallback_parser": fallback_parser,
            "parser_used": parser_used,
            "fallback_used": fallback_used,
            "markdown_path": markdown_path,
        }
    )
    return metadata


def build_parsing_quality_result(
    path: Path,
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
) -> dict[str, object]:
    fallback_parser = normalize_fallback_parser(primary_parser, fallback_parser)
    parsed_text, parser_used, fallback_used = extract_text_from_file(
        path,
        primary_parser=primary_parser,
        fallback_parser=fallback_parser,
    )
    reference_text = extract_reference_text(path)

    if not parsed_text.strip():
        raise HTTPException(status_code=400, detail="Extracted text is empty.")

    metadata = build_file_metadata(path)
    quality_checked_at = datetime.now(timezone.utc).isoformat()
    metadata.update(
        {
            "status": "quality_checked",
            "file_type": get_extension(path.name),
            "text_length": len(parsed_text),
            "reference_text_length": len(reference_text),
            "quality_checked_at": quality_checked_at,
            "primary_parser": primary_parser,
            "fallback_parser": fallback_parser,
            "parser_used": parser_used,
            "fallback_used": fallback_used,
        }
    )
    file_type = get_extension(path.name)
    quality_metrics = build_quality_metrics(parsed_text, reference_text, file_type=file_type)
    metadata.update(quality_metrics)
    update_parse_quality_summary(
        path,
        parser_used=parser_used,
        fallback_used=fallback_used,
        metrics=quality_metrics,
    )
    return metadata


def split_long_segment(segment: str, max_length: int) -> list[str]:
    normalized_segment = segment.strip()
    if not normalized_segment:
        return []

    if len(normalized_segment) <= max_length:
        return [normalized_segment]

    sentences = [
        part.strip()
        for part in re.split(r"(?<=[.!?])\s+|(?<=[.!?])\n+|\n+", normalized_segment)
        if part.strip()
    ]

    if len(sentences) == 1 and len(sentences[0]) > max_length:
        chunks: list[str] = []
        start = 0
        while start < len(normalized_segment):
            end = min(start + max_length, len(normalized_segment))
            chunks.append(normalized_segment[start:end].strip())
            start = end
        return [chunk for chunk in chunks if chunk]

    segments: list[str] = []
    current = ""
    for sentence in sentences:
        candidate = sentence if not current else f"{current} {sentence}"
        if len(candidate) <= max_length:
            current = candidate
            continue
        if current:
            segments.append(current)
        if len(sentence) > max_length:
            segments.extend(split_long_segment(sentence, max_length))
            current = ""
            continue
        current = sentence

    if current:
        segments.append(current)

    return segments


def normalize_chunk_settings(
    target_length: int | None = None,
    overlap_length: int | None = None,
) -> tuple[int, int]:
    resolved_target_length = target_length if target_length is not None else CHUNK_TARGET_LENGTH
    resolved_overlap_length = overlap_length if overlap_length is not None else CHUNK_OVERLAP_LENGTH

    if resolved_target_length < 100:
        raise HTTPException(status_code=400, detail="chunk target_length must be at least 100.")
    if resolved_overlap_length < 0:
        raise HTTPException(status_code=400, detail="chunk overlap_length must be 0 or greater.")
    if resolved_overlap_length >= resolved_target_length:
        raise HTTPException(status_code=400, detail="chunk overlap_length must be smaller than target_length.")

    return resolved_target_length, resolved_overlap_length


def build_chunk_segments(text: str, target_length: int) -> list[str]:
    raw_segments = [line.strip() for line in text.splitlines() if line.strip()]
    segments: list[str] = []

    for raw_segment in raw_segments:
        segments.extend(split_long_segment(raw_segment, target_length))

    return segments


def build_overlap_text(chunk_text: str, overlap_length: int) -> str:
    if len(chunk_text) <= overlap_length:
        return chunk_text
    return chunk_text[-overlap_length:].strip()


def create_chunks(
    text: str,
    *,
    source: str,
    target_length: int = CHUNK_TARGET_LENGTH,
    overlap_length: int = CHUNK_OVERLAP_LENGTH,
    page_number: int | None = None,
    section_header: str | None = None,
) -> list[dict[str, object]]:
    segments = build_chunk_segments(text, target_length)
    if not segments:
        return []

    chunks: list[dict[str, object]] = []
    current = ""
    current_start = 0
    cursor = 0

    for segment in segments:
        segment_start = cursor
        cursor += len(segment)

        candidate = segment if not current else f"{current}\n{segment}"
        if current and len(candidate) > target_length:
            chunk_end = current_start + len(current)
            chunks.append(
                {
                    "chunk_index": len(chunks),
                    "source": source,
                    "text": current,
                    "text_length": len(current),
                    "start_char": current_start,
                    "end_char": chunk_end,
                    "page_number": page_number,
                    "section_header": section_header,
                    "preview": current[:CHUNK_PREVIEW_LENGTH],
                }
            )
            overlap = build_overlap_text(current, overlap_length)
            current = segment if not overlap else f"{overlap}\n{segment}"
            current_start = max(chunk_end - len(overlap), 0)
            cursor = max(cursor, current_start + len(current))
            continue

        if not current:
            current_start = segment_start
        current = candidate

    if current:
        chunk_end = current_start + len(current)
        chunks.append(
            {
                "chunk_index": len(chunks),
                "source": source,
                "text": current,
                "text_length": len(current),
                "start_char": current_start,
                "end_char": chunk_end,
                "page_number": page_number,
                "section_header": section_header,
                "preview": current[:CHUNK_PREVIEW_LENGTH],
            }
        )

    return chunks


def write_chunk_summary(path: Path, chunk_count: int, *, target_length: int, overlap_length: int) -> Path:
    summary_path = get_chunk_summary_path(path)
    summary = {
        "stored_name": path.name,
        "original_name": path.name.split("__", 1)[1] if "__" in path.name else path.name,
        "chunk_count": chunk_count,
        "chunk_target_length": target_length,
        "chunk_overlap_length": overlap_length,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary_path


def build_chunking_result(
    path: Path,
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
    chunk_target_length: int | None = None,
    chunk_overlap_length: int | None = None,
) -> dict[str, object]:
    fallback_parser = normalize_fallback_parser(primary_parser, fallback_parser)
    resolved_target_length, resolved_overlap_length = normalize_chunk_settings(
        chunk_target_length,
        chunk_overlap_length,
    )
    parsed = build_parsing_result(
        path,
        primary_parser=primary_parser,
        fallback_parser=fallback_parser,
    )
    chunks = create_chunks(
        parsed["extracted_text"],
        source=parsed["original_name"],
        target_length=resolved_target_length,
        overlap_length=resolved_overlap_length,
    )

    if not chunks:
        raise HTTPException(status_code=400, detail="No chunks were created from extracted text.")

    summary_path = write_chunk_summary(
        path,
        len(chunks),
        target_length=resolved_target_length,
        overlap_length=resolved_overlap_length,
    )
    return {
        "status": "chunked",
        "stored_name": parsed["stored_name"],
        "original_name": parsed["original_name"],
        "file_type": parsed["file_type"],
        "text_length": parsed["text_length"],
        "chunk_count": len(chunks),
        "chunk_target_length": resolved_target_length,
        "chunk_overlap_length": resolved_overlap_length,
        "summary_path": str(summary_path),
        "primary_parser": primary_parser,
        "fallback_parser": fallback_parser,
        "parser_used": parsed["parser_used"],
        "fallback_used": parsed["fallback_used"],
        "chunks": chunks,
    }


class DefaultFileUploadRequest(BaseModel):
    filename: str


class ParseRequest(BaseModel):
    stored_name: str
    primary_parser: str = PRIMARY_PARSER_DOCLING
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT
    chunk_target_length: int | None = None
    chunk_overlap_length: int | None = None


class RetrieveRequest(BaseModel):
    query: str
    top_k: int = 5
    stored_name: str | None = None


class ConversationTurn(BaseModel):
    role: str
    content: str | None = None
    text: str | None = None


class ChatMetadata(BaseModel):
    product: str | None = None
    document_type: str | None = None
    locale: str | None = "ko-KR"


class ChatRequest(BaseModel):
    query: str
    top_k: int = DEFAULT_CHAT_TOP_K
    final_k: int = DEFAULT_CHAT_FINAL_K
    stored_name: str | None = None
    document_id: str | None = None
    section_hint: str | None = None
    search_api_endpoint: str | None = None
    action: str | None = "search"
    query_rewrite_model: str | None = None
    query_rewrite_base_url: str | None = None
    query_rewrite_custom_model: str | None = None
    query_rewrite_api_key: str | None = None
    answer_model: str | None = None
    answer_base_url: str | None = None
    answer_custom_model: str | None = None
    answer_api_key: str | None = None
    stream: bool = False
    conversation_context: list[ConversationTurn] = Field(default_factory=list)
    metadata: ChatMetadata | None = None


class RewriteResult(BaseModel):
    original_query: str
    rewritten_query: str
    search_queries: list[str] = Field(default_factory=list)
    document_type_filters: list[str] = Field(default_factory=list)
    query_rewrite_model: str | None = None
    intent: str | None = None
    question_type: str | None = None
    entities: dict[str, str] = Field(default_factory=dict)
    routing_hints: dict[str, str] = Field(default_factory=dict)
    normalized_conversation: list[dict[str, str]] = Field(default_factory=list)
    last_customer_message: str | None = None
    validation_reasons: list[str] = Field(default_factory=list)
    rewrite_source: str | None = None


class SearchEvaluationResult(BaseModel):
    need_more_context: bool = False
    reason_codes: list[str] = Field(default_factory=list)
    evaluation_reasons: list[str] = Field(default_factory=list)
    top_document_id: str | None = None
    top_chunk_id: str | None = None
    top_chunk_text_length: int = 0
    same_document_hit_count: int = 0


class SearchPhaseExecutionError(Exception):
    def __init__(
        self,
        *,
        cause: HTTPException,
        action: str,
        search_api_endpoint: str,
        response_time_ms: int,
    ) -> None:
        super().__init__(str(cause.detail))
        self.cause = cause
        self.action = action
        self.search_api_endpoint = search_api_endpoint
        self.response_time_ms = response_time_ms


def resolve_chat_search_limits(payload: ChatRequest) -> tuple[int, int]:
    requested_top_k = max(1, int(payload.top_k or 0))
    requested_final_k = max(1, int(payload.final_k or 0))

    resolved_final_k = max(DEFAULT_CHAT_FINAL_K, requested_final_k)
    resolved_top_k = max(DEFAULT_CHAT_TOP_K, requested_top_k, resolved_final_k)
    return resolved_top_k, resolved_final_k


def serialize_search_evaluation(result: SearchEvaluationResult) -> dict[str, object]:
    if hasattr(result, "model_dump"):
        return result.model_dump()
    return result.dict()


class RebuildIndexRequest(BaseModel):
    primary_parser: str = PRIMARY_PARSER_DOCLING
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT
    chunk_target_length: int | None = None
    chunk_overlap_length: int | None = None


def hash_token(token: str) -> int:
    return int.from_bytes(hashlib.sha256(token.encode("utf-8")).digest()[:8], "big")


def build_hash_embedding(text: str, dimension: int = EMBEDDING_DIMENSION) -> list[float]:
    vector = [0.0] * dimension
    tokens = tokenize_text(text)

    if not tokens:
        return vector

    for token in tokens:
        token_hash = hash_token(token)
        index = token_hash % dimension
        sign = -1.0 if ((token_hash >> 8) & 1) else 1.0
        vector[index] += sign

    norm = sum(value * value for value in vector) ** 0.5
    if norm == 0:
        return vector

    return [value / norm for value in vector]


def get_embedding_provider() -> str:
    provider = os.getenv("EMBEDDING_PROVIDER", EMBEDDING_PROVIDER_HASH).strip().lower()
    if provider not in {EMBEDDING_PROVIDER_HASH, EMBEDDING_PROVIDER_AZURE_OPENAI}:
        raise HTTPException(status_code=500, detail=f"Unsupported embedding provider: {provider}")
    return provider


def get_embedding_model_id() -> str:
    provider = get_embedding_provider()
    if provider == EMBEDDING_PROVIDER_HASH:
        return f"hash-{EMBEDDING_DIMENSION}"

    deployment = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", DEFAULT_AZURE_OPENAI_EMBEDDING_DEPLOYMENT).strip()
    if not deployment:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_EMBEDDING_DEPLOYMENT is not configured.")
    return deployment


def get_embedding_collection_name() -> str:
    provider = get_embedding_provider()
    model_id = re.sub(r"[^a-zA-Z0-9_-]+", "-", get_embedding_model_id()).strip("-").lower()
    return f"insurance_document_chunks__{provider}__{model_id}"


def get_embedding_dimension() -> int:
    provider = get_embedding_provider()
    if provider == EMBEDDING_PROVIDER_HASH:
        return EMBEDDING_DIMENSION
    return 1536


def get_embedding_config() -> dict[str, object]:
    provider = get_embedding_provider()
    config: dict[str, object] = {
        "provider": provider,
        "model": get_embedding_model_id(),
        "collection_name": get_embedding_collection_name(),
        "embedding_dimension": get_embedding_dimension(),
    }

    if provider == EMBEDDING_PROVIDER_AZURE_OPENAI:
        config["api_version"] = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION)

    return config


def build_azure_openai_embeddings(texts: list[str]) -> list[list[float]]:
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip()
    deployment = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", DEFAULT_AZURE_OPENAI_EMBEDDING_DEPLOYMENT).strip()
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()

    if not endpoint:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_ENDPOINT is not configured.")
    if not api_key:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_API_KEY is not configured.")
    if not deployment:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_EMBEDDING_DEPLOYMENT is not configured.")

    embeddings: list[list[float]] = []
    for start in range(0, len(texts), AZURE_OPENAI_EMBEDDING_BATCH_SIZE):
        batch = texts[start : start + AZURE_OPENAI_EMBEDDING_BATCH_SIZE]
        request_url = (
            f"{endpoint}/openai/deployments/{deployment}/embeddings"
            f"?api-version={api_version}"
        )
        payload = json.dumps({"input": batch}).encode("utf-8")
        request = urllib_request.Request(
            request_url,
            data=payload,
            method="POST",
            headers={
                "Content-Type": "application/json",
                "api-key": api_key,
            },
        )

        try:
            with urllib_request.urlopen(request, timeout=120) as response:
                raw_body = response.read()
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            raise HTTPException(status_code=502, detail=f"Azure OpenAI embedding request failed: {detail}") from exc
        except urllib_error.URLError as exc:
            raise HTTPException(status_code=502, detail=f"Azure OpenAI embedding request failed: {exc.reason}") from exc

        try:
            response_payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=502, detail="Azure OpenAI embedding response was not valid JSON.") from exc

        data = response_payload.get("data")
        if not isinstance(data, list) or len(data) != len(batch):
            raise HTTPException(status_code=502, detail="Azure OpenAI embedding response shape was invalid.")

        sorted_items = sorted(data, key=lambda item: int(item.get("index", 0)) if isinstance(item, dict) else 0)
        for item in sorted_items:
            if not isinstance(item, dict) or not isinstance(item.get("embedding"), list):
                raise HTTPException(status_code=502, detail="Azure OpenAI embedding response was missing vectors.")
            embeddings.append([float(value) for value in item["embedding"]])

    return embeddings


def get_chat_model_id() -> str:
    candidates = [
        os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT", "").strip(),
        os.getenv("AZURE_OPENAI_COMPLETION_DEPLOYMENT", "").strip(),
        os.getenv("AZURE_OPENAI_LLM_DEPLOYMENT", "").strip(),
    ]
    for candidate in candidates:
        if candidate:
            return candidate
    raise HTTPException(status_code=500, detail="AZURE_OPENAI_CHAT_DEPLOYMENT is not configured.")


def is_custom_query_rewrite_model(requested_model: str | None = None) -> bool:
    return (requested_model or "").strip().lower() == CUSTOM_QUERY_REWRITE_MODEL


def get_query_rewrite_model_id(requested_model: str | None = None) -> str:
    requested = (requested_model or "").strip()
    if not requested:
        return os.getenv("AZURE_OPENAI_QUERY_REWRITE_DEPLOYMENT", DEFAULT_QUERY_REWRITE_MODEL).strip() or DEFAULT_QUERY_REWRITE_MODEL
    if is_custom_query_rewrite_model(requested):
        return CUSTOM_QUERY_REWRITE_MODEL

    configured_models = {
        get_chat_model_id(),
        DEFAULT_QUERY_REWRITE_MODEL,
        "gpt-5.4",
        "gpt-5.4-mini",
        "gpt-4.1-mini",
        "gpt-4o",
    }
    extra_models = os.getenv("AZURE_OPENAI_QUERY_REWRITE_DEPLOYMENTS", "").strip()
    if extra_models:
        configured_models.update(model.strip() for model in extra_models.split(",") if model.strip())

    if requested not in configured_models:
        raise HTTPException(status_code=400, detail=f"Unsupported query rewrite model: {requested}")
    return requested


def get_query_rewrite_display_model(payload: ChatRequest) -> str:
    if is_custom_query_rewrite_model(payload.query_rewrite_model):
        custom_model = (payload.query_rewrite_custom_model or "").strip()
        if custom_model:
            return f"{CUSTOM_QUERY_REWRITE_MODEL}:{custom_model}"
        return CUSTOM_QUERY_REWRITE_MODEL
    return get_query_rewrite_model_id(payload.query_rewrite_model)


def get_query_rewrite_custom_config(payload: ChatRequest) -> tuple[str, str, str | None]:
    base_url = (payload.query_rewrite_base_url or "").strip()
    model_name = (payload.query_rewrite_custom_model or "").strip()
    api_key = (payload.query_rewrite_api_key or "").strip() or None

    if not base_url:
        raise HTTPException(status_code=400, detail="Custom query rewrite LLM endpoint is required.")
    if not model_name:
        raise HTTPException(status_code=400, detail="Custom query rewrite LLM model name is required.")

    return (base_url, model_name, api_key)


def is_custom_answer_model(requested_model: str | None = None) -> bool:
    return (requested_model or "").strip().lower() == CUSTOM_QUERY_REWRITE_MODEL


def get_answer_model_id(requested_model: str | None = None) -> str:
    requested = (requested_model or "").strip()
    if not requested:
        configured_default = os.getenv("AZURE_OPENAI_ANSWER_DEPLOYMENT", "").strip()
        return configured_default or DEFAULT_ANSWER_MODEL
    if is_custom_answer_model(requested):
        return CUSTOM_QUERY_REWRITE_MODEL

    configured_models = {
        get_chat_model_id(),
        DEFAULT_QUERY_REWRITE_MODEL,
        "gpt-5.4",
        "gpt-5.4-mini",
        "gpt-4.1-mini",
        "gpt-4o",
    }
    extra_models = os.getenv("AZURE_OPENAI_ANSWER_DEPLOYMENTS", "").strip()
    if extra_models:
        configured_models.update(model.strip() for model in extra_models.split(",") if model.strip())

    if requested not in configured_models:
        raise HTTPException(status_code=400, detail=f"Unsupported answer model: {requested}")
    return requested


def get_answer_display_model(payload: ChatRequest) -> str:
    if is_custom_answer_model(payload.answer_model):
        custom_model = (payload.answer_custom_model or "").strip()
        if custom_model:
            return f"{CUSTOM_QUERY_REWRITE_MODEL}:{custom_model}"
        return CUSTOM_QUERY_REWRITE_MODEL
    return get_answer_model_id(payload.answer_model)


def get_answer_custom_config(payload: ChatRequest) -> tuple[str, str, str | None]:
    base_url = (payload.answer_base_url or "").strip()
    model_name = (payload.answer_custom_model or "").strip()
    api_key = (payload.answer_api_key or "").strip() or None

    if not base_url:
        raise HTTPException(status_code=400, detail="Custom answer LLM endpoint is required.")
    if not model_name:
        raise HTTPException(status_code=400, detail="Custom answer LLM model name is required.")

    return (base_url, model_name, api_key)


def log_llm_call(
    *,
    stage: str,
    provider: str,
    endpoint: str,
    model: str,
    max_tokens: int,
    message_count: int,
    timeout_sec: int | None = None,
    extra: dict[str, object] | None = None,
) -> None:
    payload: dict[str, object] = {
        "stage": stage,
        "provider": provider,
        "endpoint": endpoint,
        "model": model,
        "temperature": DEFAULT_CHAT_TEMPERATURE,
        "top_p": DEFAULT_CHAT_TOP_P,
        "max_tokens": max_tokens,
        "message_count": message_count,
    }
    if timeout_sec is not None:
        payload["timeout_sec"] = timeout_sec
    if extra:
        payload.update(extra)
    logger.info("llm_call %s", json.dumps(payload, ensure_ascii=False, sort_keys=True))


def build_azure_openai_chat_completion(
    messages: list[dict[str, str]],
    *,
    deployment: str | None = None,
    max_tokens: int = DEFAULT_CHAT_MAX_TOKENS,
    timeout_sec: int = 120,
) -> str:
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip()
    deployment_id = deployment.strip() if deployment else get_chat_model_id()
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()

    if not endpoint:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_ENDPOINT is not configured.")
    if not api_key:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_API_KEY is not configured.")

    request_url = (
        f"{endpoint}/openai/deployments/{deployment_id}/chat/completions"
        f"?api-version={api_version}"
    )
    token_field_name = "max_completion_tokens" if deployment_id.lower().startswith("gpt-5") else "max_tokens"
    payload = json.dumps(
        {
            "messages": messages,
            "temperature": DEFAULT_CHAT_TEMPERATURE,
            "top_p": DEFAULT_CHAT_TOP_P,
            token_field_name: max_tokens,
        }
    ).encode("utf-8")
    request = urllib_request.Request(
        request_url,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "api-key": api_key,
        },
    )

    try:
        with urllib_request.urlopen(request, timeout=timeout_sec) as response:
            raw_body = response.read()
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Azure OpenAI chat request failed: {detail}") from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI chat request failed: {exc.reason}") from exc
    except (TimeoutError, socket.timeout) as exc:
        raise HTTPException(status_code=504, detail="Azure OpenAI chat request timed out.") from exc

    try:
        response_payload = json.loads(raw_body.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="Azure OpenAI chat response was not valid JSON.") from exc

    choices = response_payload.get("choices")
    if not isinstance(choices, list) or not choices:
        raise HTTPException(status_code=502, detail="Azure OpenAI chat response shape was invalid.")

    first_choice = choices[0]
    if not isinstance(first_choice, dict):
        raise HTTPException(status_code=502, detail="Azure OpenAI chat response shape was invalid.")

    message = first_choice.get("message")
    if not isinstance(message, dict):
        raise HTTPException(status_code=502, detail="Azure OpenAI chat response was missing message data.")

    content = message.get("content")
    if not isinstance(content, str) or not content.strip():
        raise HTTPException(status_code=502, detail="Azure OpenAI chat response was empty.")

    return content.strip()


def build_openai_compatible_chat_completion(
    messages: list[dict[str, str]],
    *,
    base_url: str,
    model_name: str | None = None,
    api_key: str | None = None,
    max_tokens: int = DEFAULT_CHAT_MAX_TOKENS,
    timeout_sec: int = 120,
) -> str:
    normalized_base_url = base_url.strip().rstrip("/")
    if not normalized_base_url:
        raise HTTPException(status_code=400, detail="Custom LLM endpoint is required.")

    if normalized_base_url.endswith("/chat/completions"):
        request_url = normalized_base_url
    else:
        request_url = f"{normalized_base_url}/chat/completions"

    headers = {
        "Content-Type": "application/json",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    request_payload: dict[str, object] = {
        "messages": messages,
        "temperature": DEFAULT_CHAT_TEMPERATURE,
        "top_p": DEFAULT_CHAT_TOP_P,
        "max_tokens": max_tokens,
    }
    if model_name and model_name.strip():
        request_payload["model"] = model_name.strip()

    payload = json.dumps(request_payload).encode("utf-8")
    request = urllib_request.Request(
        request_url,
        data=payload,
        method="POST",
        headers=headers,
    )

    try:
        with urllib_request.urlopen(request, timeout=timeout_sec) as response:
            raw_body = response.read()
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Custom query rewrite request failed: {detail}") from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"Custom query rewrite request failed: {exc.reason}") from exc
    except (TimeoutError, socket.timeout) as exc:
        raise HTTPException(status_code=504, detail="Custom query rewrite request timed out.") from exc

    try:
        response_payload = json.loads(raw_body.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="Custom query rewrite response was not valid JSON.") from exc

    choices = response_payload.get("choices")
    if not isinstance(choices, list) or not choices:
        raise HTTPException(status_code=502, detail="Custom query rewrite response shape was invalid.")

    first_choice = choices[0]
    if not isinstance(first_choice, dict):
        raise HTTPException(status_code=502, detail="Custom query rewrite response shape was invalid.")

    message = first_choice.get("message")
    if not isinstance(message, dict):
        raise HTTPException(status_code=502, detail="Custom query rewrite response was missing message data.")

    content = message.get("content")
    if not isinstance(content, str) or not content.strip():
        raise HTTPException(status_code=502, detail="Custom query rewrite response was empty.")

    return content.strip()


def extract_text_from_content_parts(content_parts: object) -> str:
    if not isinstance(content_parts, list):
        return ""

    fragments: list[str] = []
    for part in content_parts:
        if not isinstance(part, dict):
            continue
        text_value = part.get("text")
        if isinstance(text_value, str) and text_value:
            fragments.append(text_value)

    return "".join(fragments)


def extract_stream_choice_text(choice_payload: dict[str, object]) -> str:
    delta = choice_payload.get("delta")
    if isinstance(delta, dict):
        content = delta.get("content")
        if isinstance(content, str):
            return content
        content_parts = extract_text_from_content_parts(content)
        if content_parts:
            return content_parts

    message = choice_payload.get("message")
    if isinstance(message, dict):
        content = message.get("content")
        if isinstance(content, str):
            return content
        return extract_text_from_content_parts(content)

    return ""


def iter_openai_chat_completion_stream_chunks(response, *, provider_name: str):
    for raw_line in response:
        if not raw_line:
            continue

        decoded_line = raw_line.decode("utf-8", errors="ignore").strip()
        if not decoded_line or not decoded_line.startswith("data:"):
            continue

        payload_text = decoded_line[5:].strip()
        if not payload_text:
            continue
        if payload_text == "[DONE]":
            break

        try:
            payload = json.loads(payload_text)
        except json.JSONDecodeError:
            logger.warning("%s chat stream chunk was not valid JSON.", provider_name)
            continue

        choices = payload.get("choices")
        if not isinstance(choices, list) or not choices:
            continue

        first_choice = choices[0]
        if not isinstance(first_choice, dict):
            continue

        chunk_text = extract_stream_choice_text(first_choice)
        if chunk_text:
            yield chunk_text


def build_azure_openai_chat_completion_stream(
    messages: list[dict[str, str]],
    *,
    deployment: str | None = None,
    max_tokens: int = DEFAULT_CHAT_MAX_TOKENS,
):
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip()
    deployment_id = deployment.strip() if deployment else get_chat_model_id()
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()

    if not endpoint:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_ENDPOINT is not configured.")
    if not api_key:
        raise HTTPException(status_code=500, detail="AZURE_OPENAI_API_KEY is not configured.")

    request_url = (
        f"{endpoint}/openai/deployments/{deployment_id}/chat/completions"
        f"?api-version={api_version}"
    )
    token_field_name = "max_completion_tokens" if deployment_id.lower().startswith("gpt-5") else "max_tokens"
    payload = json.dumps(
        {
            "messages": messages,
            "temperature": DEFAULT_CHAT_TEMPERATURE,
            "top_p": DEFAULT_CHAT_TOP_P,
            token_field_name: max_tokens,
            "stream": True,
            "stream_options": {"include_usage": True},
        }
    ).encode("utf-8")
    request = urllib_request.Request(
        request_url,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "api-key": api_key,
        },
    )

    try:
        response = urllib_request.urlopen(request, timeout=120)
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Azure OpenAI chat stream request failed: {detail}") from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI chat stream request failed: {exc.reason}") from exc
    except (TimeoutError, socket.timeout) as exc:
        raise HTTPException(status_code=504, detail="Azure OpenAI chat stream request timed out.") from exc

    def iterator():
        with response:
            yield from iter_openai_chat_completion_stream_chunks(response, provider_name="Azure OpenAI")

    return iterator()


def build_openai_compatible_chat_completion_stream(
    messages: list[dict[str, str]],
    *,
    base_url: str,
    model_name: str | None = None,
    api_key: str | None = None,
    max_tokens: int = DEFAULT_CHAT_MAX_TOKENS,
):
    normalized_base_url = base_url.strip().rstrip("/")
    if not normalized_base_url:
        raise HTTPException(status_code=400, detail="Custom LLM endpoint is required.")

    if normalized_base_url.endswith("/chat/completions"):
        request_url = normalized_base_url
    else:
        request_url = f"{normalized_base_url}/chat/completions"

    headers = {
        "Content-Type": "application/json",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    request_payload: dict[str, object] = {
        "messages": messages,
        "temperature": DEFAULT_CHAT_TEMPERATURE,
        "top_p": DEFAULT_CHAT_TOP_P,
        "max_tokens": max_tokens,
        "stream": True,
        "stream_options": {"include_usage": True},
    }
    if model_name and model_name.strip():
        request_payload["model"] = model_name.strip()

    payload = json.dumps(request_payload).encode("utf-8")
    request = urllib_request.Request(
        request_url,
        data=payload,
        method="POST",
        headers=headers,
    )

    try:
        response = urllib_request.urlopen(request, timeout=120)
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Custom answer stream request failed: {detail}") from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"Custom answer stream request failed: {exc.reason}") from exc
    except (TimeoutError, socket.timeout) as exc:
        raise HTTPException(status_code=504, detail="Custom answer stream request timed out.") from exc

    def iterator():
        with response:
            yield from iter_openai_chat_completion_stream_chunks(response, provider_name="Custom answer")

    return iterator()


def build_query_rewrite_chat_completion(payload: ChatRequest, messages: list[dict[str, str]]) -> tuple[str, str]:
    model_id = get_query_rewrite_model_id(payload.query_rewrite_model)
    if model_id == CUSTOM_QUERY_REWRITE_MODEL:
        base_url, model_name, api_key = get_query_rewrite_custom_config(payload)
        log_llm_call(
            stage="query_rewrite",
            provider="custom",
            endpoint=base_url.rstrip("/") + ("" if base_url.rstrip("/").endswith("/chat/completions") else "/chat/completions"),
            model=model_name,
            max_tokens=DEFAULT_CHAT_MAX_TOKENS,
            message_count=len(messages),
            timeout_sec=DEFAULT_QUERY_REWRITE_TIMEOUT_SEC,
            extra={"api_key_configured": bool(api_key)},
        )
        return (
            build_openai_compatible_chat_completion(
                messages,
                base_url=base_url,
                model_name=model_name,
                api_key=api_key,
                timeout_sec=DEFAULT_QUERY_REWRITE_TIMEOUT_SEC,
            ),
            get_query_rewrite_display_model(payload),
        )
    azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()
    request_url = f"{azure_endpoint}/openai/deployments/{model_id}/chat/completions?api-version={api_version}" if azure_endpoint else ""
    log_llm_call(
        stage="query_rewrite",
        provider="azure_openai",
        endpoint=request_url,
        model=model_id,
        max_tokens=DEFAULT_CHAT_MAX_TOKENS,
        message_count=len(messages),
        timeout_sec=DEFAULT_QUERY_REWRITE_TIMEOUT_SEC,
        extra={"token_field": "max_completion_tokens" if model_id.lower().startswith("gpt-5") else "max_tokens"},
    )
    return build_azure_openai_chat_completion(
        messages,
        deployment=model_id,
        timeout_sec=DEFAULT_QUERY_REWRITE_TIMEOUT_SEC,
    ), model_id


def build_answer_chat_completion(payload: ChatRequest, messages: list[dict[str, str]]) -> tuple[str, str]:
    model_id = get_answer_model_id(payload.answer_model)
    if model_id == CUSTOM_QUERY_REWRITE_MODEL:
        base_url, model_name, api_key = get_answer_custom_config(payload)
        log_llm_call(
            stage="answer",
            provider="custom",
            endpoint=base_url.rstrip("/") + ("" if base_url.rstrip("/").endswith("/chat/completions") else "/chat/completions"),
            model=model_name,
            max_tokens=DEFAULT_CHAT_MAX_TOKENS,
            message_count=len(messages),
            timeout_sec=120,
            extra={"api_key_configured": bool(api_key)},
        )
        return (
            build_openai_compatible_chat_completion(
                messages,
                base_url=base_url,
                model_name=model_name,
                api_key=api_key,
            ),
            get_answer_display_model(payload),
        )
    azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()
    request_url = f"{azure_endpoint}/openai/deployments/{model_id}/chat/completions?api-version={api_version}" if azure_endpoint else ""
    log_llm_call(
        stage="answer",
        provider="azure_openai",
        endpoint=request_url,
        model=model_id,
        max_tokens=DEFAULT_CHAT_MAX_TOKENS,
        message_count=len(messages),
        timeout_sec=120,
        extra={"token_field": "max_completion_tokens" if model_id.lower().startswith("gpt-5") else "max_tokens"},
    )
    return build_azure_openai_chat_completion(messages, deployment=model_id), model_id


def build_answer_chat_completion_stream(payload: ChatRequest, messages: list[dict[str, str]]):
    model_id = get_answer_model_id(payload.answer_model)
    if model_id == CUSTOM_QUERY_REWRITE_MODEL:
        base_url, model_name, api_key = get_answer_custom_config(payload)
        log_llm_call(
            stage="answer_stream",
            provider="custom",
            endpoint=base_url.rstrip("/") + ("" if base_url.rstrip("/").endswith("/chat/completions") else "/chat/completions"),
            model=model_name,
            max_tokens=DEFAULT_CHAT_MAX_TOKENS,
            message_count=len(messages),
            timeout_sec=120,
            extra={"api_key_configured": bool(api_key), "stream": True},
        )
        return (
            build_openai_compatible_chat_completion_stream(
                messages,
                base_url=base_url,
                model_name=model_name,
                api_key=api_key,
            ),
            get_answer_display_model(payload),
        )
    azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip().rstrip("/")
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", DEFAULT_AZURE_OPENAI_API_VERSION).strip()
    request_url = f"{azure_endpoint}/openai/deployments/{model_id}/chat/completions?api-version={api_version}" if azure_endpoint else ""
    log_llm_call(
        stage="answer_stream",
        provider="azure_openai",
        endpoint=request_url,
        model=model_id,
        max_tokens=DEFAULT_CHAT_MAX_TOKENS,
        message_count=len(messages),
        timeout_sec=120,
        extra={"token_field": "max_completion_tokens" if model_id.lower().startswith("gpt-5") else "max_tokens", "stream": True},
    )
    return build_azure_openai_chat_completion_stream(messages, deployment=model_id), model_id


def build_embeddings(texts: list[str]) -> list[list[float]]:
    provider = get_embedding_provider()
    if provider == EMBEDDING_PROVIDER_HASH:
        return [build_hash_embedding(text) for text in texts]
    if provider == EMBEDDING_PROVIDER_AZURE_OPENAI:
        return build_azure_openai_embeddings(texts)
    raise HTTPException(status_code=500, detail=f"Unsupported embedding provider: {provider}")


def build_chroma_metadata(
    chunk: dict[str, object],
    *,
    stored_name: str,
    original_name: str,
    file_type: str,
) -> dict[str, str | int]:
    metadata: dict[str, str | int] = {
        "stored_name": stored_name,
        "original_name": original_name,
        "file_type": file_type,
        "chunk_index": int(chunk["chunk_index"]),
        "source": str(chunk.get("source") or original_name),
        "text_length": int(chunk["text_length"]),
        "start_char": int(chunk["start_char"]),
        "end_char": int(chunk["end_char"]),
        "preview": str(chunk["preview"]),
    }

    page_number = chunk.get("page_number")
    if isinstance(page_number, int):
        metadata["page_number"] = page_number

    section_header = chunk.get("section_header")
    if isinstance(section_header, str) and section_header.strip():
        metadata["section_header"] = section_header.strip()

    return metadata


def get_chroma_collection():
    ensure_chroma_dir()
    try:
        import chromadb
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=500, detail="chromadb is not installed in the current environment.") from exc

    client = chromadb.PersistentClient(path=str(CHROMA_DIR))
    return client.get_or_create_collection(name=get_embedding_collection_name())


def index_chunks(
    path: Path,
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
    chunk_target_length: int | None = None,
    chunk_overlap_length: int | None = None,
) -> dict[str, object]:
    chunking_result = build_chunking_result(
        path,
        primary_parser=primary_parser,
        fallback_parser=fallback_parser,
        chunk_target_length=chunk_target_length,
        chunk_overlap_length=chunk_overlap_length,
    )
    chunks = chunking_result["chunks"]

    if not isinstance(chunks, list) or not chunks:
        raise HTTPException(status_code=400, detail="No chunks available for indexing.")

    stored_name = str(chunking_result["stored_name"])
    original_name = str(chunking_result["original_name"])
    file_type = str(chunking_result["file_type"])

    documents: list[str] = []
    metadatas: list[dict[str, str | int]] = []
    ids: list[str] = []

    for chunk in chunks:
        if not isinstance(chunk, dict):
            continue

        text = str(chunk["text"])
        documents.append(text)
        metadatas.append(
            build_chroma_metadata(
                chunk,
                stored_name=stored_name,
                original_name=original_name,
                file_type=file_type,
            )
        )
        ids.append(f"{stored_name}:{int(chunk['chunk_index'])}")

    embeddings = build_embeddings(documents)
    collection = get_chroma_collection()
    if ids:
        collection.delete(ids=ids)
        collection.add(ids=ids, documents=documents, embeddings=embeddings, metadatas=metadatas)

    embedding_config = get_embedding_config()

    return {
        "status": "indexed",
        "stored_name": stored_name,
        "original_name": original_name,
        "file_type": file_type,
        "chunk_count": chunking_result["chunk_count"],
        "chunk_target_length": chunking_result["chunk_target_length"],
        "chunk_overlap_length": chunking_result["chunk_overlap_length"],
        "indexed_count": len(ids),
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "embedding_dimension": embedding_config["embedding_dimension"],
        "chroma_path": str(CHROMA_DIR),
        "primary_parser": primary_parser,
        "fallback_parser": fallback_parser,
        "parser_used": chunking_result["parser_used"],
        "fallback_used": chunking_result["fallback_used"],
    }


def build_retrieval_hits(result: dict[str, object]) -> list[dict[str, object]]:
    ids = result.get("ids")
    documents = result.get("documents")
    metadatas = result.get("metadatas")
    distances = result.get("distances")

    if not all(isinstance(item, list) and item for item in [ids, documents, metadatas, distances]):
        return []

    row_ids = ids[0]
    row_documents = documents[0]
    row_metadatas = metadatas[0]
    row_distances = distances[0]

    hits: list[dict[str, object]] = []
    for item_id, document, metadata, distance in zip(row_ids, row_documents, row_metadatas, row_distances):
        if not isinstance(metadata, dict):
            metadata = {}

        hits.append(
            {
                "id": item_id,
                "text": document,
                "contents": document,
                "score": metadata.get("score"),
                "distance": distance,
                "stored_name": metadata.get("stored_name"),
                "original_name": metadata.get("original_name"),
                "document_name": metadata.get("original_name") or metadata.get("source") or metadata.get("stored_name"),
                "source": metadata.get("source"),
                "chunk_index": metadata.get("chunk_index"),
                "text_length": metadata.get("text_length"),
                "start_char": metadata.get("start_char"),
                "end_char": metadata.get("end_char"),
                "page_number": metadata.get("page_number"),
                "section_header": metadata.get("section_header"),
                "header_path": metadata.get("section_header"),
                "preview": metadata.get("preview"),
            }
        )

    return hits


def normalize_retrieval_score(hit: dict[str, object]) -> float | None:
    rrf_score = hit.get("rrf_score")
    if isinstance(rrf_score, (int, float)):
        return float(rrf_score)

    rerank_score = hit.get("rerank_score")
    if isinstance(rerank_score, (int, float)):
        return float(rerank_score)

    raw_score = hit.get("score")
    if isinstance(raw_score, (int, float)):
        return float(raw_score)

    distance = hit.get("distance")
    if isinstance(distance, (int, float)):
        return 1.0 / (1.0 + max(float(distance), 0.0))

    return None


def normalize_product_comparison_text(text: str) -> str:
    normalized = normalize_chat_text(text)
    if not normalized:
        return ""

    normalized = re.sub(r"\.[A-Za-z0-9]+$", "", normalized)
    for token in PRODUCT_DOC_STOPWORDS:
        normalized = normalized.replace(token, " ")
    normalized = re.sub(r"[_\-\[\]\(\)]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return re.sub(r"[^0-9A-Za-z가-힣]+", "", normalized).lower()


def is_strong_product_candidate(text: str) -> bool:
    normalized = normalize_chat_text(text)
    if len(normalized) < 5:
        return False
    return contains_insurance_product_clues(normalized)


def extract_product_name_candidates(*texts: str) -> list[str]:
    raw_candidates: list[str] = []
    for text in texts:
        normalized = normalize_chat_text(text)
        if not normalized:
            continue

        for pattern in (QUOTED_PRODUCT_PATTERN, PRODUCT_QUERY_CANDIDATE_PATTERN, PRODUCT_SUFFIX_CANDIDATE_PATTERN):
            raw_candidates.extend(match.group(1).strip() for match in pattern.finditer(normalized) if match.group(1).strip())

    unique_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in raw_candidates:
        normalized_candidate = normalize_product_comparison_text(candidate)
        if not normalized_candidate or normalized_candidate in seen:
            continue
        if not is_strong_product_candidate(candidate):
            continue
        seen.add(normalized_candidate)
        unique_candidates.append(candidate.strip())
    return unique_candidates


def build_similarity_ratio(left: str, right: str) -> float:
    left_ngrams = build_character_ngrams(left)
    right_ngrams = build_character_ngrams(right)
    if not left_ngrams or not right_ngrams:
        return 0.0
    union = left_ngrams | right_ngrams
    if not union:
        return 0.0
    return len(left_ngrams & right_ngrams) / len(union)


def should_exclude_hit_for_product_mismatch(
    *,
    product_candidates: list[str],
    document_name: str,
) -> bool:
    if not product_candidates:
        return False

    normalized_document_name = normalize_product_comparison_text(document_name)
    if not normalized_document_name or not contains_insurance_product_clues(document_name):
        return False

    for candidate in product_candidates:
        normalized_candidate = normalize_product_comparison_text(candidate)
        if not normalized_candidate:
            continue
        if normalized_candidate in normalized_document_name or normalized_document_name in normalized_candidate:
            return False
        if build_similarity_ratio(normalized_candidate, normalized_document_name) >= 0.2:
            return False

    return True


def get_hit_rrf_score(hit: dict[str, object]) -> float | None:
    rrf_score = hit.get("rrf_score")
    if isinstance(rrf_score, (int, float)):
        return float(rrf_score)

    scores = hit.get("scores")
    if isinstance(scores, dict) and isinstance(scores.get("rrf_score"), (int, float)):
        return float(scores["rrf_score"])

    return UNKNOWN_RRF_SCORE


def sort_hits_for_output(hits: list[dict[str, object]]) -> list[dict[str, object]]:
    indexed_hits = [(index, hit) for index, hit in enumerate(hits) if isinstance(hit, dict)]

    def key(item: tuple[int, dict[str, object]]) -> tuple[int, float, int, float, int, float, int]:
        index, hit = item
        rrf_score = get_hit_rrf_score(hit)
        rerank_score = hit.get("rerank_score")
        score = hit.get("score")
        return (
            0 if isinstance(rrf_score, (int, float)) else 1,
            -float(rrf_score) if isinstance(rrf_score, (int, float)) else 0.0,
            0 if isinstance(rerank_score, (int, float)) else 1,
            -float(rerank_score) if isinstance(rerank_score, (int, float)) else 0.0,
            0 if isinstance(score, (int, float)) else 1,
            -float(score) if isinstance(score, (int, float)) else 0.0,
            index,
        )

    return [hit for _, hit in sorted(indexed_hits, key=key)]


def filter_hits_for_answer_generation(
    *,
    original_query: str,
    rewritten_query: str,
    hits: list[dict[str, object]],
) -> list[dict[str, object]]:
    product_candidates = extract_product_name_candidates(original_query, rewritten_query)
    if not product_candidates:
        return hits

    filtered_hits = [
        hit for hit in hits
        if not should_exclude_hit_for_product_mismatch(
            product_candidates=product_candidates,
            document_name=get_hit_document_name(hit),
        )
    ]
    return filtered_hits


def get_hit_document_name(hit: dict[str, object]) -> str:
    return str(
        hit.get("document_name")
        or hit.get("original_name")
        or hit.get("source")
        or hit.get("stored_name")
        or UNKNOWN_DOCUMENT_NAME
    ).strip() or UNKNOWN_DOCUMENT_NAME


def get_hit_header_path(hit: dict[str, object]) -> str:
    return str(
        hit.get("header_path")
        or hit.get("section_header")
        or hit.get("section")
        or UNKNOWN_HEADER_PATH
    ).strip() or UNKNOWN_HEADER_PATH


def get_hit_contents(hit: dict[str, object]) -> str:
    text = normalize_chat_text(str(hit.get("contents") or hit.get("text") or hit.get("content") or ""))
    return text or EMPTY_CONTENT_PLACEHOLDER


def build_standardized_retrieved_chunks(hits: list[dict[str, object]]) -> list[dict[str, object]]:
    standardized_chunks: list[dict[str, object]] = []
    seen_chunk_ids: set[str] = set()

    for rank, hit in enumerate(hits, start=1):
        chunk_id = str(hit.get("id") or "").strip()
        if not chunk_id or chunk_id in seen_chunk_ids:
            continue
        seen_chunk_ids.add(chunk_id)

        text = get_hit_contents(hit)
        if not text:
            continue

        document_name = get_hit_document_name(hit)
        header_path = get_hit_header_path(hit)
        section = header_path
        document_id = (
            str(hit.get("stored_name") or "").strip()
            or str(hit.get("document_id") or "").strip()
            or str(hit.get("source") or "").strip()
            or str(hit.get("original_name") or "").strip()
            or document_name
            or None
        )
        score = normalize_retrieval_score(hit)

        standardized_chunk: dict[str, object] = {
            "document_id": document_id,
            "chunk_id": chunk_id,
            "score": score,
            "section": section,
            "text": text,
            "document_name": document_name,
            "header_path": header_path,
            "contents": text,
            "rank": rank,
        }

        for source_key in ("source", "original_name", "stored_name", "matched_queries"):
            value = hit.get(source_key)
            if value:
                standardized_chunk[source_key] = value

        rrf_score = get_hit_rrf_score(hit)
        if isinstance(rrf_score, (int, float)):
            standardized_chunk["rrf_score"] = float(rrf_score)

        scores = hit.get("scores")
        if isinstance(scores, dict):
            standardized_chunk["scores"] = dict(scores)

        for numeric_key in ("distance", "rerank_score", "chunk_index", "page_number", "start_char", "end_char"):
            value = hit.get(numeric_key)
            if isinstance(value, (int, float)):
                standardized_chunk[numeric_key] = value

        standardized_chunks.append(standardized_chunk)

    return standardized_chunks


def score_hit(query_tokens: list[str], hit: dict[str, object]) -> float:
    if not query_tokens:
        return 0.0

    query_token_set = set(query_tokens)
    text = str(hit.get("text") or "")
    preview = str(hit.get("preview") or "")
    source = str(hit.get("source") or "")
    section_header = str(hit.get("section_header") or "")
    searchable_text = "\n".join([text, preview, source, section_header])
    hit_tokens = tokenize_text(searchable_text)
    if not hit_tokens:
        return 0.0

    hit_frequency = build_token_frequency(hit_tokens)
    overlap_count = sum(hit_frequency.get(token, 0) for token in query_tokens)
    unique_overlap = len(query_token_set & set(hit_tokens))
    query_coverage = unique_overlap / len(query_token_set)
    query_ngram_set = build_character_ngrams(" ".join(query_tokens))
    hit_ngram_set = build_character_ngrams(searchable_text)
    ngram_overlap = len(query_ngram_set & hit_ngram_set)
    ngram_coverage = (ngram_overlap / len(query_ngram_set)) if query_ngram_set else 0.0
    distance = float(hit.get("distance") or 0.0)

    return (query_coverage * 2.5) + (overlap_count * 0.3) + (ngram_coverage * 3.5) - distance


def rerank_hits(query: str, hits: list[dict[str, object]], top_k: int) -> list[dict[str, object]]:
    query_tokens = tokenize_text(query)
    if not query_tokens or not hits:
        return hits[:top_k]

    scored_hits: list[dict[str, object]] = []
    for hit in hits:
        rescored_hit = dict(hit)
        rescored_hit["rerank_score"] = score_hit(query_tokens, hit)
        scored_hits.append(rescored_hit)

    scored_hits.sort(
        key=lambda item: (
            float(item.get("rerank_score") or 0.0),
            -float(item.get("distance") or 0.0),
        ),
        reverse=True,
    )
    return scored_hits[:top_k]


def should_use_lexical_fallback(exc: HTTPException) -> bool:
    detail = str(exc.detail or "")
    return "Azure OpenAI embedding request failed" in detail


def build_lexical_fallback_hits(
    *,
    queries: list[str],
    top_k: int,
    stored_name: str | None,
    rerank_query: str,
) -> dict[str, object]:
    ensure_upload_dir()
    filtered_queries = [query.strip() for query in queries if query.strip()]
    if not filtered_queries:
        raise HTTPException(status_code=400, detail="Query text is required.")

    candidate_paths: list[Path] = []
    if stored_name:
        path = UPLOAD_DIR / stored_name
        if path.is_file():
            candidate_paths.append(path)
    else:
        for path in sorted(UPLOAD_DIR.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True):
            if path.is_file():
                candidate_paths.append(path)

    aggregated_hits: dict[str, dict[str, object]] = {}
    executed_queries: list[str] = []
    for path in candidate_paths:
        try:
            chunking_result = build_chunking_result(
                path,
                primary_parser=PRIMARY_PARSER_LEGACY_AUTO,
                fallback_parser=FALLBACK_PARSER_EXTENSION_DEFAULT,
            )
        except HTTPException as exc:
            logger.warning("lexical_fallback_chunking_failed stored_name=%s detail=%s", path.name, exc.detail)
            continue

        chunks = chunking_result.get("chunks")
        if not isinstance(chunks, list):
            continue

        for query in filtered_queries:
            executed_queries.append(query)
            query_tokens = tokenize_text(query)
            if not query_tokens:
                continue

            for chunk in chunks:
                if not isinstance(chunk, dict):
                    continue
                hit = {
                    "id": f"{path.name}:{int(chunk.get('chunk_index', 0))}",
                    "text": str(chunk.get("text") or ""),
                    "contents": str(chunk.get("text") or ""),
                    "distance": 0.0,
                    "stored_name": path.name,
                    "original_name": chunking_result.get("original_name"),
                    "document_name": chunking_result.get("original_name") or path.name,
                    "source": chunk.get("source") or chunking_result.get("original_name"),
                    "chunk_index": chunk.get("chunk_index"),
                    "text_length": chunk.get("text_length"),
                    "start_char": chunk.get("start_char"),
                    "end_char": chunk.get("end_char"),
                    "page_number": chunk.get("page_number"),
                    "section_header": chunk.get("section_header"),
                    "header_path": chunk.get("section_header"),
                    "preview": chunk.get("preview"),
                }
                score = score_hit(query_tokens, hit)
                if score <= 0:
                    continue

                existing_hit = aggregated_hits.get(hit["id"])
                if existing_hit is None or float(existing_hit.get("rerank_score") or 0.0) < score:
                    hit["rerank_score"] = score
                    hit["matched_queries"] = [query]
                    aggregated_hits[hit["id"]] = hit
                    continue

                matched_queries = existing_hit.get("matched_queries")
                if not isinstance(matched_queries, list):
                    matched_queries = []
                if query not in matched_queries:
                    matched_queries.append(query)
                existing_hit["matched_queries"] = matched_queries

    hits = sort_hits_for_output(rerank_hits(rerank_query.strip() or filtered_queries[0], list(aggregated_hits.values()), top_k))
    return {
        "status": "retrieved",
        "query": rerank_query.strip() or filtered_queries[0],
        "executed_queries": filtered_queries,
        "top_k": top_k,
        "hit_count": len(hits),
        "retrieved_chunks": build_standardized_retrieved_chunks(hits),
        "collection_name": "local-lexical-fallback",
        "embedding_provider": "local-lexical-fallback",
        "embedding_model": "none",
        "hits": hits,
    }


def retrieve_chunks(payload: RetrieveRequest) -> dict[str, object]:
    query = payload.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query text is required.")

    top_k = max(1, min(payload.top_k, 20))
    collection = get_chroma_collection()

    where = None
    if payload.stored_name:
        where = {"stored_name": payload.stored_name}

    candidate_count = min(max(top_k * RETRIEVAL_CANDIDATE_MULTIPLIER, top_k), RETRIEVAL_MAX_CANDIDATES)
    embedding_config = get_embedding_config()

    try:
        query_embedding = build_embeddings([query])[0]
    except HTTPException as exc:
        if not should_use_lexical_fallback(exc):
            raise
        logger.warning("retrieve_embedding_failed query=%s detail=%s", query, exc.detail)
        return build_lexical_fallback_hits(
            queries=[query],
            top_k=top_k,
            stored_name=payload.stored_name,
            rerank_query=query,
        )

    result = collection.query(
        query_embeddings=[query_embedding],
        n_results=candidate_count,
        where=where,
        include=["documents", "metadatas", "distances"],
    )
    hits = sort_hits_for_output(rerank_hits(query, build_retrieval_hits(result), top_k))
    retrieved_chunks = build_standardized_retrieved_chunks(hits)

    return {
        "status": "retrieved",
        "query": query,
        "top_k": top_k,
        "hit_count": len(hits),
        "retrieved_chunks": retrieved_chunks,
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "hits": hits,
    }


def build_query_candidates_for_chat(payload: ChatRequest, rewrite_result: RewriteResult) -> list[str]:
    document_hint = (rewrite_result.routing_hints.get("document_hint") or "").strip().lower()
    question_type = (rewrite_result.question_type or "").strip().lower()
    routing_candidates = (
        get_document_hint_expansions(document_hint)
        + get_question_type_expansions(question_type)
    )

    candidates = [
        rewrite_result.rewritten_query.strip(),
        *[query.strip() for query in rewrite_result.search_queries],
        *routing_candidates,
    ]
    unique_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        normalized_candidate = candidate.strip()
        if not normalized_candidate or normalized_candidate in seen:
            continue
        seen.add(normalized_candidate)
        unique_candidates.append(normalized_candidate)
    return unique_candidates


def retrieve_chunks_for_queries(
    *,
    queries: list[str],
    top_k: int,
    stored_name: str | None,
    rerank_query: str,
) -> dict[str, object]:
    filtered_queries = [query.strip() for query in queries if query.strip()]
    if not filtered_queries:
        raise HTTPException(status_code=400, detail="Query text is required.")

    top_k = max(1, min(top_k, 20))
    collection = get_chroma_collection()
    embedding_config = get_embedding_config()
    where = {"stored_name": stored_name} if stored_name else None
    candidate_count = min(max(top_k * RETRIEVAL_CANDIDATE_MULTIPLIER, top_k), RETRIEVAL_MAX_CANDIDATES)

    aggregated_hits: dict[str, dict[str, object]] = {}
    executed_queries: list[str] = []
    for query in filtered_queries:
        try:
            query_embedding = build_embeddings([query])[0]
        except HTTPException as exc:
            if not should_use_lexical_fallback(exc):
                raise
            logger.warning("retrieve_multi_embedding_failed query=%s detail=%s", query, exc.detail)
            return build_lexical_fallback_hits(
                queries=filtered_queries,
                top_k=top_k,
                stored_name=stored_name,
                rerank_query=rerank_query,
            )
        result = collection.query(
            query_embeddings=[query_embedding],
            n_results=candidate_count,
            where=where,
            include=["documents", "metadatas", "distances"],
        )
        executed_queries.append(query)

        for hit in build_retrieval_hits(result):
            hit_id = str(hit.get("id") or "")
            if not hit_id:
                continue

            existing_hit = aggregated_hits.get(hit_id)
            hit_copy = dict(hit)
            hit_copy["matched_queries"] = [query]
            if existing_hit is None:
                aggregated_hits[hit_id] = hit_copy
                continue

            matched_queries = existing_hit.get("matched_queries")
            if not isinstance(matched_queries, list):
                matched_queries = []
            if query not in matched_queries:
                matched_queries.append(query)
            existing_hit["matched_queries"] = matched_queries

            existing_distance = float(existing_hit.get("distance") or 0.0)
            new_distance = float(hit_copy.get("distance") or 0.0)
            if new_distance < existing_distance:
                hit_copy["matched_queries"] = matched_queries
                aggregated_hits[hit_id] = hit_copy

    hits = sort_hits_for_output(rerank_hits(rerank_query.strip() or filtered_queries[0], list(aggregated_hits.values()), top_k))
    return {
        "status": "retrieved",
        "query": rerank_query.strip() or filtered_queries[0],
        "executed_queries": executed_queries,
        "top_k": top_k,
        "hit_count": len(hits),
        "retrieved_chunks": build_standardized_retrieved_chunks(hits),
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "hits": hits,
    }


def build_chat_citations(hits: list[dict[str, object]]) -> list[dict[str, object]]:
    citations: list[dict[str, object]] = []
    for index, hit in enumerate(hits, start=1):
        citations.append(
            {
                "id": hit.get("id"),
                "rank": index,
                "source": hit.get("source") or hit.get("original_name"),
                "original_name": hit.get("original_name"),
                "stored_name": hit.get("stored_name"),
                "chunk_index": hit.get("chunk_index"),
                "page_number": hit.get("page_number"),
                "section_header": hit.get("section_header"),
            }
        )
    return citations


def format_chat_context(hits: list[dict[str, object]]) -> str:
    sections: list[str] = []
    for index, hit in enumerate(hits, start=1):
        document_name = get_hit_document_name(hit)
        header_path = get_hit_header_path(hit)
        text = get_hit_contents(hit)
        sections.append(
            f"[Context {index}]\n"
            f"document_name={document_name}\n"
            f"header_path={header_path}\n"
            f"content={text}"
        )
    return "\n\n".join(sections)


def parse_chat_completion(content: str) -> tuple[bool, str]:
    status_match = re.search(r"STATUS:\s*(grounded|insufficient)", content, flags=re.IGNORECASE)
    answer_match = re.search(r"ANSWER:\s*(.*)", content, flags=re.IGNORECASE | re.DOTALL)

    if not status_match or not answer_match:
        return False, content.strip()

    status = status_match.group(1).strip().lower()
    answer = answer_match.group(1).strip()
    return status == "insufficient", answer


def extract_streaming_answer_preview(content: str) -> str:
    answer_match = re.search(r"ANSWER:\s*(.*)", content, flags=re.IGNORECASE | re.DOTALL)
    if answer_match:
        return answer_match.group(1)

    stripped_content = content.lstrip()
    if not stripped_content:
        return ""

    # Keep STATUS/ANSWER scaffolding out of incremental UI updates.
    if stripped_content.upper().startswith("STATUS") or stripped_content.upper().startswith("ANSWER"):
        return ""

    return content


def extract_json_object(text: str) -> dict[str, object]:
    trimmed_text = text.strip()
    if not trimmed_text:
        raise ValueError("JSON payload is empty.")

    if trimmed_text.startswith("```"):
        fenced_match = re.search(r"```(?:json)?\s*(\{.*\})\s*```", trimmed_text, re.DOTALL)
        if fenced_match:
            trimmed_text = fenced_match.group(1).strip()

    start = trimmed_text.find("{")
    end = trimmed_text.rfind("}")
    if start == -1 or end == -1 or start >= end:
        raise ValueError("JSON object not found.")

    payload = json.loads(trimmed_text[start : end + 1])
    if not isinstance(payload, dict):
        raise ValueError("JSON payload must be an object.")
    return payload


def serialize_chat_metadata(metadata: ChatMetadata | None) -> dict[str, object]:
    if metadata is None:
        return {}
    if hasattr(metadata, "model_dump"):
        return metadata.model_dump(exclude_none=True)
    return metadata.dict(exclude_none=True)


def infer_question_type(query: str) -> str | None:
    return infer_question_type_from_rules(query)


def infer_document_hint(
    query: str,
    metadata: ChatMetadata | None,
    entities: dict[str, str],
    routing_hints: dict[str, str],
) -> str | None:
    combined_text = " ".join(
        [
            query,
            " ".join(f"{key}:{value}" for key, value in entities.items()),
            " ".join(f"{key}:{value}" for key, value in routing_hints.items()),
            metadata.document_type if metadata and metadata.document_type else "",
            metadata.product if metadata and metadata.product else "",
        ]
    )
    return infer_document_hint_from_rules(combined_text)


def infer_document_type_filters(
    payload: ChatRequest,
    *,
    entities: dict[str, str],
    routing_hints: dict[str, str],
    question_type: str | None,
    document_hint: str | None,
) -> list[str]:
    raw_candidates: list[str] = [
        payload.metadata.document_type if payload.metadata and payload.metadata.document_type else "",
        entities.get("document_type", ""),
        routing_hints.get("document_type", ""),
        routing_hints.get("preferred_document_type", ""),
    ]

    mapped_document_type = DOCUMENT_HINT_TO_DOCUMENT_TYPE.get(document_hint or "")
    if mapped_document_type:
        raw_candidates.append(mapped_document_type)

    if question_type in {"statistics", "numeric"}:
        raw_candidates.append("statistics_table")

    normalized_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in raw_candidates:
        normalized = normalize_search_document_type(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        normalized_candidates.append(normalized)
    return normalized_candidates


def build_query_rewrite_seed_query(trimmed_query: str, last_customer_message: str | None, normalized_conversation: list[dict[str, str]] | None = None) -> str:
    if normalized_conversation:
        focused_customer_text = build_customer_rewrite_focus_text(normalized_conversation, last_customer_message)
        if focused_customer_text:
            return focused_customer_text
    if last_customer_message and last_customer_message.strip():
        return strip_conversational_fillers(last_customer_message.strip()) or last_customer_message.strip()
    return trimmed_query


def build_query_rewrite_messages(
    *,
    seed_query: str,
    original_query: str,
    last_customer_message: str | None,
    normalized_conversation: list[dict[str, str]],
    metadata: dict[str, object],
) -> list[dict[str, str]]:
    context_lines = [f"{turn['role']}: {turn['content']}" for turn in normalized_conversation[-6:]]
    domain_focus = infer_domain_specific_focus(normalized_conversation)
    fallback_system_prompt = (
        "You rewrite financial customer-service conversations into one retrieval-friendly standalone Korean question.\n"
        "Preserve the original meaning and the user's actual search intent.\n"
        "Base the rewrite on the latest meaningful customer utterance.\n"
        "If the original conversation or question is in Korean, respond in Korean.\n"
        "Remove greetings, counselor guidance, and identity-verification details.\n"
        "If the latest customer question is statistical or numeric, keep it statistical or numeric.\n"
        "Preserve the measured target, reference year, and metric such as 평균, 비율, 건수, 금액, 진료비, 수술비, 인당.\n"
        "Do not rewrite a statistical question into coverage, claimability, terms, or exemption questions.\n"
        "Include product or service names only when they are truly required for retrieval.\n"
        "Do not include personal information.\n"
        "Output one complete standalone question sentence for rewritten_query.\n"
    )
    response_contract = (
        "Return JSON only.\n"
        "Do not output analysis, markdown, or code fences.\n"
        "The JSON schema is:\n"
        "{\n"
        '  "rewritten_query": string,\n'
        '  "search_queries": string[],\n'
        '  "intent": string,\n'
        '  "question_type": string,\n'
        '  "entities": object,\n'
        '  "routing_hints": object\n'
        "}"
    )
    system_prompt_base = load_query_rewrite_system_prompt() or fallback_system_prompt
    system_prompt = f"{system_prompt_base.strip()}\n\n{response_contract}"
    user_prompt = (
        "Rewrite the customer question into retrieval-friendly search queries.\n"
        "Rules:\n"
        "- Use the latest meaningful customer message as the primary rewrite target.\n"
        "- If the latest customer message is vague, referential, full of conversational fillers, or only contains identity verification details, use earlier customer turns to restore the full question.\n"
        "- Do not copy conversational fillers such as '그러니까요', '예를 들면', '네', or unfinished spoken fragments into rewritten_query.\n"
        "- Exclude greetings, counselor guidance, name, birth date, resident number, and other identity verification details.\n"
        "- Keep meaningful policy attributes such as 가입 시점, 보험 종류, 계약 형태 when they help retrieval.\n"
        "- If the latest customer question is statistical or numeric, remove earlier coverage, underwriting, greeting, and identity-verification context unless it is essential to the statistic itself.\n"
        "- Preserve important entities, scenario details, reference years, and numeric/statistical expressions.\n"
        "- Keep keywords such as 평균, 통계, 기준, 연도, 인당, 비율, 건수, 금액, 진료비, 수술비, 발생 건수 when they appear in the customer's real question.\n"
        "- Do not convert a statistical or numeric question into a coverage, claimability, terms, or exemption question.\n"
        "- Include product or service names only when they are actually core retrieval keys. Do not force them into statistical queries when unnecessary.\n"
        "- Infer the most likely product or document name from the context when omitted.\n"
        "- Exclude topics that the agent has already answered.\n"
        "- question_type should be one of statistics, numeric, documents, period, definition, conditions, comparison, or general when possible.\n"
        "- For statistical queries, fill entities such as year, metric, target, procedure, or topic when possible.\n"
        "- For statistical queries, routing_hints should prefer statistics_table or statistics_report.\n"
        "- When the target document family is clear, include routing_hints.document_type using one of policy, calculation_guide, business_guide, statistics_table.\n"
        "- For policy or terms questions, routing_hints may include terms, pricing_method, change_process, or general.\n"
        "- Return valid JSON only with no extra text.\n\n"
        f"Input:\n{json.dumps({'seed_query': seed_query, 'original_query': original_query, 'last_customer_message': last_customer_message, 'conversation_context': context_lines, 'metadata': metadata, 'domain_focus_hint': domain_focus}, ensure_ascii=False)}"
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def build_standalone_query_rewrite_messages(
    *,
    seed_query: str,
    original_query: str,
    metadata: dict[str, object],
) -> list[dict[str, str]]:
    fallback_system_prompt = (
        "You rewrite one Korean user question into one concise retrieval-friendly Korean question.\n"
        "Keep the original intent.\n"
        "If the original question is in Korean, respond in Korean.\n"
        "If the question is statistical or numeric, preserve the target, year, and metric.\n"
        "Do not convert statistical questions into coverage, claimability, terms, or exemption questions.\n"
        "Do not include personal information.\n"
        "rewritten_query must be one complete question sentence.\n"
    )
    response_contract = (
        "Return JSON only.\n"
        "Do not output reasoning, markdown, or code fences.\n"
        "The JSON schema is:\n"
        "{\n"
        '  "rewritten_query": string,\n'
        '  "search_queries": string[],\n'
        '  "intent": string,\n'
        '  "question_type": string,\n'
        '  "entities": object,\n'
        '  "routing_hints": object\n'
        "}"
    )
    system_prompt_base = load_query_rewrite_system_prompt() or fallback_system_prompt
    system_prompt = f"{system_prompt_base.strip()}\n\n{response_contract}"
    user_prompt = (
        "Rewrite this single user question for retrieval.\n"
        "Rules:\n"
        "- Keep product, app, service, document, and task names only when they are core retrieval keys.\n"
        "- Remove conversational phrasing only when unnecessary.\n"
        "- Preserve reference years, numbers, and keywords such as 평균, 통계, 기준, 연도, 인당, 비율, 건수, 금액, 진료비, 수술비 when they appear in the original question.\n"
        "- Do not rewrite a statistical or numeric question into a coverage, claimability, terms, or exemption question.\n"
        "- question_type should be one of statistics, numeric, documents, period, definition, conditions, comparison, or general when possible.\n"
        "- For statistical queries, fill entities such as year, metric, target, procedure, or topic when possible.\n"
        "- For statistical queries, routing_hints should prefer statistics_table or statistics_report.\n"
        "- When the target document family is clear, include routing_hints.document_type using one of policy, calculation_guide, business_guide, statistics_table.\n"
        "- Keep it short and retrieval-friendly.\n"
        "- Return valid JSON only.\n\n"
        f"Input:\n{json.dumps({'seed_query': seed_query, 'original_query': original_query, 'metadata': metadata}, ensure_ascii=False)}"
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def parse_query_rewrite_response(response_payload: dict[str, object], fallback_query: str) -> tuple[str, list[str], str | None, str | None, dict[str, str], dict[str, str]]:
    rewritten_query = ensure_question_sentence(str(response_payload.get("rewritten_query") or "").strip() or fallback_query)
    raw_search_queries = response_payload.get("search_queries")
    search_queries: list[str] = []
    if isinstance(raw_search_queries, list):
        for item in raw_search_queries:
            if isinstance(item, str) and item.strip():
                search_queries.append(item.strip())
    if rewritten_query not in search_queries:
        search_queries.insert(0, rewritten_query)

    raw_entities = response_payload.get("entities")
    entities = {str(key): str(value) for key, value in raw_entities.items()} if isinstance(raw_entities, dict) else {}
    raw_routing_hints = response_payload.get("routing_hints")
    routing_hints = (
        {str(key): str(value) for key, value in raw_routing_hints.items()}
        if isinstance(raw_routing_hints, dict)
        else {}
    )
    intent = response_payload.get("intent")
    if intent is not None:
        intent = str(intent).strip() or None
    question_type = response_payload.get("question_type")
    if question_type is not None:
        question_type = str(question_type).strip() or None

    return rewritten_query, search_queries or [fallback_query], intent, question_type, entities, routing_hints


def build_query_rewrite_fallback_result(
    payload: ChatRequest,
    *,
    trimmed_query: str,
    normalized_conversation: list[dict[str, str]],
    last_customer_message: str | None,
) -> RewriteResult:
    fallback_query = build_query_rewrite_seed_query(trimmed_query, last_customer_message, normalized_conversation)
    return enrich_rewrite_result(
        payload,
        RewriteResult(
            original_query=trimmed_query,
            rewritten_query=ensure_question_sentence(fallback_query),
            search_queries=[fallback_query, trimmed_query],
            query_rewrite_model=get_query_rewrite_display_model(payload),
            normalized_conversation=normalized_conversation,
            last_customer_message=last_customer_message,
            rewrite_source="fallback:last_customer_message",
        ),
    )


def retry_query_rewrite_once(
    payload: ChatRequest,
    *,
    seed_query: str,
    trimmed_query: str,
    last_customer_message: str | None,
    normalized_conversation: list[dict[str, str]],
    metadata: dict[str, object],
    validation_reasons: list[str],
) -> tuple[str, list[str], str | None, str | None, dict[str, str], dict[str, str]]:
    messages = build_query_rewrite_messages(
        seed_query=seed_query,
        original_query=trimmed_query,
        last_customer_message=last_customer_message,
        normalized_conversation=normalized_conversation,
        metadata=metadata,
    )
    retry_note = (
        "The previous rewritten_query failed validation.\n"
        f"Validation failures: {', '.join(validation_reasons)}.\n"
        "Return one Korean standalone question sentence between 15 and 150 characters.\n"
        "Preserve the original intent, year, and critical keywords from the customer message.\n"
        "If the source question is statistical or numeric, keep it statistical or numeric.\n"
        "Do not convert it into coverage, claimability, terms, or exemption intent.\n"
        "Keep keywords such as 평균, 통계, 기준, 연도, 인당, 비율, 건수, 금액, 진료비, 수술비 when present.\n"
        "Return JSON only."
    )
    messages.append({"role": "user", "content": retry_note})
    response_text, _ = build_query_rewrite_chat_completion(payload, messages)
    response_payload = extract_json_object(response_text)
    return parse_query_rewrite_response(response_payload, seed_query)


def apply_standalone_query_validation(
    payload: ChatRequest,
    rewrite_result: RewriteResult,
) -> RewriteResult:
    source_texts = [
        rewrite_result.last_customer_message or "",
        payload.query,
        payload.metadata.product if payload.metadata and payload.metadata.product else "",
        payload.metadata.document_type if payload.metadata and payload.metadata.document_type else "",
    ]

    validated_query, validation_reasons = validate_standalone_search_query(
        rewrite_result.rewritten_query,
        source_texts=source_texts,
    )
    if not validation_reasons:
        rewrite_result.rewritten_query = validated_query
        rewrite_result.validation_reasons = []
        rewrite_result.rewrite_source = rewrite_result.rewrite_source or "llm"
        return rewrite_result

    promoted_query, promoted_reasons, promoted_source = choose_best_rewrite_candidate(
        rewrite_result.search_queries,
        source_texts=source_texts,
    )
    if promoted_source and not promoted_reasons and promoted_query != validated_query:
        rewrite_result.rewritten_query = promoted_query
        rewrite_result.search_queries = [promoted_query, *rewrite_result.search_queries]
        rewrite_result.validation_reasons = validation_reasons
        rewrite_result.rewrite_source = "promoted:search_query"
        return rewrite_result

    fallback_query = build_query_rewrite_seed_query(
        rewrite_result.original_query,
        rewrite_result.last_customer_message,
        rewrite_result.normalized_conversation,
    )
    validated_fallback_query, fallback_reasons = validate_standalone_search_query(
        fallback_query,
        source_texts=source_texts,
    )
    if not fallback_reasons:
        rewrite_result.rewritten_query = validated_fallback_query
        rewrite_result.search_queries = [validated_fallback_query, *rewrite_result.search_queries]
        rewrite_result.validation_reasons = validation_reasons
        rewrite_result.rewrite_source = "fallback:last_customer_message"
        return rewrite_result

    metadata_augmented_query = build_metadata_augmented_query(rewrite_result.last_customer_message, payload.metadata)
    if metadata_augmented_query:
        validated_augmented_query, augmented_reasons = validate_standalone_search_query(
            metadata_augmented_query,
            source_texts=source_texts,
        )
        if not augmented_reasons:
            rewrite_result.rewritten_query = validated_augmented_query
            rewrite_result.search_queries = [validated_augmented_query, *rewrite_result.search_queries]
            rewrite_result.validation_reasons = validation_reasons + fallback_reasons
            rewrite_result.rewrite_source = "fallback:last_customer_message+metadata"
            return rewrite_result

    metadata = serialize_chat_metadata(payload.metadata)
    try:
        (
            retried_query,
            retried_search_queries,
            retried_intent,
            retried_question_type,
            retried_entities,
            retried_routing_hints,
        ) = retry_query_rewrite_once(
            payload,
            seed_query=fallback_query,
            trimmed_query=rewrite_result.original_query,
            last_customer_message=rewrite_result.last_customer_message,
            normalized_conversation=rewrite_result.normalized_conversation,
            metadata=metadata,
            validation_reasons=validation_reasons + fallback_reasons,
        )
        validated_retry_query, retry_reasons = validate_standalone_search_query(
            retried_query,
            source_texts=source_texts,
        )
        if not retry_reasons:
            rewrite_result.rewritten_query = validated_retry_query
            rewrite_result.search_queries = [validated_retry_query, *retried_search_queries, *rewrite_result.search_queries]
            rewrite_result.intent = retried_intent or rewrite_result.intent
            rewrite_result.question_type = retried_question_type or rewrite_result.question_type
            rewrite_result.entities = retried_entities or rewrite_result.entities
            rewrite_result.routing_hints = retried_routing_hints or rewrite_result.routing_hints
            rewrite_result.validation_reasons = validation_reasons + fallback_reasons
            rewrite_result.rewrite_source = "retry:llm"
            return rewrite_result
    except (HTTPException, ValueError, json.JSONDecodeError) as exc:
        logger.warning("query_rewrite_retry_failed original_query=%s error=%s", rewrite_result.original_query, exc)

    rewrite_result.rewritten_query = validated_fallback_query or ensure_question_sentence(fallback_query)
    rewrite_result.search_queries = [rewrite_result.rewritten_query, *rewrite_result.search_queries]
    rewrite_result.validation_reasons = validation_reasons + fallback_reasons
    rewrite_result.rewrite_source = "fallback:last_customer_message"
    return rewrite_result


def enrich_rewrite_result(payload: ChatRequest, rewrite_result: RewriteResult) -> RewriteResult:
    entities = dict(rewrite_result.entities)
    routing_hints = dict(rewrite_result.routing_hints)
    search_queries = [query.strip() for query in rewrite_result.search_queries if query.strip()]
    enrichment_source = normalize_chat_text(
        " ".join(
            [
                rewrite_result.rewritten_query,
                rewrite_result.last_customer_message or "",
                payload.query,
            ]
        )
    )
    inferred_question_type = infer_question_type(payload.query)
    if is_statistics_or_numeric_query(enrichment_source):
        question_type = "statistics" if "통계" in enrichment_source or "자료" in enrichment_source else "numeric"
    else:
        question_type = rewrite_result.question_type or inferred_question_type

    statistics_entities = extract_statistics_entities(enrichment_source)
    for key, value in statistics_entities.items():
        entities.setdefault(key, value)

    if question_type and "question_type" not in entities:
        entities["question_type"] = question_type

    document_hint = infer_document_hint(payload.query, payload.metadata, entities, routing_hints)
    if question_type in {"statistics", "numeric"} and not document_hint:
        document_hint = STATISTICS_DOCUMENT_HINTS[1]
        routing_hints.setdefault("preferred_document_hint", STATISTICS_DOCUMENT_HINTS[0])
        routing_hints.setdefault("chunk_types", ",".join(STATISTICS_ROUTING_CHUNK_TYPES))
        if entities.get("year"):
            routing_hints.setdefault("year", entities["year"])

    if document_hint:
        routing_hints["document_hint"] = document_hint

    document_type_filters = infer_document_type_filters(
        payload,
        entities=entities,
        routing_hints=routing_hints,
        question_type=question_type,
        document_hint=document_hint,
    )
    if document_type_filters:
        routing_hints["document_type"] = ",".join(document_type_filters)

    search_queries.extend(get_document_hint_expansions(document_hint))
    if question_type == "statistics":
        search_queries.extend(get_statistics_query_expansions(rewrite_result.rewritten_query, entities))
    else:
        search_queries.extend(get_question_type_expansions(question_type))

    unique_queries: list[str] = []
    seen: set[str] = set()
    for candidate in [rewrite_result.rewritten_query, *search_queries]:
        normalized = candidate.strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique_queries.append(normalized)

    return RewriteResult(
        original_query=rewrite_result.original_query,
        rewritten_query=rewrite_result.rewritten_query,
        search_queries=unique_queries or [rewrite_result.rewritten_query],
        document_type_filters=document_type_filters,
        query_rewrite_model=rewrite_result.query_rewrite_model,
        intent=rewrite_result.intent,
        question_type=question_type,
        entities=entities,
        routing_hints=routing_hints,
        normalized_conversation=rewrite_result.normalized_conversation,
        last_customer_message=rewrite_result.last_customer_message,
        validation_reasons=rewrite_result.validation_reasons,
        rewrite_source=rewrite_result.rewrite_source,
    )


# Normalize user input into retrieval-ready signals for the chat-side Hybrid RAG flow.
def rewrite_chat_query(payload: ChatRequest) -> RewriteResult:
    trimmed_query = payload.query.strip()
    if not trimmed_query:
        raise HTTPException(status_code=400, detail="Query text is required.")

    normalized_conversation = normalize_conversation_context(payload.conversation_context)
    if not normalized_conversation:
        normalized_conversation = parse_conversation_context_from_query_text(trimmed_query)
    metadata = serialize_chat_metadata(payload.metadata)
    if is_standalone_query_rewrite(normalized_conversation):
        last_customer_message = None
        seed_query = normalize_direct_user_question(trimmed_query) or trimmed_query
        messages = build_standalone_query_rewrite_messages(
            seed_query=seed_query,
            original_query=trimmed_query,
            metadata=metadata,
        )
    else:
        last_customer_message = extract_last_meaningful_customer_message(normalized_conversation)
        seed_query = build_query_rewrite_seed_query(trimmed_query, last_customer_message, normalized_conversation)
        messages = build_query_rewrite_messages(
            seed_query=seed_query,
            original_query=trimmed_query,
            last_customer_message=last_customer_message,
            normalized_conversation=normalized_conversation,
            metadata=metadata,
        )

    try:
        response_text, query_rewrite_model = build_query_rewrite_chat_completion(payload, messages)
        try:
            response_payload = extract_json_object(response_text)
        except (ValueError, json.JSONDecodeError) as exc:
            logger.warning(
                "query_rewrite_invalid_response original_query=%s model=%s preview=%s error=%s",
                trimmed_query,
                query_rewrite_model,
                summarize_llm_response_preview(response_text),
                exc,
            )
            raise
        rewritten_query, search_queries, intent, question_type, entities, routing_hints = parse_query_rewrite_response(
            response_payload,
            seed_query,
        )

        rewrite_result = RewriteResult(
            original_query=trimmed_query,
            rewritten_query=rewritten_query or seed_query,
            search_queries=search_queries,
            query_rewrite_model=query_rewrite_model,
            intent=intent,
            question_type=question_type,
            entities=entities,
            routing_hints=routing_hints,
            normalized_conversation=normalized_conversation,
            last_customer_message=last_customer_message,
            rewrite_source="llm",
        )
        rewrite_result = apply_standalone_query_validation(payload, rewrite_result)
        return enrich_rewrite_result(
            payload,
            rewrite_result,
        )
    except HTTPException as exc:
        if exc.status_code == 400:
            raise
        logger.warning("query_rewrite_failed original_query=%s error=%s", trimmed_query, exc)
        return build_query_rewrite_fallback_result(
            payload,
            trimmed_query=trimmed_query,
            normalized_conversation=normalized_conversation,
            last_customer_message=last_customer_message,
        )
    except (ValueError, json.JSONDecodeError) as exc:
        logger.warning("query_rewrite_failed original_query=%s error=%s", trimmed_query, exc)
        return build_query_rewrite_fallback_result(
            payload,
            trimmed_query=trimmed_query,
            normalized_conversation=normalized_conversation,
            last_customer_message=last_customer_message,
        )


def validate_retrieval_response(payload: object) -> dict[str, object]:
    if not isinstance(payload, dict):
        raise HTTPException(status_code=502, detail="RAG retrieval response must be a JSON object.")

    raw_results = payload.get("results")
    raw_hits = payload.get("hits")
    results = raw_results if isinstance(raw_results, list) else None
    hits = raw_hits if isinstance(raw_hits, list) else None

    selected_items = results if results is not None else hits
    if selected_items is None:
        raise HTTPException(status_code=502, detail="RAG retrieval response is missing hits or results.")

    if results is not None and hits is not None and len(results) != len(hits):
        logger.info(
            "search_api_response_normalized source=results results_count=%s hits_count=%s",
            len(results),
            len(hits),
        )

    standardized_hits = [
        normalize_external_search_hit(item, original_rank=index + 1)
        for index, item in enumerate(selected_items)
        if isinstance(item, dict)
    ]
    standardized_hits = sort_hits_for_output(standardized_hits)
    returned_count = len(standardized_hits)
    return {
        "status": str(payload.get("status") or "retrieved"),
        "query": str(payload.get("query") or ""),
        "top_k": returned_count or int(payload.get("final_k") or payload.get("top_k") or 0),
        "hit_count": returned_count or int(payload.get("hit_count") or 0),
        "retrieved_chunks": build_standardized_retrieved_chunks(standardized_hits),
        "collection_name": str(payload.get("collection_name") or ""),
        "embedding_provider": payload.get("embedding_provider"),
        "embedding_model": payload.get("embedding_model"),
        "hits": standardized_hits,
    }


def normalize_external_search_hit(hit: dict[str, object], *, original_rank: int | None = None) -> dict[str, object]:
    metadata = hit.get("metadata") if isinstance(hit.get("metadata"), dict) else {}
    scores = hit.get("scores") if isinstance(hit.get("scores"), dict) else {}

    normalized_hit: dict[str, object] = dict(hit)
    normalized_hit["id"] = str(hit.get("id") or hit.get("chunk_id") or "").strip()
    contents = get_hit_contents(hit)
    document_name = str(
        hit.get("source")
        or hit.get("document_name")
        or metadata.get("source_file")
        or hit.get("document_id")
        or UNKNOWN_DOCUMENT_NAME
    ).strip() or UNKNOWN_DOCUMENT_NAME
    header_path = str(
        hit.get("header_path")
        or hit.get("section_header")
        or hit.get("section")
        or metadata.get("header_path")
        or UNKNOWN_HEADER_PATH
    ).strip() or UNKNOWN_HEADER_PATH
    normalized_hit["text"] = contents
    normalized_hit["contents"] = contents
    normalized_hit["source"] = document_name
    normalized_hit["stored_name"] = str(hit.get("stored_name") or hit.get("document_id") or "").strip()
    normalized_hit["original_name"] = str(hit.get("original_name") or hit.get("document_name") or document_name).strip()
    normalized_hit["document_name"] = document_name
    normalized_hit["section_header"] = header_path
    normalized_hit["header_path"] = header_path
    normalized_hit["scores"] = dict(scores) if scores else {}
    if original_rank is not None:
        normalized_hit["original_rank"] = original_rank

    if isinstance(scores.get("rrf_score"), (int, float)):
        normalized_hit["rrf_score"] = scores["rrf_score"]

    if isinstance(hit.get("score"), (int, float)):
        normalized_hit["score"] = hit["score"]
    elif isinstance(scores.get("score"), (int, float)):
        normalized_hit["score"] = scores["score"]

    if isinstance(scores.get("rerank_score"), (int, float)):
        normalized_hit["rerank_score"] = scores["rerank_score"]
    elif isinstance(hit.get("score"), (int, float)):
        normalized_hit["rerank_score"] = hit["score"]
    elif isinstance(scores.get("vector_score"), (int, float)):
        normalized_hit["rerank_score"] = scores["vector_score"]

    if isinstance(metadata.get("chunk_index"), int):
        normalized_hit["chunk_index"] = metadata["chunk_index"]
    if isinstance(metadata.get("page_number"), int):
        normalized_hit["page_number"] = metadata["page_number"]

    return normalized_hit


def build_external_search_payload(
    endpoint: str,
    *,
    question: str,
    query: str,
    top_k: int,
    final_k: int,
    stored_name: str | None,
    rewrite_result: RewriteResult | None = None,
) -> dict[str, object]:
    # Keep document type hints in the rewrite trace only. Do not constrain
    # Search API recall with filters.document_type.
    filters: dict[str, object] = {}

    if endpoint.rstrip("/").endswith("/api/search"):
        # `docs/retrieval_api_design.md` defines request.top_k as the
        # intermediate candidate count after RRF and request.final_k as the
        # final returned result count. Keep that contract in the outbound
        # request and leave the API's internal candidate expansion
        # (`max(20, 2 * request.top_k)`) to the Search API implementation.
        requested_top_k = max(1, top_k)
        requested_final_k = max(1, final_k)
        external_top_k = max(requested_top_k, requested_final_k)
        payload: dict[str, object] = {
            "query": query,
            "top_k": external_top_k,
            "final_k": min(requested_final_k, external_top_k),
            "use_rerank": False,
            "include_source_metadata": True,
            "include_scores": True,
            "keyword_vector_weight": 0.3,
            "return_format": "json",
        }
        if filters:
            payload["filters"] = filters
        return payload

    payload = {
        "question": question,
        "query": query,
        "rewritten_query": query,
        "top_k": top_k,
        "stored_name": stored_name,
    }
    if filters:
        payload["filters"] = filters
    return payload


def call_rag_retrieval_endpoint(
    endpoint: str,
    *,
    question: str,
    query: str,
    top_k: int,
    final_k: int,
    stored_name: str | None,
    rewrite_result: RewriteResult | None = None,
) -> dict[str, object]:
    normalized_endpoint = endpoint.strip()
    if not normalized_endpoint:
        raise HTTPException(status_code=400, detail="RAG endpoint is required when calling an external retrieval API.")
    if not normalized_endpoint.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="RAG endpoint must start with http:// or https://.")

    payload = build_external_search_payload(
        normalized_endpoint,
        question=question,
        query=query,
        top_k=top_k,
        final_k=final_k,
        stored_name=stored_name,
        rewrite_result=rewrite_result,
    )
    logger.info(
        "search_api_call %s",
        json.dumps(
            {
                "endpoint": normalized_endpoint,
                "timeout_sec": DEFAULT_EXTERNAL_SEARCH_TIMEOUT_SEC,
                "payload": payload,
            },
            ensure_ascii=False,
            sort_keys=True,
        ),
    )
    request = urllib_request.Request(
        normalized_endpoint,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json"},
    )

    try:
        with urllib_request.urlopen(request, timeout=DEFAULT_EXTERNAL_SEARCH_TIMEOUT_SEC) as response:
            raw_body = response.read()
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"RAG retrieval request failed: {detail}") from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"RAG retrieval request failed: {exc.reason}") from exc
    except (TimeoutError, socket.timeout) as exc:
        raise HTTPException(
            status_code=504,
            detail=f"RAG retrieval request timed out after {DEFAULT_EXTERNAL_SEARCH_TIMEOUT_SEC} seconds.",
        ) from exc

    try:
        response_payload = json.loads(raw_body.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="RAG retrieval response was not valid JSON.") from exc

    return validate_retrieval_response(response_payload)


def should_fallback_to_internal_retrieval(exc: HTTPException) -> bool:
    return exc.status_code in {502, 504}


def resolve_retrieval_for_chat(payload: ChatRequest, rewrite_result: RewriteResult) -> tuple[dict[str, object], str, str | None]:
    retrieval_query = rewrite_result.rewritten_query
    search_api_endpoint = (payload.search_api_endpoint or DEFAULT_EXTERNAL_SEARCH_API_ENDPOINT).strip()
    resolved_top_k, resolved_final_k = resolve_chat_search_limits(payload)
    fallback_detail: str | None = None
    if search_api_endpoint:
        try:
            retrieval = call_rag_retrieval_endpoint(
                search_api_endpoint,
                question=payload.query.strip(),
                query=retrieval_query,
                top_k=resolved_top_k,
                final_k=resolved_final_k,
                stored_name=payload.stored_name,
                rewrite_result=rewrite_result,
            )
            return retrieval, search_api_endpoint, None
        except HTTPException as exc:
            if not should_fallback_to_internal_retrieval(exc):
                raise
            fallback_detail = (
                f"External search API failed ({exc.detail}). "
                "Fell back to internal retrieval."
            )
            logger.warning(
                "chat_search_external_fallback endpoint=%s query=%s detail=%s",
                search_api_endpoint,
                retrieval_query,
                exc.detail,
            )

    retrieval = retrieve_chunks_for_queries(
        queries=build_query_candidates_for_chat(payload, rewrite_result),
        top_k=resolved_top_k,
        stored_name=payload.stored_name,
        rerank_query=retrieval_query,
    )
    return retrieval, "internal:/retrieve", fallback_detail


def execute_search_for_chat(payload: ChatRequest, rewrite_result: RewriteResult) -> dict[str, object]:
    resolved_top_k, resolved_final_k = resolve_chat_search_limits(payload)
    logger.info(
        "chat_search_limits incoming_top_k=%s incoming_final_k=%s resolved_top_k=%s resolved_final_k=%s",
        payload.top_k,
        payload.final_k,
        resolved_top_k,
        resolved_final_k,
    )
    retrieval, rag_endpoint, detail = resolve_retrieval_for_chat(payload, rewrite_result)
    hits = retrieval.get("hits")
    if not isinstance(hits, list):
        hits = []
    hits = sort_hits_for_output([hit for hit in hits if isinstance(hit, dict)])

    retrieved_chunks = retrieval.get("retrieved_chunks")
    if not isinstance(retrieved_chunks, list):
        retrieved_chunks = build_standardized_retrieved_chunks(hits)
    else:
        retrieved_chunks = build_standardized_retrieved_chunks(hits)

    return {
        "query": str(retrieval.get("query") or rewrite_result.rewritten_query),
        "executed_queries": retrieval.get("executed_queries") if isinstance(retrieval.get("executed_queries"), list) else [],
        "top_k": int(retrieval.get("top_k") or resolved_top_k or resolved_final_k),
        "hit_count": int(retrieval.get("hit_count") or len(hits)),
        "collection_name": retrieval.get("collection_name"),
        "embedding_provider": retrieval.get("embedding_provider"),
        "embedding_model": retrieval.get("embedding_model"),
        "rag_endpoint": rag_endpoint,
        "detail": detail,
        "hits": hits,
        "retrieved_chunks": retrieved_chunks,
    }


def evaluate_retrieved_chunks(retrieved_chunks: list[dict[str, object]]) -> SearchEvaluationResult:
    if not retrieved_chunks:
        return SearchEvaluationResult(
            need_more_context=True,
            reason_codes=["no_hits"],
            evaluation_reasons=["검색 결과가 없어 추가 문맥 확보가 필요합니다."],
        )

    top_chunk = retrieved_chunks[0]
    top_text = normalize_chat_text(str(top_chunk.get("text") or ""))
    top_document_id = str(top_chunk.get("document_id") or "").strip() or None
    top_chunk_id = str(top_chunk.get("chunk_id") or "").strip() or None
    same_document_hit_count = 0
    if top_document_id:
        same_document_hit_count = sum(
            1
            for chunk in retrieved_chunks
            if str(chunk.get("document_id") or "").strip() == top_document_id
        )

    reason_codes: list[str] = []
    evaluation_reasons: list[str] = []
    keyword_hits: list[str] = []

    if len(top_text) < SEARCH_EVAL_MIN_TOP_CHUNK_TEXT_LENGTH:
        reason_codes.append("top_chunk_short")
        evaluation_reasons.append("상위 chunk 길이가 짧아 앞뒤 문맥이 더 필요할 수 있습니다.")

    for keyword in SEARCH_EVAL_CONTEXT_KEYWORDS:
        if keyword in top_text:
            keyword_hits.append(keyword)

    if keyword_hits:
        reason_codes.append("top_chunk_contains_condition_keywords")
        evaluation_reasons.append(
            f"상위 chunk에 조건/예외 신호({', '.join(keyword_hits)})가 있어 추가 문맥이 필요할 수 있습니다."
        )

    if top_document_id and same_document_hit_count >= SEARCH_EVAL_MULTI_HIT_SAME_DOCUMENT_THRESHOLD:
        reason_codes.append("same_document_cluster")
        evaluation_reasons.append("상위 결과가 같은 문서에 몰려 있어 주변 문맥 조회가 유리할 수 있습니다.")

    return SearchEvaluationResult(
        need_more_context=bool(reason_codes),
        reason_codes=reason_codes,
        evaluation_reasons=evaluation_reasons,
        top_document_id=top_document_id,
        top_chunk_id=top_chunk_id,
        top_chunk_text_length=len(top_text),
        same_document_hit_count=same_document_hit_count,
    )


def build_answer_generation_messages(
    *,
    original_query: str,
    rewritten_query: str,
    hits: list[dict[str, object]],
) -> list[dict[str, str]]:
    system_prompt = (
        "You answer questions using only the provided document context.\n"
        "Do not use outside knowledge.\n"
        "If the context is insufficient, say so.\n"
        "If the original question is in Korean, answer in Korean.\n"
        "Use each context item's document_name, header_path, and content together when judging evidence.\n"
        "Use document titles and section paths as part of your grounding, not just the content body.\n"
        "When the context includes conditional requirements, keep common requirements and conditional requirements separate.\n"
        "Never present conditional requirements as unconditional facts.\n"
        "If multiple documents or sections conflict, explicitly describe the difference or separate common points from conditional differences.\n"
        "If information is conditional, explain it as conditional.\n"
        "Do not guess anything not stated in the metadata or content.\n"
        "Use this exact format:\n"
        "STATUS: grounded or insufficient\n"
        "ANSWER: <final answer>"
    )
    user_prompt = (
        f"Original Question:\n{original_query}\n\n"
        f"Interpreted Retrieval Query:\n{rewritten_query}\n\n"
        f"Context:\n{format_chat_context(hits)}\n\n"
        "Answer generation spec (from docs/answer-generation-spec.md):\n"
        f"{load_answer_generation_spec() or 'answer generation spec is unavailable.'}\n\n"
        "Rules:\n"
        "- Answer only from the context.\n"
        "- If the context does not support a clear answer, return STATUS: insufficient.\n"
        "- Keep the answer concise and factual.\n"
        "- Do not mention information not present in the context.\n"
        "- Review document_name, header_path, and content for every context item before answering.\n"
        "- Use document titles and section locations as explicit grounding signals when comparing evidence.\n"
        "- If the document lists items that apply only in certain cases, separate them into '공통 서류' and '조건부 추가 서류'.\n"
        "- If the user's question is broad and the required documents vary by condition, explicitly say that additional documents depend on case.\n"
        "- If different documents or sections disagree, name the difference and distinguish common points from conditional or document-specific points.\n"
        "- Do not merge multiple scenario-specific document lists into one unconditional checklist.\n"
        "- Do not infer missing facts from partial metadata; if metadata or content is missing, use only what is present."
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def select_hits_for_answer_prompt(
    payload: ChatRequest,
    rewrite_result: RewriteResult,
    hits: list[dict[str, object]],
) -> list[dict[str, object]]:
    # Exclude only clear product mismatches based on the user query/rewrite and each hit's document_name.
    selected_hits = filter_hits_for_answer_generation(
        original_query=payload.query.strip(),
        rewritten_query=rewrite_result.rewritten_query,
        hits=hits,
    )
    if len(selected_hits) != len(hits):
        logger.info(
            "answer_context_product_filter original_query=%s rewritten_query=%s original_hit_count=%s filtered_hit_count=%s",
            payload.query.strip(),
            rewrite_result.rewritten_query,
            len(hits),
            len(selected_hits),
        )
    return selected_hits


def format_sse_event(event: str, payload: dict[str, object]) -> str:
    return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"


def build_sse_response(events) -> StreamingResponse:
    return StreamingResponse(
        events,
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def execute_query_rewrite_phase(payload: ChatRequest) -> tuple[RewriteResult, int]:
    rewrite_started_at = time.perf_counter()
    rewrite_result = rewrite_chat_query(payload)
    query_rewrite_time_ms = round((time.perf_counter() - rewrite_started_at) * 1000)
    return rewrite_result, query_rewrite_time_ms


def execute_search_phase(
    payload: ChatRequest,
    rewrite_result: RewriteResult,
    retrieval_query_override: str | None = None,
) -> tuple[dict[str, object], int, str, str]:
    search_started_at = time.perf_counter()
    search_api_endpoint = (payload.search_api_endpoint or DEFAULT_EXTERNAL_SEARCH_API_ENDPOINT).strip()
    action = "search"
    try:
        if retrieval_query_override:
            temporary_rewrite_result = rewrite_result.model_copy(
                update={
                    "rewritten_query": retrieval_query_override,
                    "search_queries": [retrieval_query_override, *rewrite_result.search_queries],
                }
            )
            search_result = execute_search_for_chat(payload, temporary_rewrite_result)
        else:
            search_result = execute_search_for_chat(payload, rewrite_result)
    except HTTPException as exc:
        search_api_response_time_ms = round((time.perf_counter() - search_started_at) * 1000)
        raise SearchPhaseExecutionError(
            cause=exc,
            action=action,
            search_api_endpoint=search_api_endpoint,
            response_time_ms=search_api_response_time_ms,
        ) from exc

    search_api_response_time_ms = round((time.perf_counter() - search_started_at) * 1000)
    return search_result, search_api_response_time_ms, action, search_api_endpoint


def choose_retry_search_query(rewrite_result: RewriteResult, current_query: str) -> str | None:
    current_normalized = normalize_chat_text(current_query).strip()
    for candidate in rewrite_result.search_queries:
        normalized_candidate = normalize_chat_text(candidate).strip()
        if not normalized_candidate or normalized_candidate == current_normalized:
            continue
        return normalized_candidate
    return None


def maybe_retry_search_phase(
    payload: ChatRequest,
    rewrite_result: RewriteResult,
    *,
    search_result: dict[str, object],
    insufficient_context: bool,
    current_search_time_ms: int,
) -> tuple[dict[str, object], int, str | None]:
    search_api_endpoint = (payload.search_api_endpoint or DEFAULT_EXTERNAL_SEARCH_API_ENDPOINT).strip()
    if not search_api_endpoint:
        return search_result, current_search_time_ms, None
    if not insufficient_context:
        return search_result, current_search_time_ms, None

    retry_query = choose_retry_search_query(rewrite_result, str(search_result.get("query") or rewrite_result.rewritten_query))
    if not retry_query:
        return search_result, current_search_time_ms, None

    logger.info(
        "chat_search_retry_started original_query=%s retry_query=%s",
        rewrite_result.rewritten_query,
        retry_query,
    )
    retry_result, retry_time_ms, _, _ = execute_search_phase(
        payload,
        rewrite_result,
        retrieval_query_override=retry_query,
    )
    merged_time_ms = current_search_time_ms + retry_time_ms
    retry_detail = f"Initial answer was insufficient. Retried search with alternate query: {retry_query}."
    existing_detail = search_result.get("detail") if isinstance(search_result.get("detail"), str) else None
    retry_result["detail"] = " ".join(part for part in [existing_detail, retry_detail] if part)
    return retry_result, merged_time_ms, retry_query


def generate_grounded_answer_stream(payload: ChatRequest) -> StreamingResponse:
    original_query = payload.query.strip()
    rewrite_result, query_rewrite_time_ms = execute_query_rewrite_phase(payload)
    answer_model_id = get_answer_display_model(payload)

    try:
        search_result, search_api_response_time_ms, action, search_api_endpoint = execute_search_phase(
            payload,
            rewrite_result,
        )
    except SearchPhaseExecutionError as exc:
        action = exc.action
        search_api_endpoint = exc.search_api_endpoint
        search_api_response_time_ms = exc.response_time_ms
        error_reason = "검색 API 연결에 실패해 검색 결과를 가져오지 못했습니다."
        error_answer = "검색 API 연결에 실패했습니다."
        logger.warning(
            "chat_%s_failed query=%s rewrite_query=%s detail=%s",
            action,
            original_query,
            rewrite_result.rewritten_query,
            exc.cause.detail,
        )
        error_payload = {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": [],
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": rewrite_result.rewritten_query,
            "executed_search_queries": [],
            "top_k": payload.final_k,
            "hit_count": 0,
            "collection_name": None,
            "embedding_provider": None,
            "embedding_model": None,
            "rag_endpoint": search_api_endpoint or "internal:/retrieve",
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": True,
            "search_evaluation": {
                "need_more_context": True,
                "reason_codes": [f"{action}_request_failed"],
                "evaluation_reasons": [error_reason],
            },
            "answer": error_answer,
            "insufficient_context": True,
            "citations": [],
            "hits": [],
            "chat_model": None,
            "answer_model": answer_model_id,
            "detail": str(exc.cause.detail),
        }

        def error_events():
            yield format_sse_event("done", error_payload)

        return build_sse_response(error_events())

    hits = search_result["hits"]
    retrieved_chunks = search_result["retrieved_chunks"]
    search_result_detail = search_result.get("detail") if isinstance(search_result.get("detail"), str) else None
    search_evaluation = evaluate_retrieved_chunks(retrieved_chunks)

    if not hits:
        empty_payload = {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": retrieved_chunks,
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": search_result["query"],
            "executed_search_queries": search_result["executed_queries"],
            "top_k": search_result["top_k"],
            "hit_count": 0,
            "collection_name": search_result["collection_name"],
            "embedding_provider": search_result["embedding_provider"],
            "embedding_model": search_result["embedding_model"],
            "rag_endpoint": search_result["rag_endpoint"],
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": search_evaluation.need_more_context,
            "search_evaluation": serialize_search_evaluation(search_evaluation),
            "answer": "문서에서 충분한 근거를 찾지 못했습니다.",
            "insufficient_context": True,
            "citations": [],
            "hits": [],
            "chat_model": None,
            "answer_model": answer_model_id,
            "detail": search_result_detail,
        }

        def empty_events():
            yield format_sse_event("done", empty_payload)

        return build_sse_response(empty_events())

    answer_context_hits = select_hits_for_answer_prompt(payload, rewrite_result, hits)
    if not answer_context_hits:
        empty_answer_payload = {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": retrieved_chunks,
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": search_result["query"],
            "executed_search_queries": search_result["executed_queries"],
            "top_k": search_result["top_k"],
            "hit_count": search_result["hit_count"],
            "collection_name": search_result["collection_name"],
            "embedding_provider": search_result["embedding_provider"],
            "embedding_model": search_result["embedding_model"],
            "rag_endpoint": search_result["rag_endpoint"],
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": True,
            "search_evaluation": serialize_search_evaluation(search_evaluation),
            "answer": "질문과 일치하는 상품 문맥을 찾지 못했습니다.",
            "insufficient_context": True,
            "citations": build_chat_citations(hits),
            "hits": hits,
            "chat_model": None,
            "answer_model": answer_model_id,
            "detail": search_result_detail,
        }

        def filtered_empty_events():
            yield format_sse_event("done", empty_answer_payload)

        return build_sse_response(filtered_empty_events())

    messages = build_answer_generation_messages(
        original_query=original_query,
        rewritten_query=rewrite_result.rewritten_query,
        hits=answer_context_hits,
    )

    def stream_events():
        base_payload: dict[str, object] = {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": retrieved_chunks,
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": search_result["query"],
            "executed_search_queries": search_result["executed_queries"],
            "top_k": search_result["top_k"],
            "hit_count": search_result["hit_count"],
            "collection_name": search_result["collection_name"],
            "embedding_provider": search_result["embedding_provider"],
            "embedding_model": search_result["embedding_model"],
            "rag_endpoint": search_result["rag_endpoint"],
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": search_evaluation.need_more_context,
            "search_evaluation": serialize_search_evaluation(search_evaluation),
            "answer": "",
            "insufficient_context": False,
            "citations": build_chat_citations(hits),
            "hits": hits,
            "chat_model": answer_model_id,
            "answer_model": answer_model_id,
            "detail": search_result_detail,
        }

        rewrite_text = rewrite_result.rewritten_query.strip()
        if rewrite_text:
            pending_rewrite_delta = ""
            for char in rewrite_text:
                pending_rewrite_delta += char
                if len(pending_rewrite_delta) >= STREAM_REWRITE_DELTA_MIN_CHARS or char in {" ", "\n", "\t", "?", ".", "!"}:
                    yield format_sse_event("rewrite_delta", {"content": pending_rewrite_delta})
                    pending_rewrite_delta = ""
            if pending_rewrite_delta:
                yield format_sse_event("rewrite_delta", {"content": pending_rewrite_delta})
            yield format_sse_event("rewrite_done", {"rewritten_query": rewrite_text})

        yield format_sse_event("meta", base_payload)

        streamed_response_text = ""
        streamed_answer_preview = ""
        pending_delta_text = ""
        last_delta_flush_at = time.perf_counter()
        streamed_model_id = answer_model_id
        try:
            answer_stream, streamed_model_id = build_answer_chat_completion_stream(payload, messages)
            for chunk_text in answer_stream:
                if not chunk_text:
                    continue
                streamed_response_text += chunk_text
                next_answer_preview = extract_streaming_answer_preview(streamed_response_text)
                if not next_answer_preview:
                    continue
                if not next_answer_preview.startswith(streamed_answer_preview):
                    # Stream deltas are append-only, so skip incompatible rewrites and rely on final `done`.
                    continue
                delta_text = next_answer_preview[len(streamed_answer_preview) :]
                if not delta_text:
                    continue
                streamed_answer_preview = next_answer_preview
                pending_delta_text += delta_text
                now = time.perf_counter()
                should_flush = (
                    len(pending_delta_text) >= STREAM_DELTA_MIN_CHARS
                    or "\n" in pending_delta_text
                    or (now - last_delta_flush_at) >= STREAM_DELTA_FLUSH_INTERVAL_SEC
                )
                if should_flush:
                    yield format_sse_event("delta", {"content": pending_delta_text})
                    pending_delta_text = ""
                    last_delta_flush_at = now
            if pending_delta_text:
                yield format_sse_event("delta", {"content": pending_delta_text})
        except HTTPException as exc:
            logger.warning(
                "chat_answer_stream_failed query=%s rewrite_query=%s detail=%s",
                original_query,
                rewrite_result.rewritten_query,
                exc.detail,
            )
            yield format_sse_event("error", {"detail": str(exc.detail)})
            failed_payload = {
                **base_payload,
                "answer": "답변 생성 스트림에 실패했습니다.",
                "insufficient_context": True,
                "chat_model": streamed_model_id,
                "answer_model": streamed_model_id,
                "detail": str(exc.detail),
            }
            yield format_sse_event("done", failed_payload)
            return

        response_text = streamed_response_text
        insufficient_context, answer = parse_chat_completion(response_text)
        if insufficient_context:
            yield format_sse_event("answer_replace", {"content": CHAT_RETRY_ANSWER_TEXT})
            try:
                retry_search_result, retried_search_time_ms, retry_query = maybe_retry_search_phase(
                    payload,
                    rewrite_result,
                    search_result=search_result,
                    insufficient_context=True,
                    current_search_time_ms=search_api_response_time_ms,
                )
                retry_hits = retry_search_result["hits"]
                if retry_hits:
                    retry_retrieved_chunks = retry_search_result["retrieved_chunks"]
                    retry_search_evaluation = evaluate_retrieved_chunks(retry_retrieved_chunks)
                    retry_answer_context_hits = select_hits_for_answer_prompt(payload, rewrite_result, retry_hits)
                    if not retry_answer_context_hits:
                        completed_payload = {
                            **base_payload,
                            "answer": "질문과 일치하는 상품 문맥을 찾지 못했습니다.",
                            "insufficient_context": True,
                            "search_query": retry_search_result["query"],
                            "executed_search_queries": retry_search_result["executed_queries"],
                            "top_k": retry_search_result["top_k"],
                            "hit_count": retry_search_result["hit_count"],
                            "collection_name": retry_search_result["collection_name"],
                            "embedding_provider": retry_search_result["embedding_provider"],
                            "embedding_model": retry_search_result["embedding_model"],
                            "rag_endpoint": retry_search_result["rag_endpoint"],
                            "search_api_response_time_ms": retried_search_time_ms,
                            "need_more_context": True,
                            "search_evaluation": serialize_search_evaluation(retry_search_evaluation),
                            "retrieved_chunks": retry_retrieved_chunks,
                            "citations": build_chat_citations(retry_hits),
                            "hits": retry_hits,
                            "chat_model": answer_model_id,
                            "answer_model": answer_model_id,
                            "detail": retry_search_result.get("detail"),
                        }
                        yield format_sse_event("done", completed_payload)
                        return
                    retry_messages = build_answer_generation_messages(
                        original_query=original_query,
                        rewritten_query=str(retry_search_result["query"] or rewrite_result.rewritten_query),
                        hits=retry_answer_context_hits,
                    )
                    retry_response_text, retry_answer_model_id = build_answer_chat_completion(payload, retry_messages)
                    retry_insufficient_context, retry_answer = parse_chat_completion(retry_response_text)
                    completed_payload = {
                        **base_payload,
                        "answer": retry_answer,
                        "insufficient_context": retry_insufficient_context,
                        "search_query": retry_search_result["query"],
                        "executed_search_queries": retry_search_result["executed_queries"],
                        "top_k": retry_search_result["top_k"],
                        "hit_count": retry_search_result["hit_count"],
                        "collection_name": retry_search_result["collection_name"],
                        "embedding_provider": retry_search_result["embedding_provider"],
                        "embedding_model": retry_search_result["embedding_model"],
                        "rag_endpoint": retry_search_result["rag_endpoint"],
                        "search_api_response_time_ms": retried_search_time_ms,
                        "need_more_context": retry_search_evaluation.need_more_context,
                        "search_evaluation": serialize_search_evaluation(retry_search_evaluation),
                        "retrieved_chunks": retry_retrieved_chunks,
                        "citations": build_chat_citations(retry_hits),
                        "hits": retry_hits,
                        "chat_model": retry_answer_model_id,
                        "answer_model": retry_answer_model_id,
                        "detail": retry_search_result.get("detail"),
                    }
                    if retry_query:
                        completed_payload["answer"] = retry_answer
                    yield format_sse_event("done", completed_payload)
                    return
            except (HTTPException, SearchPhaseExecutionError) as exc:
                logger.warning(
                    "chat_search_retry_failed query=%s rewrite_query=%s detail=%s",
                    original_query,
                    rewrite_result.rewritten_query,
                    exc.detail if isinstance(exc, HTTPException) else exc.cause.detail,
                )
        completed_payload = {
            **base_payload,
            "answer": answer,
            "insufficient_context": insufficient_context,
            "chat_model": streamed_model_id,
            "answer_model": streamed_model_id,
        }
        yield format_sse_event("done", completed_payload)

    return build_sse_response(stream_events())


def generate_grounded_answer(payload: ChatRequest) -> dict[str, object]:
    original_query = payload.query.strip()
    rewrite_result, query_rewrite_time_ms = execute_query_rewrite_phase(payload)
    answer_model_id = get_answer_display_model(payload)
    try:
        search_result, search_api_response_time_ms, action, search_api_endpoint = execute_search_phase(
            payload,
            rewrite_result,
        )
    except SearchPhaseExecutionError as exc:
        action = exc.action
        search_api_endpoint = exc.search_api_endpoint
        search_api_response_time_ms = exc.response_time_ms
        error_reason = "검색 API 연결에 실패해 검색 결과를 가져오지 못했습니다."
        error_answer = "검색 API 연결에 실패했습니다."
        logger.warning(
            "chat_%s_failed query=%s rewrite_query=%s detail=%s",
            action,
            original_query,
            rewrite_result.rewritten_query,
            exc.cause.detail,
        )
        return {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": [],
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": rewrite_result.rewritten_query,
            "executed_search_queries": [],
            "top_k": payload.final_k,
            "hit_count": 0,
            "collection_name": None,
            "embedding_provider": None,
            "embedding_model": None,
            "rag_endpoint": search_api_endpoint or "internal:/retrieve",
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": True,
            "search_evaluation": {
                "need_more_context": True,
                "reason_codes": [f"{action}_request_failed"],
                "evaluation_reasons": [error_reason],
            },
            "answer": error_answer,
            "insufficient_context": True,
            "citations": [],
            "hits": [],
            "chat_model": None,
            "answer_model": answer_model_id,
            "detail": str(exc.cause.detail),
        }
    hits = search_result["hits"]
    retrieved_chunks = search_result["retrieved_chunks"]
    search_result_detail = search_result.get("detail") if isinstance(search_result.get("detail"), str) else None

    search_evaluation = evaluate_retrieved_chunks(retrieved_chunks)

    if not hits:
        if not hits:
            return {
                "status": "answered",
                "query": original_query,
                "interpreted_query": rewrite_result.rewritten_query,
                "rewritten_query": rewrite_result.rewritten_query,
                "search_queries": rewrite_result.search_queries,
                "retrieved_chunks": retrieved_chunks,
                "normalized_conversation": rewrite_result.normalized_conversation,
                "last_customer_message": rewrite_result.last_customer_message,
                "validation_reasons": rewrite_result.validation_reasons,
                "rewrite_source": rewrite_result.rewrite_source,
                "query_rewrite_model": rewrite_result.query_rewrite_model,
                "intent": rewrite_result.intent,
                "question_type": rewrite_result.question_type,
                "entities": rewrite_result.entities,
                "routing_hints": rewrite_result.routing_hints,
                "search_query": search_result["query"],
                "executed_search_queries": search_result["executed_queries"],
                "top_k": search_result["top_k"],
                "hit_count": 0,
                "collection_name": search_result["collection_name"],
                "embedding_provider": search_result["embedding_provider"],
                "embedding_model": search_result["embedding_model"],
                "rag_endpoint": search_result["rag_endpoint"],
                "search_api_endpoint": search_api_endpoint or None,
                "action": action,
                "query_rewrite_time_ms": query_rewrite_time_ms,
                "search_api_response_time_ms": search_api_response_time_ms,
                "need_more_context": search_evaluation.need_more_context,
                "search_evaluation": serialize_search_evaluation(search_evaluation),
                "answer": "문서에서 충분한 근거를 찾지 못했습니다.",
                "insufficient_context": True,
                "citations": [],
                "hits": [],
                "chat_model": None,
                "answer_model": answer_model_id,
                "detail": search_result_detail,
            }

    answer_context_hits = select_hits_for_answer_prompt(payload, rewrite_result, hits)
    if not answer_context_hits:
        return {
            "status": "answered",
            "query": original_query,
            "interpreted_query": rewrite_result.rewritten_query,
            "rewritten_query": rewrite_result.rewritten_query,
            "search_queries": rewrite_result.search_queries,
            "retrieved_chunks": retrieved_chunks,
            "normalized_conversation": rewrite_result.normalized_conversation,
            "last_customer_message": rewrite_result.last_customer_message,
            "validation_reasons": rewrite_result.validation_reasons,
            "rewrite_source": rewrite_result.rewrite_source,
            "query_rewrite_model": rewrite_result.query_rewrite_model,
            "intent": rewrite_result.intent,
            "question_type": rewrite_result.question_type,
            "entities": rewrite_result.entities,
            "routing_hints": rewrite_result.routing_hints,
            "search_query": search_result["query"],
            "executed_search_queries": search_result["executed_queries"],
            "top_k": search_result["top_k"],
            "hit_count": search_result["hit_count"],
            "collection_name": search_result["collection_name"],
            "embedding_provider": search_result["embedding_provider"],
            "embedding_model": search_result["embedding_model"],
            "rag_endpoint": search_result["rag_endpoint"],
            "search_api_endpoint": search_api_endpoint or None,
            "action": action,
            "query_rewrite_time_ms": query_rewrite_time_ms,
            "search_api_response_time_ms": search_api_response_time_ms,
            "need_more_context": True,
            "search_evaluation": serialize_search_evaluation(search_evaluation),
            "answer": "질문과 일치하는 상품 문맥을 찾지 못했습니다.",
            "insufficient_context": True,
            "citations": build_chat_citations(hits),
            "hits": hits,
            "chat_model": None,
            "answer_model": answer_model_id,
            "detail": search_result_detail,
        }

    messages = build_answer_generation_messages(
        original_query=original_query,
        rewritten_query=rewrite_result.rewritten_query,
        hits=answer_context_hits,
    )
    response_text, answer_model_id = build_answer_chat_completion(
        payload,
        messages,
    )
    insufficient_context, answer = parse_chat_completion(response_text)
    if insufficient_context:
        try:
            retry_search_result, search_api_response_time_ms, retry_query = maybe_retry_search_phase(
                payload,
                rewrite_result,
                search_result=search_result,
                insufficient_context=True,
                current_search_time_ms=search_api_response_time_ms,
            )
            retry_hits = retry_search_result["hits"]
            if retry_hits:
                retry_retrieved_chunks = retry_search_result["retrieved_chunks"]
                retry_search_evaluation = evaluate_retrieved_chunks(retry_retrieved_chunks)
                retry_answer_context_hits = select_hits_for_answer_prompt(payload, rewrite_result, retry_hits)
                if not retry_answer_context_hits:
                    search_result = retry_search_result
                    hits = retry_hits
                    retrieved_chunks = retry_retrieved_chunks
                    search_result_detail = retry_search_result.get("detail") if isinstance(retry_search_result.get("detail"), str) else search_result_detail
                    search_evaluation = retry_search_evaluation
                    insufficient_context = True
                    answer = "질문과 일치하는 상품 문맥을 찾지 못했습니다."
                else:
                    retry_messages = build_answer_generation_messages(
                        original_query=original_query,
                        rewritten_query=str(retry_search_result["query"] or rewrite_result.rewritten_query),
                        hits=retry_answer_context_hits,
                    )
                    retry_response_text, answer_model_id = build_answer_chat_completion(
                        payload,
                        retry_messages,
                    )
                    insufficient_context, answer = parse_chat_completion(retry_response_text)
                    search_result = retry_search_result
                    hits = retry_hits
                    retrieved_chunks = retry_retrieved_chunks
                    search_result_detail = retry_search_result.get("detail") if isinstance(retry_search_result.get("detail"), str) else search_result_detail
                    search_evaluation = retry_search_evaluation
        except (HTTPException, SearchPhaseExecutionError) as exc:
            logger.warning(
                "chat_search_retry_failed query=%s rewrite_query=%s detail=%s",
                original_query,
                rewrite_result.rewritten_query,
                exc.detail if isinstance(exc, HTTPException) else exc.cause.detail,
            )

    return {
        "status": "answered",
        "query": original_query,
        "interpreted_query": rewrite_result.rewritten_query,
        "rewritten_query": rewrite_result.rewritten_query,
        "search_queries": rewrite_result.search_queries,
        "retrieved_chunks": retrieved_chunks,
        "normalized_conversation": rewrite_result.normalized_conversation,
        "last_customer_message": rewrite_result.last_customer_message,
        "validation_reasons": rewrite_result.validation_reasons,
        "rewrite_source": rewrite_result.rewrite_source,
        "query_rewrite_model": rewrite_result.query_rewrite_model,
        "intent": rewrite_result.intent,
        "question_type": rewrite_result.question_type,
        "entities": rewrite_result.entities,
        "routing_hints": rewrite_result.routing_hints,
        "search_query": search_result["query"],
        "executed_search_queries": search_result["executed_queries"],
        "top_k": search_result["top_k"],
        "hit_count": search_result["hit_count"],
        "collection_name": search_result["collection_name"],
        "embedding_provider": search_result["embedding_provider"],
        "embedding_model": search_result["embedding_model"],
        "rag_endpoint": search_result["rag_endpoint"],
        "search_api_endpoint": search_api_endpoint or None,
        "action": action,
        "query_rewrite_time_ms": query_rewrite_time_ms,
        "search_api_response_time_ms": search_api_response_time_ms,
        "need_more_context": search_evaluation.need_more_context,
        "search_evaluation": serialize_search_evaluation(search_evaluation),
        "answer": answer,
        "insufficient_context": insufficient_context,
        "citations": build_chat_citations(hits),
        "hits": hits,
        "chat_model": answer_model_id,
        "answer_model": answer_model_id,
        "detail": search_result_detail,
    }


def delete_indexed_chunks(stored_name: str) -> int:
    collection = get_chroma_collection()
    result = collection.get(where={"stored_name": stored_name}, include=[])
    ids = result.get("ids") if isinstance(result, dict) else None

    if not isinstance(ids, list) or not ids:
        return 0

    collection.delete(ids=ids)
    return len(ids)


def clear_indexed_chunks() -> int:
    collection = get_chroma_collection()
    result = collection.get(include=[])
    ids = result.get("ids") if isinstance(result, dict) else None

    if not isinstance(ids, list) or not ids:
        return 0

    collection.delete(ids=ids)
    return len(ids)


def delete_uploaded_file_and_index(stored_name: str) -> dict[str, object]:
    path = get_uploaded_file_path(stored_name)
    metadata = build_file_metadata(path)
    removed_index_count = delete_indexed_chunks(path.name)
    markdown_output_path = get_markdown_output_path(path)
    removed_markdown_output = markdown_output_path.is_file()
    if removed_markdown_output:
        markdown_output_path.unlink()

    parse_summary_path = get_parse_summary_path(path)
    parse_summary_removed = parse_summary_path.is_file()
    if parse_summary_removed:
        parse_summary_path.unlink()

    chunk_summary_path = get_chunk_summary_path(path)
    chunk_summary_removed = chunk_summary_path.is_file()
    if chunk_summary_removed:
        chunk_summary_path.unlink()

    path.unlink()

    return {
        "status": "deleted",
        "stored_name": metadata["stored_name"],
        "original_name": metadata["original_name"],
        "removed_index_count": removed_index_count,
        "removed_parse_summary": parse_summary_removed,
        "removed_chunk_summary": chunk_summary_removed,
        "removed_markdown_output": removed_markdown_output,
    }


def list_indexed_files() -> list[dict[str, object]]:
    collection = get_chroma_collection()
    result = collection.get(include=["metadatas"])

    metadatas = result.get("metadatas") if isinstance(result, dict) else None
    if not isinstance(metadatas, list):
        return []

    indexed: dict[str, dict[str, object]] = {}
    for metadata in metadatas:
        if not isinstance(metadata, dict):
            continue

        stored_name = metadata.get("stored_name")
        if not isinstance(stored_name, str) or not stored_name:
            continue

        if stored_name in indexed:
            indexed[stored_name]["chunk_count"] = int(indexed[stored_name]["chunk_count"]) + 1
            continue

        indexed[stored_name] = {
            "stored_name": stored_name,
            "original_name": metadata.get("original_name") or stored_name,
            "file_type": metadata.get("file_type") or "",
            "chunk_count": 1,
        }

    return sorted(indexed.values(), key=lambda item: str(item["original_name"]))


def rebuild_all_indexes(
    *,
    primary_parser: str = PRIMARY_PARSER_DOCLING,
    fallback_parser: str = FALLBACK_PARSER_EXTENSION_DEFAULT,
    chunk_target_length: int | None = None,
    chunk_overlap_length: int | None = None,
) -> dict[str, object]:
    uploaded_files = list_uploaded_files()
    removed_count = clear_indexed_chunks()
    rebuilt: list[dict[str, object]] = []

    for uploaded_file in uploaded_files:
        path = get_uploaded_file_path(str(uploaded_file["stored_name"]))
        rebuilt.append(
            index_chunks(
                path,
                primary_parser=primary_parser,
                fallback_parser=fallback_parser,
                chunk_target_length=chunk_target_length,
                chunk_overlap_length=chunk_overlap_length,
            )
        )

    embedding_config = get_embedding_config()
    return {
        "status": "rebuilt",
        "file_count": len(rebuilt),
        "removed_count": removed_count,
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "files": [
            {
                "stored_name": item["stored_name"],
                "original_name": item["original_name"],
                "indexed_count": item["indexed_count"],
            }
            for item in rebuilt
        ],
    }


def list_pipeline_files() -> list[dict[str, object]]:
    uploaded_files = list_uploaded_files()
    parse_history = get_parse_history_from_logs()
    parsed_files = {
        str(item.get("stored_name")): item
        for item in list_parsed_files()
        if isinstance(item.get("stored_name"), str)
    }
    chunked_files = {
        str(item.get("stored_name")): item
        for item in list_chunked_files()
        if isinstance(item.get("stored_name"), str)
    }
    indexed_files = {
        str(item.get("stored_name")): item
        for item in list_indexed_files()
        if isinstance(item.get("stored_name"), str)
    }

    pipeline_files: list[dict[str, object]] = []
    for uploaded in uploaded_files:
        stored_name = str(uploaded["stored_name"])
        history = parse_history.get(stored_name, {})
        parsed = parsed_files.get(stored_name)
        chunked = chunked_files.get(stored_name)
        indexed = indexed_files.get(stored_name)

        pipeline_files.append(
            {
                **uploaded,
                "upload_status": "completed",
                "parse_status": (
                    "completed"
                    if chunked or indexed or (parsed and parsed.get("status") != "failed")
                    else "failed"
                    if parsed and parsed.get("status") == "failed"
                    else "pending"
                ),
                "chunk_status": "completed" if chunked else "pending",
                "index_status": "completed" if indexed else "pending",
                "parse_text_length": parsed.get("text_length") if parsed else None,
                "parse_parser_used": parsed.get("parser_used") if parsed and parsed.get("parser_used") else None,
                "parse_fallback_used": parsed.get("fallback_used") if parsed else None,
                "parse_error_detail": parsed.get("error_detail") if parsed else None,
                "parse_preview": parsed.get("preview") if parsed else None,
                "markdown_path": parsed.get("markdown_path") if parsed else None,
                "last_successful_parse_parser_used": (
                    parsed.get("parser_used")
                    if parsed and parsed.get("status") == "completed" and parsed.get("parser_used")
                    else history.get("last_successful_parser_used")
                ),
                "last_successful_parse_fallback_used": (
                    parsed.get("fallback_used")
                    if parsed and parsed.get("status") == "completed"
                    else history.get("last_successful_fallback_used")
                ),
                "last_failed_parse_parser_used": (
                    parsed.get("parser_used")
                    if parsed and parsed.get("status") == "failed" and parsed.get("parser_used")
                    else history.get("last_failed_parser_used")
                ),
                "last_failed_parse_fallback_used": (
                    parsed.get("fallback_used")
                    if parsed and parsed.get("status") == "failed"
                    else history.get("last_failed_fallback_used")
                ),
                "chunk_count": chunked.get("chunk_count") if chunked else None,
                "indexed_chunk_count": indexed.get("chunk_count") if indexed else None,
                "quality_checked_at": parsed.get("quality_checked_at") if parsed else None,
                "jaccard_similarity": parsed.get("jaccard_similarity") if parsed else None,
                "levenshtein_distance": parsed.get("levenshtein_distance") if parsed else None,
                "quality_warning": parsed.get("quality_warning") if parsed else None,
                "quality_warning_message": parsed.get("quality_warning_message") if parsed else None,
                "quality_warning_reasons": parsed.get("quality_warning_reasons") if parsed else None,
                "pdf_garbled_detected": parsed.get("pdf_garbled_detected") if parsed else None,
                "pdf_suspicious_char_ratio": parsed.get("pdf_suspicious_char_ratio") if parsed else None,
                "pdf_reference_suspicious_char_ratio": parsed.get("pdf_reference_suspicious_char_ratio") if parsed else None,
                "pdf_suspicious_char_ratio_delta": parsed.get("pdf_suspicious_char_ratio_delta") if parsed else None,
                "pdf_replacement_character_count": parsed.get("pdf_replacement_character_count") if parsed else None,
                "pdf_control_character_count": parsed.get("pdf_control_character_count") if parsed else None,
                "pdf_text_length_ratio": parsed.get("pdf_text_length_ratio") if parsed else None,
            }
        )

    return pipeline_files


@app.get("/health")
def health() -> dict[str, str]:
    embedding_config = get_embedding_config()
    return {
        "status": "ok",
        "embedding_provider": str(embedding_config["provider"]),
        "embedding_model": str(embedding_config["model"]),
        "collection_name": str(embedding_config["collection_name"]),
    }


@app.get("/upload")
def get_upload_status() -> dict[str, object]:
    files = list_uploaded_files()
    return {
        "page": "upload",
        "status": "ready",
        "message": "Upload API is available.",
        "file_count": len(files),
    }


@app.get("/upload/files")
def get_uploaded_files() -> dict[str, object]:
    files = list_uploaded_files()
    return {"files": files, "count": len(files)}


@app.get("/pipeline/files")
def get_pipeline_files() -> dict[str, object]:
    files = list_pipeline_files()
    return {"files": files, "count": len(files)}


@app.delete("/upload/files")
def clear_uploaded_files() -> dict[str, object]:
    ensure_upload_dir()
    removed_count = 0

    for path in list(UPLOAD_DIR.iterdir()):
        if not path.is_file():
            continue
        delete_uploaded_file_and_index(path.name)
        removed_count += 1

    return {"status": "cleared", "removed_count": removed_count}


@app.delete("/upload/files/{stored_name}")
def delete_uploaded_file(stored_name: str) -> dict[str, object]:
    return delete_uploaded_file_and_index(stored_name)


@app.get("/upload/default-files")
def get_default_files() -> dict[str, object]:
    files = list_default_files()
    return {"files": files, "count": len(files), "directory": str(DEFAULT_FILE_DIR)}


@app.post("/upload")
async def upload_document(file: UploadFile = File(...)) -> dict[str, object]:
    filename = file.filename or ""
    extension = get_extension(filename)
    logger.info("upload_started filename=%s extension=%s", filename, extension)

    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=get_supported_file_types_message())

    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    normalized_name = Path(filename).name
    validate_uploaded_file_type(normalized_name, content)
    ensure_no_completed_duplicate_upload(normalized_name)
    saved = save_content_to_uploads(normalized_name, content, file.content_type)
    logger.info(
        "upload_completed stored_name=%s original_name=%s size_bytes=%s",
        saved["stored_name"],
        saved["original_name"],
        saved["size_bytes"],
    )
    return saved


@app.post("/upload/default-file")
def upload_default_file(payload: DefaultFileUploadRequest) -> dict[str, object]:
    logger.info("default_upload_started filename=%s", payload.filename)
    ensure_default_file_dir()
    filename = Path(payload.filename).name
    source = DEFAULT_FILE_DIR / filename

    if get_extension(filename) not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=get_supported_file_types_message())

    if not source.is_file():
        raise HTTPException(status_code=404, detail="Default file not found.")

    content = source.read_bytes()
    if not content:
        raise HTTPException(status_code=400, detail="Default file is empty.")

    validate_uploaded_file_type(filename, content)
    ensure_no_completed_duplicate_upload(filename)
    saved = save_content_to_uploads(filename, content, None)
    logger.info(
        "default_upload_completed stored_name=%s original_name=%s size_bytes=%s",
        saved["stored_name"],
        saved["original_name"],
        saved["size_bytes"],
    )
    return saved


@app.get("/parse")
def get_parse_status() -> dict[str, object]:
    files = list_uploaded_files()
    return {
        "page": "parse",
        "status": "ready",
        "message": "Parsing API is available for uploaded PDF, DOC, DOCX, XLS, and XLSX files.",
        "file_count": len(files),
    }


@app.get("/parse/parsers")
def get_parse_parsers() -> dict[str, object]:
    return get_parser_catalog()


@app.post("/parse")
def parse_uploaded_file(payload: ParseRequest) -> dict[str, object]:
    path = get_uploaded_file_path(payload.stored_name)
    fallback_parser = normalize_fallback_parser(payload.primary_parser, payload.fallback_parser)
    logger.info(
        "parse_started stored_name=%s original_name=%s %s",
        payload.stored_name,
        path.name.split("__", 1)[1] if "__" in path.name else path.name,
        summarize_parser_selection(payload.primary_parser, fallback_parser),
    )
    try:
        result = build_parsing_result(
            path,
            primary_parser=payload.primary_parser,
            fallback_parser=fallback_parser,
        )
    except HTTPException as exc:
        write_parse_failure_summary(
            path,
            file_type=get_extension(path.name),
            parser_used=f"{payload.primary_parser} -> {fallback_parser}",
            fallback_used=fallback_parser != FALLBACK_PARSER_NONE,
            error_detail=str(exc.detail),
        )
        raise
    logger.info(
        "parse_completed stored_name=%s parser_used=%s fallback_used=%s text_length=%s",
        payload.stored_name,
        result["parser_used"],
        result["fallback_used"],
        result["text_length"],
    )
    return result


@app.get("/parse/{stored_name}")
def parse_uploaded_file_by_name(stored_name: str) -> dict[str, object]:
    path = get_uploaded_file_path(stored_name)
    return build_parsing_result(path)


@app.post("/parse/quality")
def parse_uploaded_file_quality(payload: ParseRequest) -> dict[str, object]:
    path = get_uploaded_file_path(payload.stored_name)
    fallback_parser = normalize_fallback_parser(payload.primary_parser, payload.fallback_parser)
    logger.info(
        "parse_quality_started stored_name=%s %s",
        payload.stored_name,
        summarize_parser_selection(payload.primary_parser, fallback_parser),
    )
    result = build_parsing_quality_result(
        path,
        primary_parser=payload.primary_parser,
        fallback_parser=fallback_parser,
    )
    logger.info(
        "parse_quality_completed stored_name=%s parser_used=%s fallback_used=%s jaccard_similarity=%s",
        payload.stored_name,
        result["parser_used"],
        result["fallback_used"],
        result["jaccard_similarity"],
    )
    return result


@app.get("/chunk")
def get_chunk_status() -> dict[str, object]:
    files = list_uploaded_files()
    return {
        "page": "chunk",
        "status": "ready",
        "message": "Chunking API is available for parsed uploaded files.",
        "file_count": len(files),
        "chunk_target_length": CHUNK_TARGET_LENGTH,
        "chunk_overlap_length": CHUNK_OVERLAP_LENGTH,
    }


@app.post("/chunk")
def chunk_uploaded_file(payload: ParseRequest) -> dict[str, object]:
    path = get_uploaded_file_path(payload.stored_name)
    fallback_parser = normalize_fallback_parser(payload.primary_parser, payload.fallback_parser)
    resolved_target_length, resolved_overlap_length = normalize_chunk_settings(
        payload.chunk_target_length,
        payload.chunk_overlap_length,
    )
    logger.info(
        "chunk_started stored_name=%s %s chunk_target_length=%s chunk_overlap_length=%s",
        payload.stored_name,
        summarize_parser_selection(payload.primary_parser, fallback_parser),
        resolved_target_length,
        resolved_overlap_length,
    )
    result = build_chunking_result(
        path,
        primary_parser=payload.primary_parser,
        fallback_parser=fallback_parser,
        chunk_target_length=resolved_target_length,
        chunk_overlap_length=resolved_overlap_length,
    )
    logger.info(
        "chunk_completed stored_name=%s parser_used=%s fallback_used=%s chunk_count=%s chunk_target_length=%s chunk_overlap_length=%s",
        payload.stored_name,
        result["parser_used"],
        result["fallback_used"],
        result["chunk_count"],
        result["chunk_target_length"],
        result["chunk_overlap_length"],
    )
    return result


@app.get("/chunk/{stored_name}")
def chunk_uploaded_file_by_name(stored_name: str) -> dict[str, object]:
    path = get_uploaded_file_path(stored_name)
    return build_chunking_result(path)


@app.get("/index")
def get_index_status() -> dict[str, object]:
    files = list_uploaded_files()
    embedding_config = get_embedding_config()
    return {
        "page": "index",
        "status": "ready",
        "message": "Indexing API is available for chunked uploaded files.",
        "file_count": len(files),
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "embedding_dimension": embedding_config["embedding_dimension"],
    }


@app.get("/index/files")
def get_indexed_files() -> dict[str, object]:
    files = list_indexed_files()
    embedding_config = get_embedding_config()
    return {
        "files": files,
        "count": len(files),
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
    }


@app.delete("/index/files")
def delete_indexed_files() -> dict[str, object]:
    removed_count = clear_indexed_chunks()
    embedding_config = get_embedding_config()
    return {
        "status": "cleared",
        "removed_count": removed_count,
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
    }


@app.post("/index")
def index_uploaded_file(payload: ParseRequest) -> dict[str, object]:
    path = get_uploaded_file_path(payload.stored_name)
    fallback_parser = normalize_fallback_parser(payload.primary_parser, payload.fallback_parser)
    resolved_target_length, resolved_overlap_length = normalize_chunk_settings(
        payload.chunk_target_length,
        payload.chunk_overlap_length,
    )
    logger.info(
        "index_started stored_name=%s %s chunk_target_length=%s chunk_overlap_length=%s",
        payload.stored_name,
        summarize_parser_selection(payload.primary_parser, fallback_parser),
        resolved_target_length,
        resolved_overlap_length,
    )
    result = index_chunks(
        path,
        primary_parser=payload.primary_parser,
        fallback_parser=fallback_parser,
        chunk_target_length=resolved_target_length,
        chunk_overlap_length=resolved_overlap_length,
    )
    logger.info(
        "index_completed stored_name=%s parser_used=%s fallback_used=%s indexed_count=%s chunk_target_length=%s chunk_overlap_length=%s",
        payload.stored_name,
        result["parser_used"],
        result["fallback_used"],
        result["indexed_count"],
        result["chunk_target_length"],
        result["chunk_overlap_length"],
    )
    return result


@app.post("/index/rebuild")
def rebuild_indexed_files(payload: RebuildIndexRequest) -> dict[str, object]:
    fallback_parser = normalize_fallback_parser(payload.primary_parser, payload.fallback_parser)
    resolved_target_length, resolved_overlap_length = normalize_chunk_settings(
        payload.chunk_target_length,
        payload.chunk_overlap_length,
    )
    logger.info(
        "index_rebuild_started %s chunk_target_length=%s chunk_overlap_length=%s",
        summarize_parser_selection(payload.primary_parser, fallback_parser),
        resolved_target_length,
        resolved_overlap_length,
    )
    result = rebuild_all_indexes(
        primary_parser=payload.primary_parser,
        fallback_parser=fallback_parser,
        chunk_target_length=resolved_target_length,
        chunk_overlap_length=resolved_overlap_length,
    )
    logger.info(
        "index_rebuild_completed file_count=%s removed_count=%s collection_name=%s embedding_provider=%s embedding_model=%s",
        result["file_count"],
        result["removed_count"],
        result["collection_name"],
        result["embedding_provider"],
        result["embedding_model"],
    )
    return result


@app.get("/index/{stored_name}")
def index_uploaded_file_by_name(stored_name: str) -> dict[str, object]:
    path = get_uploaded_file_path(stored_name)
    return index_chunks(
        path,
        primary_parser=PRIMARY_PARSER_LEGACY_AUTO,
        fallback_parser=FALLBACK_PARSER_EXTENSION_DEFAULT,
    )


@app.get("/retrieve")
def get_retrieve_status() -> dict[str, object]:
    embedding_config = get_embedding_config()
    return {
        "page": "retrieve",
        "status": "ready",
        "message": "Retrieval API is available for indexed chunks.",
        "collection_name": embedding_config["collection_name"],
        "embedding_provider": embedding_config["provider"],
        "embedding_model": embedding_config["model"],
        "embedding_dimension": embedding_config["embedding_dimension"],
    }


@app.post("/retrieve")
def retrieve_indexed_chunks(payload: RetrieveRequest) -> dict[str, object]:
    return retrieve_chunks(payload)


@app.get("/chat")
def get_chat_status() -> dict[str, object]:
    chat_model = None
    try:
        chat_model = get_chat_model_id()
    except HTTPException:
        chat_model = None

    return {
        "page": "chat",
        "status": "ready",
        "message": "Chat API supports retrieval-based grounded answers.",
        "chat_model": chat_model,
    }


@app.post("/chat")
def answer_with_retrieval(payload: ChatRequest):
    if payload.stream:
        return generate_grounded_answer_stream(payload)
    return generate_grounded_answer(payload)


@app.get("/evaluation")
def get_evaluation_status() -> dict[str, object]:
    return {
        "page": "evaluation",
        "status": "not_implemented",
        "message": "Evaluation API skeleton is ready.",
    }
